from __future__ import annotations

import argparse
import re
import uuid
from datetime import datetime
from html.parser import HTMLParser
from io import BytesIO
from urllib.parse import urljoin, urlparse
from urllib.request import Request, build_opener

from openpyxl import Workbook, load_workbook

from excel_transform_1c.adapters.excel import ADO_INDICATOR_HEADERS, ADO_OPIU_HEADERS


class SelectOptionsParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.options: dict[str, list[tuple[str, str]]] = {}
        self._select = ""
        self._option_value: str | None = None
        self._option_text: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        values = dict(attrs)
        if tag == "select":
            self._select = values.get("name") or values.get("id") or ""
            self.options.setdefault(self._select, [])
        elif tag == "option" and self._select:
            self._option_value = values.get("value") or ""
            self._option_text = []

    def handle_data(self, data: str) -> None:
        if self._option_value is not None:
            self._option_text.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag == "option" and self._option_value is not None and self._select:
            self.options[self._select].append(
                (self._option_value, " ".join("".join(self._option_text).split()))
            )
            self._option_value = None
            self._option_text = []
        elif tag == "select":
            self._select = ""


def get(opener, url: str) -> tuple[str, bytes]:
    with opener.open(url, timeout=30) as response:
        return response.geturl(), response.read()


def multipart_body(
    fields: dict[str, str],
    files: dict[str, tuple[str, bytes, str]],
) -> tuple[bytes, str]:
    boundary = f"----fitera-owner-smoke-{uuid.uuid4().hex}"
    chunks: list[bytes] = []
    for name, value in fields.items():
        chunks.extend(
            [
                f"--{boundary}\r\n".encode(),
                f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode(),
                value.encode("utf-8"),
                b"\r\n",
            ]
        )
    for name, (filename, payload, content_type) in files.items():
        chunks.extend(
            [
                f"--{boundary}\r\n".encode(),
                (
                    f'Content-Disposition: form-data; name="{name}"; '
                    f'filename="{filename}"\r\n'
                ).encode(),
                f"Content-Type: {content_type}\r\n\r\n".encode(),
                payload,
                b"\r\n",
            ]
        )
    chunks.append(f"--{boundary}--\r\n".encode())
    return b"".join(chunks), f"multipart/form-data; boundary={boundary}"


def post_multipart(
    opener,
    url: str,
    fields: dict[str, str],
    files: dict[str, tuple[str, bytes, str]],
) -> tuple[str, bytes]:
    body, content_type = multipart_body(fields, files)
    request = Request(url, data=body, method="POST")
    request.add_header("Content-Type", content_type)
    with opener.open(request, timeout=60) as response:
        return response.geturl(), response.read()


def home_context(opener, base_url: str) -> tuple[str, str]:
    _, payload = get(opener, urljoin(base_url, "/"))
    html = payload.decode("utf-8")
    for marker in ("ООО «ФИТЭРА»", "ERP-статьи", "ЦФО Инталев", "Загрузить / дополнить"):
        assert marker in html, f"Home marker missing: {marker}"

    parser = SelectOptionsParser()
    parser.feed(html)
    organization_options = [
        (value, label)
        for value, label in parser.options.get("organization_node_id", [])
        if value
    ]
    organization = next(
        (
            value
            for value, _ in organization_options
            if value == "000000001"
        ),
        organization_options[0][0],
    )
    scenario = next(
        value
        for value, label in parser.options.get("scenario_id", [])
        if value and "ПЛАН 2026" in label and "2026" in label
    )
    return organization, scenario


def budget_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Синтетический бизнес-свод"
    sheet.cell(2, 1, "АЮ Административный Отдел")
    sheet.cell(2, 5, "Отдел")
    for month in range(1, 13):
        sheet.cell(2, 22 + month, datetime(2026, month, 1))
        sheet.cell(3, 22 + month, "план")
    indicators = (
        (10, "Оборот в кг"),
        (11, "Выручка за 1 кг"),
        (12, "Итого расходов на 1 кг"),
        (13, "Валовая прибыль на 1 кг"),
        (20, "Выручка ИТОГО"),
        (21, "Прочие доходы по основной деятельности"),
        (22, "Валовая прибыль"),
        (30, "Расходы по основной деятельности ИТОГО"),
        (31, "Административные расходы"),
        (32, "Коммерческие расходы"),
        (33, "Расходы на транспортную логистику"),
        (34, "Расходы на складскую логистику"),
        (40, "EBITDA"),
        (41, "Операционная прибыль"),
    )
    for row_number, indicator in indicators:
        sheet.cell(row_number, 5, "АЮ Административный Отдел")
        sheet.cell(row_number, 7, indicator)
        for month in range(1, 13):
            sheet.cell(row_number, 22 + month, row_number * month)
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    return payload.getvalue()


def organization_reference_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист_1"
    organization_name = 'ООО "Айс Юнион"'
    cfo_name = "АЮ Административный Отдел"
    sheet.cell(7, 1, "Организация")
    sheet.cell(8, 7, "Головная организация")
    sheet.cell(8, 32, "Верхний уровень иерархии")
    sheet.cell(8, 39, "Код")
    sheet.cell(9, 1, organization_name)
    sheet.cell(10, 1, cfo_name)
    sheet.cell(11, 1, "Административный департамент")
    sheet.cell(11, 7, cfo_name)
    sheet.cell(11, 32, organization_name)
    sheet.cell(11, 39, "000000173")
    sheet.cell(12, 1, organization_name)
    sheet.cell(13, 7, organization_name)
    sheet.cell(13, 32, "4 Владивосток")
    sheet.cell(13, 39, "000000001")
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    return payload.getvalue()


def upload_organization_reference(opener, base_url: str) -> None:
    final_url, _ = post_multipart(
        opener,
        urljoin(base_url, "/references"),
        {"kind": "organizations"},
        {
            "reference_file": (
                "ОрганизациииерархияЕРП.xlsx",
                organization_reference_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert urlparse(final_url).path == "/"


def classifier_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Полный путь статьи", "Статья", "Показатель", "Канал сбыта"])
    sheet.append(
        [
            "Административные → Связь → Интернет",
            "Интернет",
            "Услуги связи",
            "Основной канал",
        ]
    )
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    return payload.getvalue()


def upload_budget(opener, base_url: str, organization: str, scenario: str) -> tuple[str, str]:
    final_url, payload = post_multipart(
        opener,
        urljoin(base_url, "/uploads"),
        {
            "reporting_unit": "АЮ Административный Отдел",
            "organization_node_id": organization,
            "scenario_id": scenario,
            "year": "2026",
            "period_selector_present": "1",
            "all_year": "1",
            "workbook_password": "",
        },
        {
            "budget_file": (
                "owner-smoke-http.xlsx",
                budget_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    match = re.search(r"/runs/([^/?#]+)", urlparse(final_url).path)
    assert match, f"Budget upload did not reach a RUN: {final_url}"
    return match.group(1), payload.decode("utf-8")


def indicator_counts(html: str) -> dict[str, int]:
    labels = {
        "automatic": "Найдено автоматически:",
        "attention": "Требует внимания:",
        "not_found": "Не найдено:",
    }
    result: dict[str, int] = {}
    for key, label in labels.items():
        match = re.search(
            rf"<dt>\s*{re.escape(label)}\s*</dt>\s*<dd>\s*(\d+)\s*</dd>",
            html,
        )
        assert match, f"Indicator counter missing: {label}"
        result[key] = int(match.group(1))
    return result


def export_payload(opener, base_url: str, run_id: str) -> bytes:
    _, payload = get(opener, urljoin(base_url, f"/runs/{run_id}/export"))
    return payload


def sheet_snapshot(payload: bytes, sheet_name: str) -> tuple[tuple[object, ...], ...]:
    workbook = load_workbook(BytesIO(payload), data_only=True, read_only=True)
    try:
        sheet = workbook[sheet_name]
        return tuple(tuple(cell.value for cell in row) for row in sheet.iter_rows())
    finally:
        workbook.close()


def assert_export(payload: bytes) -> None:
    workbook = load_workbook(BytesIO(payload), data_only=True, read_only=True)
    try:
        assert workbook.sheetnames == ["OPIU Light", "ОПИУ", "Показатели"]
        assert tuple(cell.value for cell in workbook["ОПИУ"][1]) == ADO_OPIU_HEADERS
        assert tuple(cell.value for cell in workbook["Показатели"][1]) == ADO_INDICATOR_HEADERS
        assert workbook["ОПИУ"].max_column == len(ADO_OPIU_HEADERS)
        assert workbook["Показатели"].max_column == 14
        assert workbook["Показатели"].max_row == 14 * 12 + 1
        rows = list(
            workbook["Показатели"].iter_rows(min_row=2, values_only=True)
        )
        assert {row[10] for row in rows} == {
            "Доход",
            "Расход",
            "KPI",
        }
        assert "Оборот в кг" in {row[12] for row in rows}
    finally:
        workbook.close()


def initial_owner_smoke(opener, base_url: str) -> None:
    upload_organization_reference(opener, base_url)
    organization, scenario = home_context(opener, base_url)
    run_id, initial_html = upload_budget(opener, base_url, organization, scenario)
    assert 'data-testid="bdr-diagnostics"' in initial_html
    for marker in (
        "Прочитано строк БДР:",
        "Месячных ячеек прочитано:",
        "Числовых значений:",
        "Ошибок Excel:",
        "Строк требует внимания:",
        "Доходные показатели:",
        "Расходные показатели:",
        "KPI найдено:",
        "KPI с организацией:",
        "KPI с периодом:",
        "KPI со значением:",
        "KPI экспортировано:",
        "Количество показателей:",
        "Сопоставлено:",
        "Экспортировано:",
    ):
        assert marker in initial_html
    assert re.search(r"Прочитано строк БДР:</dt><dd>14</dd>", initial_html)
    assert re.search(r"Доходные показатели:</dt><dd>3</dd>", initial_html)
    assert re.search(r"Расходные показатели:</dt><dd>5</dd>", initial_html)
    assert re.search(r"KPI найдено:</dt><dd>6</dd>", initial_html)
    assert re.search(r"KPI с организацией:</dt><dd>6</dd>", initial_html)
    assert re.search(r"KPI с периодом:</dt><dd>6</dd>", initial_html)
    assert re.search(r"KPI со значением:</dt><dd>6</dd>", initial_html)
    assert re.search(r"KPI экспортировано:</dt><dd>6</dd>", initial_html)
    assert re.search(r"Сопоставлено:</dt><dd>14</dd>", initial_html)
    assert 'data-testid="indicator-classifier-summary"' not in initial_html
    assert "Проверка завершена" in initial_html
    lowered = initial_html.lower()
    assert "sql id" not in lowered
    assert "internal key" not in lowered
    assert_export(export_payload(opener, base_url, run_id))
    print("OWNER_SMOKE_HTTP_INITIAL_PASS")


def post_restart_owner_smoke(opener, base_url: str) -> None:
    organization, scenario = home_context(opener, base_url)
    run_id, html = upload_budget(opener, base_url, organization, scenario)
    assert 'data-testid="bdr-diagnostics"' in html
    assert re.search(r"Количество показателей:</dt><dd>14</dd>", html)
    assert_export(export_payload(opener, base_url, run_id))
    print("OWNER_SMOKE_HTTP_POST_RESTART_PASS")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--base-url", required=True)
    parser.add_argument("--phase", choices=("initial", "post-restart"), required=True)
    args = parser.parse_args()

    base_url = args.base_url.rstrip("/") + "/"
    opener = build_opener()
    if args.phase == "initial":
        initial_owner_smoke(opener, base_url)
    else:
        post_restart_owner_smoke(opener, base_url)


if __name__ == "__main__":
    main()
