from __future__ import annotations

import os
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

from excel_transform_1c.application.service import WorkflowService
from excel_transform_1c.core.opiu_rules.opiu_rule_models import OPIURule, opiu_rule_to_payload


def _budget_bytes() -> bytes:
    headers = [
        "ПОДРАЗДЕЛЕНИЕ (ЦФО 1)",
        "ТИП РАСХОДОВ",
        "ДЕПАРТАМЕНТ (ЦФО 2)",
        "Вид организации",
        "ОТДЕЛ",
        "НАЛОГООБЛОЖЕНИЕ",
        "ГРУППА РАСХОДОВ",
        "СТАТЬЯ",
        "Январь",
        "Февраль",
        "Март",
        "Апрель",
        "Май",
        "Июнь",
        "Июль",
        "Август",
        "Сентябрь",
        "Октябрь",
        "Ноябрь",
        "Декабрь",
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Тест"
    sheet.append(["synthetic package smoke"])
    sheet.append(headers)
    sheet.append(
        [
            "ТЕСТ",
            "Административные",
            "Департамент",
            "ТК",
            "ЦФО",
            "БЕЗ НДС",
            "Связь",
            "Интернет",
            100,
            *([0] * 11),
        ]
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def main() -> None:
    runtime = Path(os.environ["OPIU_RUNTIME_DIR"])
    service = WorkflowService(runtime)
    counts = service.reference_counts()
    # A preceding HTTP smoke may explicitly supplement packaged catalogs.
    # Baselines must remain present, while exact totals may grow.
    assert counts["erp_articles"] >= 271
    assert counts["organizations"] >= 357
    assert counts["scenarios"] >= 12
    assert counts["intalev_cfos"] >= 16

    scenario = next(
        item
        for item in service.store.list_scenarios()
        if item.name == "ПЛАН 2026" and item.year == 2026
    )
    node = service.organization_nodes()[0]
    context = service.build_context(
        "ТЕСТ",
        node.node_id,
        scenario.scenario_id,
        2026,
        [],
    )
    pending = service.prepare_upload("package-smoke.xlsx", _budget_bytes(), context)
    run = service.process_upload(
        pending.upload_id,
        pending.candidates[0].candidate_id,
    )
    assert len(run.records) == 12

    formula_rule = OPIURule(
        rule_id="package-opiu-rule",
        report_line="Расходы",
        report_indicator="Административные расходы сумма",
        disclosure_group="Административные",
        article="Интернет",
        article_code=run.records[0].erp_code,
        source="Synthetic package smoke",
        formula_condition="",
        required_analytics=(),
        region_required=False,
        network_required=False,
        cfo_required=False,
        organization_required=False,
        indicator_code="IND-ADMIN",
        sales_channel="",
    )
    service.store.save_opiu_rules([opiu_rule_to_payload(formula_rule)])
    service._apply_indicator_matches(run)
    assert service.opiu_rule_source_count() >= 1
    assert service.indicator_counts(run.run_id) == {
        "automatic": 1,
        "attention": 0,
        "not_found": 0,
    }

    exported = service.export_run(run.run_id)
    workbook = load_workbook(BytesIO(exported), read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == ["OPIU Light", "ОПИУ", "Показатели"]
        assert workbook["OPIU Light"].max_row == 13
        assert workbook["ОПИУ"].max_row == 13
        assert workbook["Показатели"].max_row == 13
        headers = {
            cell.value: cell.column
            for cell in workbook["Показатели"][1]
        }
        assert (
            workbook["Показатели"].cell(
                2, headers["Канал сбыта"]
            ).value
            is None
        )
        assert (
            workbook["Показатели"].cell(
                2, headers["Показатель"]
            ).value
            == "Административные расходы сумма"
        )
        assert (
            workbook["Показатели"].cell(
                2, headers["Значение"]
            ).value
            == 100
        )
    finally:
        workbook.close()

    restarted = WorkflowService(runtime)
    assert restarted.opiu_rule_source_count() >= 1
    assert len(restarted.opiu_rules()) >= 1
    print("OPIU_RULES_PACKAGE_SMOKE_PASS")


if __name__ == "__main__":
    main()
