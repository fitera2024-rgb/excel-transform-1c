from __future__ import annotations

import re
import tempfile
import zipfile
from contextlib import contextmanager
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from pathlib import Path
from posixpath import normpath
from typing import Iterator
from xml.etree import ElementTree

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.utils import get_column_letter

from excel_transform_1c.adapters.workbook_repair import prepare_workbook
from excel_transform_1c.core.detection import detect_candidate_ranges, read_source_rows
from excel_transform_1c.core.indicator_matching import (
    IndicatorExportRow,
    aggregate_indicator_rows,
)
from excel_transform_1c.core.models import CandidateRange, IndicatorType, PreviewRecord


# Business export sheet. Reference names and codes intentionally use separate
# columns; a display path with a parenthesized code must never leak here.
EXPORT_HEADERS = (
    "Единица отчёта",
    "Организация",
    "Код организации",
    "Сценарий",
    "Год",
    "Месяц",
    "Период",
    "Департамент",
    "Вид организации",
    "Отдел",
    "ЦФО",
    "Код ЦФО",
    "Тип показателя",
    "Показатель",
    "Тип расходов",
    "Группа расходов",
    "Исходное название статьи",
    "ERP-код статьи",
    "Официальное название статьи ERP",
    "Налогообложение",
    "Значение",
    "Статус",
    "Комментарий",
    "Номер исходной строки",
)

# ADO-oriented sheet requested by the owner. Missing reference codes remain
# empty until the corresponding catalogs are uploaded; business rows are not
# dropped because a code is not available yet.
ADO_OPIU_HEADERS = (
    "Организация",
    "Код организации",
    "Сценарий",
    "Год",
    "Месяц",
    "Период",
    "Департамент",
    "Отдел",
    "ЦФО",
    "Код ЦФО",
    "Тип показателя",
    "Показатель",
    "Тип расходов",
    "Код статьи",
    "Название статьи",
    "Инт Номенклатура",
    "Код номенклатуры",
    "Регион продаж",
    "Код региона продаж",
    "Значение",
)

ADO_INDICATOR_HEADERS = (
    "Организация",
    "Код организации",
    "Департамент",
    "Отдел",
    "ЦФО",
    "Код ЦФО",
    "Сценарий",
    "Год",
    "Месяц",
    "Период",
    "Тип показателя",
    "Канал сбыта",
    "Показатель",
    "Значение",
)


def _separate_organization_reference(value: str) -> tuple[str, str]:
    """Split the exact RunContext display contract into export name and code."""

    text = value.strip()
    match = re.fullmatch(r"(?P<path>.+) \((?P<code>[^()]*)\)", text)
    if match is None:
        path = text
        code = ""
    else:
        path = match.group("path").strip()
        code = match.group("code").strip()
    name = path.rsplit(" → ", 1)[-1].strip()
    return name, code


def _record_organization_reference(record: PreviewRecord) -> tuple[str, str]:
    """Prefer the exact root reference produced by the hierarchy resolver."""

    if record.organization_unit or record.organization_unit_code:
        return (
            record.organization_unit.strip(),
            record.organization_unit_code.strip(),
        )
    return _separate_organization_reference(record.organization)


def _indicator_organization_references(
    records: list[PreviewRecord],
) -> dict[str, tuple[str, str]]:
    """Resolve one exact root reference for every aggregated context value."""

    candidates: dict[str, set[tuple[str, str]]] = {}
    for record in records:
        if not (record.organization_unit or record.organization_unit_code):
            continue
        candidates.setdefault(record.organization, set()).add(
            _record_organization_reference(record)
        )

    resolved: dict[str, tuple[str, str]] = {}
    for organization, references in candidates.items():
        if len(references) != 1:
            raise ValueError(
                "Для одного контекста экспорта найдены разные exact ERP-организации"
            )
        resolved[organization] = next(iter(references))
    return resolved


@contextmanager
def load_cached_workbook(
    path: str | Path,
    *,
    read_only: bool = True,
) -> Iterator[object]:
    """Open cached values through a disposable OOXML compatibility copy when needed."""

    source_path = Path(path)
    # Keep the loader injectable for focused adapter tests that use a virtual
    # path. Real files are first staged into a separate validated/repaired
    # OOXML working copy; the original upload remains byte-for-byte unchanged.
    prepared_path = (
        prepare_workbook(source_path).working_path
        if source_path.exists()
        else source_path
    )
    workbook = None
    try:
        load_options = {"data_only": True, "read_only": read_only}
        if not read_only:
            load_options["keep_links"] = False
        workbook = load_workbook(prepared_path, **load_options)
        yield workbook
        return
    except KeyError as exc:
        if "sharedStrings.xml" not in str(exc):
            raise
    finally:
        if workbook is not None:
            workbook.close()

    with tempfile.TemporaryDirectory(prefix="excel_transform_1c_") as temp_dir:
        compatible = Path(temp_dir) / prepared_path.name
        _normalized_ooxml_copy(prepared_path, compatible)
        workbook = load_workbook(
            compatible,
            data_only=True,
            read_only=read_only,
            keep_links=False,
        )
        try:
            yield workbook
        finally:
            workbook.close()


def _normalized_ooxml_copy(source: Path, target: Path) -> None:
    with zipfile.ZipFile(source, "r") as archive, zipfile.ZipFile(
        target,
        "w",
        compression=zipfile.ZIP_DEFLATED,
    ) as output:
        for info in archive.infolist():
            name = info.filename
            data = archive.read(name)
            if name.lower() == "xl/sharedstrings.xml":
                name = "xl/sharedStrings.xml"
            if name.lower() == "[content_types].xml":
                text = data.decode("utf-8", errors="strict")
                text = re.sub(
                    r'PartName="/xl/SharedStrings\\.xml"',
                    'PartName="/xl/sharedStrings.xml"',
                    text,
                    flags=re.IGNORECASE,
                )
                data = text.encode("utf-8")
            output.writestr(name, data)


def detect_path(path: str | Path) -> list[CandidateRange]:
    with load_cached_workbook(path) as workbook:
        return detect_candidate_ranges(workbook)


def read_path(path: str | Path, candidate: CandidateRange, source_file: str):
    # Intalev hierarchy can be encoded in row outline levels.  Open only that
    # bounded workflow in normal mode because ReadOnlyWorksheet deliberately
    # omits row_dimensions; prepared-budget workbooks stay streaming/read-only.
    read_only = candidate.source_kind != "intalev_opiu"
    cached_formula_values = (
        _bdr_cached_formula_values(path, candidate)
        if candidate.source_kind == "bdr_full"
        else {}
    )
    with load_cached_workbook(path, read_only=read_only) as workbook:
        return read_source_rows(
            workbook,
            candidate,
            source_file,
            cached_formula_values,
        )


def _bdr_cached_formula_values(
    path: str | Path,
    candidate: CandidateRange,
) -> dict[str, object]:
    """Read only saved formula results for exact KPI month coordinates.

    ``openpyxl(data_only=True)`` remains the primary value reader. This small
    OOXML fallback covers formula variants where the cached ``<v>`` exists but
    is not surfaced by the library. Formula text is never returned.
    """

    source_path = Path(path)
    if not source_path.exists():
        return {}
    kpi_rows = {
        row_number
        for indicator_type, first_row, last_row in candidate.indicator_blocks
        if indicator_type == IndicatorType.KPI
        for row_number in range(first_row, last_row + 1)
    }
    target_cells = {
        f"{get_column_letter(candidate.columns[f'month_{month}'])}{row_number}"
        for row_number in kpi_rows
        for month in range(1, 13)
    }
    result: dict[str, object] = {}
    try:
        prepared_path = prepare_workbook(source_path).working_path
        with zipfile.ZipFile(prepared_path) as archive:
            result.update(
                _ooxml_cached_cell_values(
                    archive,
                    candidate.sheet,
                    target_cells,
                    formulas_only=True,
                )
            )

            if (
                candidate.bdr_value_sheet
                and candidate.bdr_value_columns
                and candidate.bdr_value_rows
            ):
                source_to_target: dict[str, str] = {}
                for target_row, source_row in candidate.bdr_value_rows:
                    for month in range(1, 13):
                        target_coordinate = (
                            f"{get_column_letter(candidate.columns[f'month_{month}'])}"
                            f"{target_row}"
                        )
                        source_coordinate = (
                            f"{get_column_letter(candidate.bdr_value_columns[f'month_{month}'])}"
                            f"{source_row}"
                        )
                        source_to_target[source_coordinate] = target_coordinate
                source_values = _ooxml_cached_cell_values(
                    archive,
                    candidate.bdr_value_sheet,
                    set(source_to_target),
                    formulas_only=False,
                    use_display_precision=True,
                )
                result.update(
                    {
                        source_to_target[source_coordinate]: value
                        for source_coordinate, value in source_values.items()
                    }
                )
    except (OSError, KeyError, zipfile.BadZipFile, ElementTree.ParseError):
        return {}
    return result


def _ooxml_cached_cell_values(
    archive: zipfile.ZipFile,
    sheet_name: str,
    coordinates: set[str],
    *,
    formulas_only: bool,
    use_display_precision: bool = False,
) -> dict[str, object]:
    """Read saved OOXML scalars from one relationship-resolved worksheet."""

    sheet_part = _worksheet_part(archive, sheet_name)
    if not sheet_part:
        return {}
    result: dict[str, object] = {}
    number_formats = (
        _number_formats_by_style(archive) if use_display_precision else {}
    )
    with archive.open(sheet_part) as sheet_xml:
        for _, cell in ElementTree.iterparse(sheet_xml, events=("end",)):
            element_name = _local_name(cell.tag)
            if element_name != "c":
                if element_name == "row":
                    cell.clear()
                continue
            coordinate = cell.attrib.get("r", "")
            if coordinate in coordinates:
                has_formula = any(
                    _local_name(child.tag) == "f" for child in cell
                )
                cached = next(
                    (
                        child
                        for child in cell
                        if _local_name(child.tag) == "v"
                    ),
                    None,
                )
                if (
                    cached is not None
                    and cached.text is not None
                    and (has_formula or not formulas_only)
                ):
                    value = _cached_formula_scalar(
                        cell.attrib.get("t", "n"),
                        cached.text,
                    )
                    if number_formats and isinstance(value, Decimal):
                        try:
                            style_id = int(cell.attrib.get("s", "0"))
                        except ValueError:
                            style_id = 0
                        value = _apply_display_precision(
                            value,
                            number_formats.get(style_id, ""),
                        )
                    result[coordinate] = value
            cell.clear()
    return result


def _number_formats_by_style(archive: zipfile.ZipFile) -> dict[int, str]:
    """Resolve worksheet style indexes to Excel number-format codes."""

    try:
        root = ElementTree.fromstring(archive.read("xl/styles.xml"))
    except (KeyError, ElementTree.ParseError):
        return {}
    custom_formats: dict[int, str] = {}
    cell_xfs = None
    for element in root.iter():
        name = _local_name(element.tag)
        if name == "numFmt":
            try:
                custom_formats[int(element.attrib["numFmtId"])] = element.attrib.get(
                    "formatCode",
                    "",
                )
            except (KeyError, ValueError):
                continue
        elif name == "cellXfs":
            cell_xfs = element
    if cell_xfs is None:
        return {}
    result: dict[int, str] = {}
    for style_id, style in enumerate(cell_xfs):
        try:
            number_format_id = int(style.attrib.get("numFmtId", "0"))
        except ValueError:
            continue
        result[style_id] = custom_formats.get(
            number_format_id,
            BUILTIN_FORMATS.get(number_format_id, ""),
        )
    return result


def _apply_display_precision(
    value: Decimal,
    number_format: str,
) -> Decimal:
    """Apply plain-number display precision without converting dates/percentages."""

    positive_section = number_format.split(";", 1)[0]
    simplified = re.sub(r'"[^"]*"|\\.', "", positive_section)
    if not simplified or "%" in simplified or re.search(r"[dmyhs]", simplified, re.I):
        return value
    placeholders = re.sub(r"\[[^]]*]", "", simplified)
    if not re.search(r"[0#]", placeholders):
        return value
    decimal_part = placeholders.split(".", 1)[1] if "." in placeholders else ""
    decimal_places = len(re.match(r"[0#]*", decimal_part).group(0))
    quantum = Decimal(1).scaleb(-decimal_places)
    return value.quantize(quantum, rounding=ROUND_HALF_UP)


def _worksheet_part(archive: zipfile.ZipFile, sheet_name: str) -> str:
    workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    relation_id = ""
    for element in workbook_root.iter():
        if _local_name(element.tag) != "sheet" or element.attrib.get("name") != sheet_name:
            continue
        relation_id = next(
            (
                value
                for name, value in element.attrib.items()
                if _local_name(name) == "id"
            ),
            "",
        )
        break
    if not relation_id:
        return ""

    relations_root = ElementTree.fromstring(
        archive.read("xl/_rels/workbook.xml.rels")
    )
    target = next(
        (
            element.attrib.get("Target", "")
            for element in relations_root.iter()
            if _local_name(element.tag) == "Relationship"
            and element.attrib.get("Id") == relation_id
        ),
        "",
    )
    if not target:
        return ""
    if target.startswith("/"):
        return target.lstrip("/")
    if target.startswith("xl/"):
        return target
    return normpath(f"xl/{target}")


def _cached_formula_scalar(cell_type: str, value: str) -> object:
    if cell_type == "e":
        return value
    if cell_type == "b":
        return value == "1"
    if cell_type in {"", "n"}:
        try:
            return Decimal(value)
        except InvalidOperation:
            return value
    return value


def _local_name(name: str) -> str:
    return name.rsplit("}", 1)[-1]


def _style_header(sheet, headers: tuple[str, ...], last_column: str) -> None:
    sheet.append(headers)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{last_column}1"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_legacy_sheet(sheet, records: list[PreviewRecord]) -> None:
    sheet.title = "OPIU Light"
    _style_header(sheet, EXPORT_HEADERS, "X")
    for record in records:
        organization, organization_code = _record_organization_reference(record)
        department = record.department or None
        cfo = record.erp_department or record.cfo or None
        sheet.append(
            (
                record.reporting_unit,
                organization or None,
                organization_code or None,
                record.scenario,
                record.year,
                record.month,
                f"{record.month:02d}.{record.year}",
                department,
                record.organization_type,
                cfo,
                cfo,
                record.cfo_code or None,
                record.indicator_type_label,
                record.indicator or record.source_article,
                record.expense_type,
                record.expense_group,
                record.source_article,
                record.erp_code,
                record.erp_article_name,
                record.tax,
                float(record.amount) if record.amount is not None else None,
                record.status,
                record.comment,
                record.source_row,
            )
        )
    sheet.auto_filter.ref = f"A1:X{max(len(records) + 1, 1)}"
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["G"].width = 12
    for column in (
        "H", "I", "J", "K", "M", "N", "O", "P", "Q", "S", "T", "V", "W"
    ):
        sheet.column_dimensions[column].width = 22
    sheet.column_dimensions["L"].width = 16
    sheet.column_dimensions["R"].width = 18
    sheet.column_dimensions["U"].width = 16
    sheet.column_dimensions["X"].width = 18
    for column in ("C", "L", "R"):
        for cell in sheet[column]:
            cell.number_format = "@"
    sheet["U1"].alignment = Alignment(horizontal="center", vertical="center")
    for cell in sheet["U"][1:]:
        cell.number_format = '#,##0.00;[Red](#,##0.00);-'


def _write_ado_opiu_sheet(sheet, records: list[PreviewRecord]) -> None:
    sheet.title = "ОПИУ"
    _style_header(sheet, ADO_OPIU_HEADERS, "T")
    for record in records:
        organization, organization_code = _record_organization_reference(record)
        department = record.department or None
        cfo = record.erp_department or record.cfo or None
        sheet.append(
            (
                organization or None,
                organization_code or None,
                record.scenario,
                record.year,
                record.month,
                f"{record.month:02d}.{record.year}",
                department,
                cfo,
                cfo,
                record.cfo_code or None,
                record.indicator_type_label,
                record.indicator or record.source_article,
                record.expense_type,
                record.erp_code or None,
                record.erp_article_name or record.source_article,
                record.nomenclature or None,
                None,  # Код номенклатуры.
                record.sales_region or None,
                None,  # Код региона продаж.
                float(record.amount) if record.amount is not None else None,
            )
        )
    sheet.auto_filter.ref = f"A1:T{max(len(records) + 1, 1)}"
    widths = {
        "A": 28,
        "B": 20,
        "C": 18,
        "D": 10,
        "E": 10,
        "F": 12,
        "G": 28,
        "H": 28,
        "I": 28,
        "J": 16,
        "K": 18,
        "L": 30,
        "M": 24,
        "N": 18,
        "O": 30,
        "P": 24,
        "Q": 18,
        "R": 20,
        "S": 18,
        "T": 16,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for column in ("B", "J", "N", "Q", "S"):
        for cell in sheet[column]:
            cell.number_format = "@"
    for cell in sheet["T"][1:]:
        cell.number_format = '#,##0.00;[Red](#,##0.00);-'


def _write_ado_indicators_sheet(
    sheet,
    rows: list[IndicatorExportRow],
    records: list[PreviewRecord],
) -> None:
    sheet.title = "Показатели"
    _style_header(sheet, ADO_INDICATOR_HEADERS, "N")
    organization_references = _indicator_organization_references(records)
    for row in rows:
        organization, organization_code = organization_references.get(
            row.organization,
            _separate_organization_reference(row.organization),
        )
        sheet.append(
            (
                organization or None,
                organization_code or None,
                row.department or None,
                row.cfo or None,
                row.cfo or None,
                row.cfo_code or None,
                row.scenario,
                row.year,
                row.month,
                row.period,
                row.indicator_type,
                row.sales_channel,
                row.indicator,
                float(row.amount),
            )
        )
    sheet.auto_filter.ref = f"A1:N{max(len(rows) + 1, 1)}"
    widths = {
        "A": 28,
        "B": 20,
        "C": 28,
        "D": 28,
        "E": 28,
        "F": 16,
        "G": 18,
        "H": 10,
        "I": 10,
        "J": 12,
        "K": 18,
        "L": 26,
        "M": 24,
        "N": 16,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for column in ("B", "F"):
        for cell in sheet[column]:
            cell.number_format = "@"
    sheet["N1"].alignment = Alignment(horizontal="center", vertical="center")
    for cell in sheet["N"][1:]:
        cell.number_format = '#,##0.00;[Red](#,##0.00);-'


def export_opiu_light(records: list[PreviewRecord]) -> bytes:
    """Export the legacy sheet plus the two-sheet ADO schema in one workbook."""

    workbook = Workbook()
    try:
        legacy_sheet = workbook.active
        _write_legacy_sheet(legacy_sheet, records)
        _write_ado_opiu_sheet(workbook.create_sheet(), records)
        _write_ado_indicators_sheet(
            workbook.create_sheet(),
            aggregate_indicator_rows(records),
            records,
        )

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
    finally:
        workbook.close()
