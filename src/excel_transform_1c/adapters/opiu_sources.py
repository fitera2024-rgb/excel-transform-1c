from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from excel_transform_1c.core.opiu_rules.opiu_formula_parser import (
    parse_analytic_rows,
    parse_erp_source_rules,
    parse_formula_rows,
)
from excel_transform_1c.core.opiu_rules.opiu_rule_builder import (
    build_opiu_rule_catalog,
)
from excel_transform_1c.core.opiu_rules.opiu_rule_models import (
    AnalyticInputRow,
    CatalogEntry,
    ERPIndicatorCatalogEntry,
    FormulaInputRow,
    OPIURuleCatalog,
)


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _open(content: bytes, *, read_only: bool):
    return load_workbook(
        BytesIO(content),
        data_only=False,
        read_only=read_only,
        keep_links=False,
    )


def parse_formula_workbook(content: bytes):
    workbook = _open(content, read_only=False)
    try:
        for sheet in workbook.worksheets:
            if _clean(sheet.cell(1, 1).value) != "Строка":
                continue
            measure = _clean(sheet.cell(1, 2).value)
            if not measure:
                continue
            rows = []
            for row_number in range(2, sheet.max_row + 1):
                cell = sheet.cell(row_number, 1)
                rows.append(
                    FormulaInputRow(
                        source_row=row_number,
                        name=_clean(cell.value),
                        formula=_clean(sheet.cell(row_number, 2).value),
                        indent=int(cell.alignment.indent or 0),
                        measure=measure,
                    )
                )
            return parse_formula_rows(rows)
    finally:
        workbook.close()
    raise ValueError("ОПИУ ФОРМУЛЫ: не найдены точные заголовки Строка/Сумма")


def parse_analytic_workbook(content: bytes):
    workbook = _open(content, read_only=False)
    try:
        for sheet in workbook.worksheets:
            if _clean(sheet.cell(1, 1).value) != "Строка":
                continue
            headers = [_clean(sheet.cell(1, column).value) for column in range(1, 10)]
            if "Аналитика 1" not in headers:
                continue
            rows = []
            for row_number in range(2, sheet.max_row + 1):
                cell = sheet.cell(row_number, 1)
                rows.append(
                    AnalyticInputRow(
                        source_row=row_number,
                        name=_clean(cell.value),
                        analytics=tuple(
                            _clean(sheet.cell(row_number, column).value)
                            for column in range(2, 10)
                        ),
                        indent=int(cell.alignment.indent or 0),
                    )
                )
            return parse_analytic_rows(rows)
    finally:
        workbook.close()
    raise ValueError("ОПИУ АНАЛИТИКИ: не найдены структурные заголовки")


def _indicator_header(sheet) -> tuple[int, dict[str, int]]:
    required = {"Показатели отчетов", "Строка", "Колонка", "Код"}
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30), values_only=True),
        start=1,
    ):
        headers = {_clean(value): column for column, value in enumerate(values, start=1)}
        if required.issubset(headers):
            return row_number, headers
    raise ValueError("ERP-каталог: не найдены структурные заголовки показателей")


def parse_indicator_catalog(content: bytes) -> tuple[ERPIndicatorCatalogEntry, ...]:
    workbook = _open(content, read_only=True)
    try:
        for sheet in workbook.worksheets:
            try:
                header_row, columns = _indicator_header(sheet)
            except ValueError:
                continue
            entries: list[ERPIndicatorCatalogEntry] = []
            for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                value = lambda header: _clean(values[columns[header] - 1])
                name = value("Показатели отчетов")
                report_line = value("Строка")
                code = value("Код")
                if not name or not report_line or not code:
                    continue
                entries.append(
                    ERPIndicatorCatalogEntry(
                        indicator_code=code,
                        name=name,
                        report_line=report_line,
                        indicator_group=(
                            value("Группа аналитик")
                            if "Группа аналитик" in columns
                            else ""
                        ),
                        column=value("Колонка"),
                        normalized_code=(
                            value("Нормализованный код")
                            if "Нормализованный код" in columns
                            else ""
                        ),
                    )
                )
            return tuple(dict.fromkeys(entries))
    finally:
        workbook.close()
    raise ValueError("ERP-каталог показателей не распознан")


def parse_code_name_catalog(content: bytes, label: str) -> tuple[CatalogEntry, ...]:
    workbook = _open(content, read_only=True)
    try:
        for sheet in workbook.worksheets:
            header_row = None
            name_column = None
            code_column = None
            for row_number, values in enumerate(
                sheet.iter_rows(
                    min_row=1,
                    max_row=min(sheet.max_row, 30),
                    values_only=True,
                ),
                start=1,
            ):
                headers = {
                    _clean(value): column
                    for column, value in enumerate(values, start=1)
                }
                if "Наименование" in headers and "Код" in headers:
                    header_row = row_number
                    name_column = headers["Наименование"]
                    code_column = headers["Код"]
                    break
            if header_row is None or name_column is None or code_column is None:
                continue
            entries = []
            for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                name = _clean(values[name_column - 1])
                code = _clean(values[code_column - 1])
                if name or code:
                    entries.append(CatalogEntry(code=code, name=name))
            return tuple(dict.fromkeys(entries))
    finally:
        workbook.close()
    raise ValueError(f"{label}: не найдены точные заголовки Код/Наименование")


def build_catalog_from_source_bytes(
    *,
    formulas_xlsx: bytes,
    analytics_xlsx: bytes,
    indicators_xlsx: bytes,
    sources_mxl: bytes,
    regions_xlsx: bytes,
    networks_xlsx: bytes,
) -> OPIURuleCatalog:
    return build_opiu_rule_catalog(
        formula_rules=parse_formula_workbook(formulas_xlsx),
        analytic_rules=parse_analytic_workbook(analytics_xlsx),
        indicator_catalog=parse_indicator_catalog(indicators_xlsx),
        source_rules=parse_erp_source_rules(sources_mxl),
        region_catalog=parse_code_name_catalog(regions_xlsx, "Регионы"),
        network_catalog=parse_code_name_catalog(networks_xlsx, "Сети"),
    )
