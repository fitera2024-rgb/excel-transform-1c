from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from excel_transform_1c.core.detection import normalize_header
from excel_transform_1c.core.models import ERPArticle, OrganizationNode


ERP_HEADERS = {
    "code": {"код"},
    "name": {"официальное наименование", "наименование erp"},
    "expense_type": {"тип расходов"},
    "expense_group": {"группа расходов"},
    "source_article": {"статья", "исходная статья"},
}

ORG_HEADERS = {
    "node_id": {"id", "идентификатор"},
    "code": {"код"},
    "name": {"наименование", "узел"},
    "parent_id": {"родитель id", "parent id"},
    "full_path": {"полный путь", "путь"},
}

SCENARIO_HEADERS = {
    "name": {"наименование", "сценарий"},
    "year": {"год"},
    "erp_code": {"erp-код", "код"},
    "comment": {"комментарий"},
}

REAL_EXPORT_HEADERS = {
    "erp_articles": {
        "name": {
            "статья доходов и расходов",
            "статьи доходов и расходов",
            "иерархия статей доходов и расходов",
        },
        "code": {"код", "код элемента", "код справочника", "код записи"},
    },
    "organizations": {
        "name": {
            "организации",
            "организация",
            "структура организаций",
            "иерархия организаций",
            "наименование",
        },
        "code": {"код", "код элемента", "код справочника", "код записи"},
    },
    "scenarios": {
        "name": {
            "сценарии",
            "сценарий",
            "сценарии бюджетирования",
            "наименование",
        },
        "code": {"код", "erp-код", "код элемента", "код справочника", "код записи"},
    },
}

PATH_ALIASES = {"полный путь", "путь", "полное наименование"}
MAX_HEADER_SCAN_ROWS = 120


def parse_reference_workbook(content: bytes, kind: str) -> list[dict[str, Any]]:
    if kind not in {"erp_articles", "organizations", "scenarios"}:
        raise ValueError("Неизвестный тип справочника")

    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=False)
    except Exception as exc:
        raise ValueError("Файл справочника не открывается или повреждён") from exc

    flat = _parse_flat_interchange(workbook, kind)
    if flat is not None:
        return flat

    if kind == "erp_articles":
        result = _parse_real_erp_articles(workbook)
    elif kind == "organizations":
        result = _parse_real_organizations(workbook)
    else:
        result = _parse_real_scenarios(workbook)

    if not result:
        raise ValueError(
            "Не найден распознаваемый диапазон справочника. "
            "Загрузите известную ERP-выгрузку либо документированный плоский interchange-файл."
        )
    return result


def _parse_flat_interchange(workbook: Any, kind: str) -> list[dict[str, Any]] | None:
    required = {
        "erp_articles": ERP_HEADERS,
        "organizations": ORG_HEADERS,
        "scenarios": SCENARIO_HEADERS,
    }[kind]
    candidates: list[tuple[Any, int, dict[str, int]]] = []
    for sheet in workbook.worksheets:
        for row_number in range(1, min(sheet.max_row, 80) + 1):
            values = [cell.value for cell in sheet[row_number]]
            columns = _match_headers(values, required)
            if columns:
                candidates.append((sheet, row_number, columns))
    if not candidates:
        return None
    if len(candidates) != 1:
        raise ValueError("Справочник содержит несколько плоских структурно подходящих диапазонов")

    sheet, header_row, columns = candidates[0]
    rows: list[dict[str, Any]] = []
    for row_number in range(header_row + 1, sheet.max_row + 1):
        cells = {field: sheet.cell(row_number, column) for field, column in columns.items()}
        if all(_is_blank(cell.value) for cell in cells.values()):
            continue
        rows.append(
            {
                field: _clean_flat_cell(field, cell)
                for field, cell in cells.items()
            }
        )
    return rows


def _parse_real_erp_articles(workbook: Any) -> list[dict[str, Any]]:
    sheet, first_data_row, name_col, code_col = _best_real_layout(workbook, "erp_articles")
    stack: list[str] = []
    result: list[dict[str, Any]] = []
    seen_codes: set[str] = set()

    for row_number in range(first_data_row, sheet.max_row + 1):
        raw_name = sheet.cell(row_number, name_col).value
        code_cell = sheet.cell(row_number, code_col)

        if not _is_blank(raw_name):
            name = _hierarchy_text(raw_name)
            if not _looks_like_header(name, REAL_EXPORT_HEADERS["erp_articles"]["name"]):
                _set_stack(stack, _row_level(sheet, row_number, name_col, raw_name), name)

        code = _code_text(code_cell)
        if not code or _looks_like_header(code, REAL_EXPORT_HEADERS["erp_articles"]["code"]):
            continue
        if code in seen_codes:
            raise ValueError(f"ERP-справочник статей содержит повторяющийся код: {code}")
        if not stack:
            continue

        article = stack[-1]
        expense_type = stack[0] if len(stack) >= 2 else ""
        expense_group = stack[-2] if len(stack) >= 3 else ""
        result.append(
            {
                "code": code,
                "name": article,
                "expense_type": expense_type,
                "expense_group": expense_group,
                "source_article": article,
            }
        )
        seen_codes.add(code)

    return result


def _parse_real_organizations(workbook: Any) -> list[dict[str, Any]]:
    sheet, first_data_row, name_col, code_col = _best_real_layout(workbook, "organizations")
    path_col = _find_optional_column(
        sheet,
        first_data_row,
        PATH_ALIASES,
        excluded_columns={name_col, code_col},
    )

    stack: list[str] = []
    coded_stack: dict[int, str] = {}
    raw_nodes: list[dict[str, Any]] = []
    seen_codes: set[str] = set()

    for row_number in range(first_data_row, sheet.max_row + 1):
        raw_name = sheet.cell(row_number, name_col).value
        code_cell = sheet.cell(row_number, code_col)

        level = 0
        if not _is_blank(raw_name):
            name = _hierarchy_text(raw_name)
            if _looks_like_header(name, REAL_EXPORT_HEADERS["organizations"]["name"]):
                continue
            level = _row_level(sheet, row_number, name_col, raw_name)
            _set_stack(stack, level, name)
        elif stack:
            name = stack[-1]
            level = max(len(stack) - 1, 0)
        else:
            continue

        code = _code_text(code_cell)
        if not code or _looks_like_header(code, REAL_EXPORT_HEADERS["organizations"]["code"]):
            continue
        if code in seen_codes:
            raise ValueError(f"Справочник организаций содержит повторяющийся код: {code}")

        explicit_path = ""
        if path_col is not None:
            explicit_path = _clean_scalar(sheet.cell(row_number, path_col).value)
        full_path = explicit_path or " → ".join(stack)

        parent_id = None
        for parent_level in range(level - 1, -1, -1):
            if parent_level in coded_stack:
                parent_id = coded_stack[parent_level]
                break

        raw_nodes.append(
            {
                "node_id": code,
                "code": code,
                "name": name,
                "parent_id": parent_id,
                "full_path": full_path,
            }
        )
        seen_codes.add(code)
        coded_stack[level] = code
        for deeper in [item for item in coded_stack if item > level]:
            del coded_stack[deeper]

    if path_col is not None:
        by_path = {node["full_path"]: node["node_id"] for node in raw_nodes}
        for node in raw_nodes:
            separator = " → " if " → " in node["full_path"] else None
            if separator:
                parent_path = node["full_path"].rsplit(separator, 1)[0]
                node["parent_id"] = by_path.get(parent_path)

    return raw_nodes


def _parse_real_scenarios(workbook: Any) -> list[dict[str, Any]]:
    sheet, first_data_row, name_col, code_col = _best_real_layout(workbook, "scenarios")
    result: list[dict[str, Any]] = []
    seen_names: set[str] = set()

    for row_number in range(first_data_row, sheet.max_row + 1):
        raw_name = sheet.cell(row_number, name_col).value
        if _is_blank(raw_name):
            continue
        name = str(raw_name).strip()
        if not name or _looks_like_header(name, REAL_EXPORT_HEADERS["scenarios"]["name"]):
            continue

        code = _code_text(sheet.cell(row_number, code_col))
        year_match = re.search(r"(?<!\d)(20\d{2}|21\d{2})(?!\d)", name.replace("_", " "))
        year = int(year_match.group(1)) if year_match else 0
        identity = f"{name}\0{year}"
        if identity in seen_names:
            continue
        seen_names.add(identity)
        result.append(
            {
                "name": name,
                "year": str(year),
                "erp_code": code,
                "comment": "",
            }
        )
    return result


def _best_real_layout(workbook: Any, kind: str) -> tuple[Any, int, int, int]:
    aliases = REAL_EXPORT_HEADERS[kind]
    candidates: list[tuple[int, Any, int, int, int]] = []

    for sheet in workbook.worksheets:
        name_headers = _header_positions(sheet, aliases["name"], allow_contains=True)
        code_headers = _header_positions(sheet, aliases["code"], allow_contains=True)

        for name_row, name_col, name_quality in name_headers:
            for code_row, code_col, code_quality in code_headers:
                if name_col == code_col:
                    continue
                first_data_row = max(name_row, code_row) + 1
                score = _layout_score(sheet, first_data_row, name_col, code_col, kind)
                if score <= 0:
                    continue
                score += name_quality * 50 + code_quality * 50
                score += max(0, 30 - abs(name_row - code_row))
                if name_col < code_col:
                    score += 20
                candidates.append((score, sheet, first_data_row, name_col, code_col))

    if not candidates:
        expected_name = " / ".join(sorted(aliases["name"]))
        expected_code = " / ".join(sorted(aliases["code"]))
        raise ValueError(
            "Не найден заголовок известной ERP-выгрузки. "
            f"Ожидается поле наименования ({expected_name}) и поле кода ({expected_code}); "
            "они могут находиться на соседних строках заголовка."
        )

    candidates.sort(key=lambda item: item[0], reverse=True)
    _, sheet, first_data_row, name_col, code_col = candidates[0]
    return sheet, first_data_row, name_col, code_col


def _header_positions(
    sheet: Any,
    aliases: set[str],
    *,
    allow_contains: bool,
) -> list[tuple[int, int, int]]:
    results: list[tuple[int, int, int]] = []
    max_row = min(sheet.max_row, MAX_HEADER_SCAN_ROWS)
    for row_number in range(1, max_row + 1):
        for column_number, cell in enumerate(sheet[row_number], start=1):
            quality = _header_quality(cell.value, aliases, allow_contains=allow_contains)
            if quality:
                results.append((row_number, column_number, quality))
    return results


def _header_quality(value: Any, aliases: set[str], *, allow_contains: bool) -> int:
    normalized = normalize_header(value)
    if not normalized:
        return 0

    normalized_aliases = {normalize_header(alias) for alias in aliases}
    if normalized in normalized_aliases:
        return 3

    if "код" in normalized_aliases and re.match(r"^код(?:\s|$|\()", normalized):
        return 2

    if allow_contains:
        for alias in normalized_aliases:
            if len(alias) >= 4 and (alias in normalized or normalized in alias):
                return 2

    compact = re.sub(r"[^0-9a-zа-я]+", " ", normalized).strip()
    for alias in normalized_aliases:
        compact_alias = re.sub(r"[^0-9a-zа-я]+", " ", alias).strip()
        if compact == compact_alias:
            return 2
        if allow_contains and len(compact_alias) >= 4 and compact_alias in compact:
            return 1
    return 0


def _layout_score(
    sheet: Any,
    first_data_row: int,
    name_col: int,
    code_col: int,
    kind: str,
) -> int:
    if first_data_row > sheet.max_row:
        return 0

    code_values: list[str] = []
    name_count = 0
    rows_with_both = 0
    for row_number in range(first_data_row, sheet.max_row + 1):
        name_value = sheet.cell(row_number, name_col).value
        code_value = _code_text(sheet.cell(row_number, code_col))
        if not _is_blank(name_value):
            name_count += 1
        if code_value and not _looks_like_header(code_value, REAL_EXPORT_HEADERS[kind]["code"]):
            code_values.append(code_value)
            if not _is_blank(name_value):
                rows_with_both += 1

    if not code_values or not name_count:
        return 0

    unique_codes = len(set(code_values))
    score = min(len(code_values), 1000) * 4 + min(unique_codes, 1000) * 6
    score += min(name_count, 1000) + min(rows_with_both, 500) * 2
    if unique_codes == len(code_values):
        score += 30
    return score


def _find_optional_column(
    sheet: Any,
    first_data_row: int,
    aliases: set[str],
    *,
    excluded_columns: set[int] | None = None,
) -> int | None:
    matches: list[tuple[int, int]] = []
    excluded = excluded_columns or set()
    scan_to = min(max(first_data_row, 1), MAX_HEADER_SCAN_ROWS)
    for row_number in range(1, scan_to + 1):
        for index, cell in enumerate(sheet[row_number], start=1):
            if index in excluded:
                continue
            quality = _header_quality(cell.value, aliases, allow_contains=True)
            if quality:
                matches.append((quality, index))
    if not matches:
        return None
    matches.sort(reverse=True)
    best_quality = matches[0][0]
    best_columns = {column for quality, column in matches if quality == best_quality}
    return next(iter(best_columns)) if len(best_columns) == 1 else None


def _row_level(sheet: Any, row_number: int, name_col: int, value: Any) -> int:
    cell = sheet.cell(row_number, name_col)
    indent = int(cell.alignment.indent or 0)
    outline = int(sheet.row_dimensions[row_number].outlineLevel or 0)
    text = str(value)
    leading = len(text) - len(text.lstrip(" \t"))
    leading_level = leading // 2
    return max(indent, outline, leading_level)


def _set_stack(stack: list[str], level: int, name: str) -> None:
    level = max(level, 0)
    if level > len(stack):
        level = len(stack)
    del stack[level:]
    stack.append(name)


def _hierarchy_text(value: Any) -> str:
    return str(value).lstrip(" \t\r\n")


def _looks_like_header(value: Any, aliases: set[str]) -> bool:
    return _header_quality(value, aliases, allow_contains=True) > 0


def _match_headers(values: list[Any], required: dict[str, set[str]]) -> dict[str, int] | None:
    normalized = {index: normalize_header(value) for index, value in enumerate(values, start=1)}
    columns: dict[str, int] = {}
    for field, aliases in required.items():
        normalized_aliases = {normalize_header(alias) for alias in aliases}
        matches = [index for index, value in normalized.items() if value in normalized_aliases]
        if len(matches) != 1:
            return None
        columns[field] = matches[0]
    return columns


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _clean_scalar(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _clean_flat_cell(field: str, cell: Any) -> str:
    if cell.value is None:
        return ""
    if field in {"code", "node_id", "parent_id", "erp_code"}:
        return _code_text(cell)
    text = str(cell.value)
    if field in {"year", "comment"}:
        return text.strip()
    return text


def _code_text(cell: Any) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        number_format = str(cell.number_format or "")
        if re.fullmatch(r"0+", number_format):
            return f"{value:0{len(number_format)}d}"
        return str(value)
    if isinstance(value, float) and value.is_integer():
        number_format = str(cell.number_format or "")
        if re.fullmatch(r"0+", number_format):
            return f"{int(value):0{len(number_format)}d}"
        return str(int(value))
    return str(value).strip()


def erp_articles(payload: list[dict[str, Any]]) -> list[ERPArticle]:
    return [ERPArticle(**item) for item in payload]


def organization_nodes(payload: list[dict[str, Any]]) -> list[OrganizationNode]:
    return [
        OrganizationNode(
            node_id=item["node_id"],
            code=item["code"],
            name=item["name"],
            parent_id=item["parent_id"] or None,
            full_path=item["full_path"],
        )
        for item in payload
    ]