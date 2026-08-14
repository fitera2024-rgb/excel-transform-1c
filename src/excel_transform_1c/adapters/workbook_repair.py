from __future__ import annotations

import hashlib
import os
import re
import shutil
import struct
import zlib
from dataclasses import dataclass
from enum import StrEnum
from pathlib import Path, PurePosixPath
from tempfile import NamedTemporaryFile
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, BadZipFile, ZipFile, ZipInfo

import olefile
import xlrd
from openpyxl import Workbook


OLE_COMPOUND_FILE_SIGNATURE = bytes.fromhex("D0CF11E0A1B11AE1")
ZIP_SIGNATURES = (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")
SPREADSHEETML_NAMESPACE = "urn:schemas-microsoft-com:office:spreadsheet"

_LOCAL_FILE_HEADER = b"PK\x03\x04"
_DATA_DESCRIPTOR = b"PK\x07\x08"
_CENTRAL_DIRECTORY = b"PK\x01\x02"
_END_OF_CENTRAL_DIRECTORY = b"PK\x05\x06"
_ILLEGAL_XML_BYTES = re.compile(rb"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_UNESCAPED_AMPERSAND = re.compile(
    rb"&(?!amp;|lt;|gt;|apos;|quot;|#[0-9]+;|#x[0-9A-Fa-f]+;)"
)
_OOXML_REQUIRED_PARTS = frozenset(
    {"[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"}
)
_MAX_ARCHIVE_MEMBERS = 20_000
_MAX_UNCOMPRESSED_BYTES = 512 * 1024 * 1024


class WorkbookFormat(StrEnum):
    OOXML = "ooxml"
    LEGACY_BIFF = "legacy_biff"
    ENCRYPTED_OOXML = "encrypted_ooxml"
    XML_SPREADSHEET = "xml_spreadsheet"
    UNKNOWN = "unknown"


class WorkbookRepairError(ValueError):
    """The selected file cannot be prepared without guessing at its contents."""


@dataclass(frozen=True)
class PreparedWorkbook:
    original_path: Path
    working_path: Path
    format: WorkbookFormat
    repaired: bool = False


@dataclass(frozen=True)
class RepairResult:
    path: Path
    changed: bool
    repaired_parts: tuple[str, ...] = ()


def detect_workbook_format(path: str | Path) -> WorkbookFormat:
    """Detect an Excel container from bytes and internal parts, never its suffix."""

    source_path = Path(path)
    try:
        with source_path.open("rb") as source:
            prefix = source.read(4096)
    except OSError:
        return WorkbookFormat.UNKNOWN

    if prefix.startswith(OLE_COMPOUND_FILE_SIGNATURE):
        return _detect_ole_format(source_path)
    if prefix.startswith(ZIP_SIGNATURES):
        return (
            WorkbookFormat.OOXML
            if _has_ooxml_parts(source_path)
            else WorkbookFormat.UNKNOWN
        )
    if _looks_like_spreadsheetml(prefix):
        return WorkbookFormat.XML_SPREADSHEET
    return WorkbookFormat.UNKNOWN


def prepare_workbook(
    source_path: str | Path,
    working_path: str | Path | None = None,
) -> PreparedWorkbook:
    """Return a readable OOXML path while preserving the input byte-for-byte."""

    source = Path(source_path)
    workbook_format = detect_workbook_format(source)
    if workbook_format is WorkbookFormat.ENCRYPTED_OOXML:
        raise WorkbookRepairError(
            "Файл защищён. Укажите пароль и выберите файл повторно"
        )
    if workbook_format is WorkbookFormat.UNKNOWN:
        raise WorkbookRepairError(
            "Файл не распознан как поддерживаемая книга Excel"
        )

    if workbook_format is WorkbookFormat.OOXML:
        target = _resolve_working_path(source, working_path)
        try:
            _validate_ooxml(source)
        except WorkbookRepairError:
            result = repair_ooxml(source, target)
            return PreparedWorkbook(source, result.path, workbook_format, repaired=True)
        result = repair_ooxml(source, target)
        return PreparedWorkbook(source, result.path, workbook_format)

    target = _resolve_working_path(source, working_path)
    if workbook_format is WorkbookFormat.LEGACY_BIFF:
        _convert_legacy_biff(source, target)
    else:
        _convert_spreadsheetml(source, target)
    _validate_ooxml(target)
    return PreparedWorkbook(source, target, workbook_format)


def repair_ooxml(source_path: str | Path, target_path: str | Path) -> RepairResult:
    """Conservatively rebuild a recoverable OOXML package into a separate file."""

    source = Path(source_path)
    target = Path(target_path)
    _require_separate_paths(source, target)

    try:
        _validate_ooxml(source)
    except WorkbookRepairError:
        pass
    else:
        _atomic_copy(source, target)
        return RepairResult(target, changed=False)

    entries = _read_zip_entries(source)
    if not _OOXML_REQUIRED_PARTS.issubset(entries):
        raise WorkbookRepairError(
            "Файл Excel повреждён и не содержит обязательных частей книги"
        )

    repaired_parts: list[str] = []
    normalized_entries: dict[str, bytes] = {}
    for name, payload in entries.items():
        normalized = payload
        if _is_xml_part(name):
            normalized = _repair_xml(payload)
            if normalized != payload:
                repaired_parts.append(name)
        normalized_entries[name] = normalized

    _write_deterministic_zip(normalized_entries, target)
    try:
        _validate_ooxml(target)
    except WorkbookRepairError:
        target.unlink(missing_ok=True)
        raise
    return RepairResult(target, changed=True, repaired_parts=tuple(repaired_parts))


def default_working_path(source_path: str | Path) -> Path:
    source = Path(source_path)
    if source.stem == "source-original":
        return source.with_name("source-working.xlsx")
    digest = _sha256_path(source)[:12]
    return source.with_name(f".{source.stem}.{digest}.working.xlsx")


def _sha256_path(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        while chunk := source.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest()


def _detect_ole_format(path: Path) -> WorkbookFormat:
    try:
        with olefile.OleFileIO(path) as container:
            streams = {
                "/".join(part.casefold() for part in item)
                for item in container.listdir(streams=True, storages=False)
            }
    except (OSError, IOError, ValueError, TypeError, olefile.OleFileError):
        return WorkbookFormat.UNKNOWN

    if {"encryptioninfo", "encryptedpackage"}.issubset(streams):
        return WorkbookFormat.ENCRYPTED_OOXML
    if "workbook" in streams or "book" in streams:
        return WorkbookFormat.LEGACY_BIFF
    return WorkbookFormat.UNKNOWN


def _has_ooxml_parts(path: Path) -> bool:
    try:
        with ZipFile(path) as archive:
            names = set(archive.namelist())
    except (BadZipFile, OSError, EOFError):
        names = _local_header_names(path)
    return _OOXML_REQUIRED_PARTS.issubset(names)


def _local_header_names(path: Path) -> set[str]:
    try:
        payload = path.read_bytes()
    except OSError:
        return set()
    names: set[str] = set()
    position = 0
    while len(names) < _MAX_ARCHIVE_MEMBERS:
        position = payload.find(_LOCAL_FILE_HEADER, position)
        if position < 0 or position + 30 > len(payload):
            break
        try:
            fields = struct.unpack_from("<IHHHHHIIIHH", payload, position)
        except struct.error:
            break
        flags = fields[2]
        name_length = fields[9]
        extra_length = fields[10]
        name_start = position + 30
        name_end = name_start + name_length
        if name_end + extra_length > len(payload):
            break
        encoding = "utf-8" if flags & 0x800 else "cp437"
        try:
            name = payload[name_start:name_end].decode(encoding)
        except UnicodeDecodeError:
            position += 4
            continue
        if _safe_member_name(name):
            names.add(name)
        position = name_end + extra_length
    return names


def _looks_like_spreadsheetml(prefix: bytes) -> bool:
    sample = prefix
    if sample.startswith((b"\xff\xfe", b"\xfe\xff")):
        try:
            text = sample.decode("utf-16", errors="ignore")
        except UnicodeError:
            return False
    else:
        text = sample.decode("utf-8-sig", errors="ignore")
    lowered = text.casefold()
    return "<?xml" in lowered and SPREADSHEETML_NAMESPACE in lowered


def _resolve_working_path(source: Path, requested: str | Path | None) -> Path:
    target = Path(requested) if requested is not None else default_working_path(source)
    _require_separate_paths(source, target)
    target.parent.mkdir(parents=True, exist_ok=True)
    return target


def _require_separate_paths(source: Path, target: Path) -> None:
    if source.resolve() == target.resolve():
        raise WorkbookRepairError("Рабочая копия должна быть отдельным файлом")


def _validate_ooxml(path: Path) -> None:
    try:
        with ZipFile(path) as archive:
            infos = archive.infolist()
            if len(infos) > _MAX_ARCHIVE_MEMBERS:
                raise WorkbookRepairError("В книге Excel слишком много внутренних частей")
            names = {info.filename for info in infos}
            if not _OOXML_REQUIRED_PARTS.issubset(names):
                raise WorkbookRepairError(
                    "Файл Excel не содержит обязательных частей книги"
                )
            total = 0
            for info in infos:
                if not _safe_member_name(info.filename):
                    raise WorkbookRepairError("Книга Excel содержит небезопасный путь")
                total += info.file_size
                if total > _MAX_UNCOMPRESSED_BYTES:
                    raise WorkbookRepairError("Книга Excel слишком велика после распаковки")
                payload = archive.read(info)
                if _is_xml_part(info.filename):
                    _parse_xml(payload)
    except WorkbookRepairError:
        raise
    except (BadZipFile, OSError, EOFError, RuntimeError, ElementTree.ParseError) as exc:
        raise WorkbookRepairError("Пакет Excel повреждён") from exc


def _read_zip_entries(path: Path) -> dict[str, bytes]:
    try:
        with ZipFile(path) as archive:
            entries = _read_standard_zip(archive)
    except (BadZipFile, OSError, EOFError, RuntimeError, zlib.error):
        entries = _recover_local_entries(path)
    if not entries:
        raise WorkbookRepairError("Пакет Excel не удалось восстановить")
    return entries


def _read_standard_zip(archive: ZipFile) -> dict[str, bytes]:
    infos = archive.infolist()
    if len(infos) > _MAX_ARCHIVE_MEMBERS:
        raise WorkbookRepairError("В книге Excel слишком много внутренних частей")
    result: dict[str, bytes] = {}
    total = 0
    for info in infos:
        name = info.filename
        if name.endswith("/"):
            continue
        if not _safe_member_name(name) or name in result:
            raise WorkbookRepairError("Книга Excel содержит небезопасную структуру")
        total += info.file_size
        if total > _MAX_UNCOMPRESSED_BYTES:
            raise WorkbookRepairError("Книга Excel слишком велика после распаковки")
        result[name] = archive.read(info)
    return result


def _recover_local_entries(path: Path) -> dict[str, bytes]:
    try:
        payload = path.read_bytes()
    except OSError as exc:
        raise WorkbookRepairError("Файл Excel не удалось прочитать") from exc

    result: dict[str, bytes] = {}
    total = 0
    position = payload.find(_LOCAL_FILE_HEADER)
    while position >= 0 and position + 30 <= len(payload):
        if len(result) >= _MAX_ARCHIVE_MEMBERS:
            raise WorkbookRepairError("В книге Excel слишком много внутренних частей")
        try:
            fields = struct.unpack_from("<IHHHHHIIIHH", payload, position)
        except struct.error as exc:
            raise WorkbookRepairError("Локальная структура ZIP повреждена") from exc
        flags = fields[2]
        method = fields[3]
        expected_crc = fields[6]
        compressed_size = fields[7]
        uncompressed_size = fields[8]
        name_length = fields[9]
        extra_length = fields[10]
        if flags & 0x1:
            raise WorkbookRepairError("Внутренний ZIP-пакет зашифрован")
        name_start = position + 30
        name_end = name_start + name_length
        data_start = name_end + extra_length
        if data_start > len(payload):
            raise WorkbookRepairError("Локальная структура ZIP обрезана")
        encoding = "utf-8" if flags & 0x800 else "cp437"
        try:
            name = payload[name_start:name_end].decode(encoding)
        except UnicodeDecodeError as exc:
            raise WorkbookRepairError("Имя внутренней части Excel повреждено") from exc
        if not _safe_member_name(name) or name in result:
            raise WorkbookRepairError("Книга Excel содержит небезопасную структуру")

        if flags & 0x8:
            data, consumed = _inflate_unknown_size(payload[data_start:], method)
            descriptor_start = data_start + consumed
            position = _skip_data_descriptor(payload, descriptor_start)
        else:
            data_end = data_start + compressed_size
            if data_end > len(payload):
                raise WorkbookRepairError("Содержимое ZIP обрезано")
            compressed = payload[data_start:data_end]
            data = _decompress(compressed, method)
            position = data_end

        if not (flags & 0x8):
            if len(data) != uncompressed_size:
                raise WorkbookRepairError("Размер внутренней части Excel не совпадает")
            if (zlib.crc32(data) & 0xFFFFFFFF) != expected_crc:
                raise WorkbookRepairError("Контрольная сумма внутренней части неверна")
        total += len(data)
        if total > _MAX_UNCOMPRESSED_BYTES:
            raise WorkbookRepairError("Книга Excel слишком велика после распаковки")
        if name and not name.endswith("/"):
            result[name] = data

        next_header = payload.find(_LOCAL_FILE_HEADER, position)
        terminal_positions = [
            value
            for value in (
                payload.find(_CENTRAL_DIRECTORY, position),
                payload.find(_END_OF_CENTRAL_DIRECTORY, position),
            )
            if value >= 0
        ]
        if next_header < 0 or (terminal_positions and min(terminal_positions) < next_header):
            break
        position = next_header
    return result


def _inflate_unknown_size(payload: bytes, method: int) -> tuple[bytes, int]:
    if method != ZIP_DEFLATED:
        raise WorkbookRepairError(
            "ZIP с неизвестным размером поддерживается только для deflate"
        )
    inflater = zlib.decompressobj(-15)
    try:
        data = inflater.decompress(payload)
        data += inflater.flush()
    except zlib.error as exc:
        raise WorkbookRepairError("Сжатая часть Excel повреждена") from exc
    if not inflater.eof:
        raise WorkbookRepairError("Сжатая часть Excel обрезана")
    return data, len(payload) - len(inflater.unused_data)


def _skip_data_descriptor(payload: bytes, position: int) -> int:
    if payload[position:position + 4] == _DATA_DESCRIPTOR:
        return position + 16
    return position + 12


def _decompress(payload: bytes, method: int) -> bytes:
    if method == 0:
        return payload
    if method == ZIP_DEFLATED:
        try:
            return zlib.decompress(payload, -15)
        except zlib.error as exc:
            raise WorkbookRepairError("Сжатая часть Excel повреждена") from exc
    raise WorkbookRepairError("Метод сжатия ZIP не поддерживается")


def _safe_member_name(name: str) -> bool:
    pure = PurePosixPath(name)
    return bool(name) and not pure.is_absolute() and ".." not in pure.parts and "\\" not in name


def _is_xml_part(name: str) -> bool:
    lowered = name.casefold()
    return lowered.endswith(".xml") or lowered.endswith(".rels")


def _parse_xml(payload: bytes) -> None:
    ElementTree.fromstring(payload)


def _repair_xml(payload: bytes) -> bytes:
    try:
        _parse_xml(payload)
        return payload
    except ElementTree.ParseError:
        pass

    candidate = _strip_illegal_xml_characters(payload)
    try:
        _parse_xml(candidate)
        return candidate
    except ElementTree.ParseError:
        candidate = _UNESCAPED_AMPERSAND.sub(b"&amp;", candidate)
        try:
            _parse_xml(candidate)
        except ElementTree.ParseError as exc:
            raise WorkbookRepairError(
                "XML-часть книги повреждена и не может быть восстановлена безопасно"
            ) from exc
        return candidate


def _strip_illegal_xml_characters(payload: bytes) -> bytes:
    if payload.startswith((b"\xff\xfe", b"\xfe\xff")):
        encoding = "utf-16"
        text = payload.decode(encoding)
        cleaned = "".join(
            character
            for character in text
            if character in "\t\n\r" or ord(character) >= 0x20
        )
        return cleaned.encode(encoding)
    return _ILLEGAL_XML_BYTES.sub(b"", payload)


def _write_deterministic_zip(entries: dict[str, bytes], target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    temp_path: Path | None = None
    try:
        with NamedTemporaryFile(
            mode="wb",
            prefix=f".{target.name}.",
            suffix=".tmp",
            dir=target.parent,
            delete=False,
        ) as temp:
            temp_path = Path(temp.name)
        with ZipFile(temp_path, "w", compression=ZIP_DEFLATED, compresslevel=6) as archive:
            for name in sorted(entries):
                info = ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
                info.compress_type = ZIP_DEFLATED
                info.external_attr = 0o600 << 16
                archive.writestr(info, entries[name])
        os.replace(temp_path, target)
    except Exception:
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)
        target.unlink(missing_ok=True)
        raise


def _atomic_copy(source: Path, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    temp_path: Path | None = None
    try:
        with NamedTemporaryFile(
            mode="wb",
            prefix=f".{target.name}.",
            suffix=".tmp",
            dir=target.parent,
            delete=False,
        ) as temp:
            temp_path = Path(temp.name)
        shutil.copyfile(source, temp_path)
        os.replace(temp_path, target)
    except Exception:
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)
        target.unlink(missing_ok=True)
        raise


def _convert_legacy_biff(source: Path, target: Path) -> None:
    book = None
    workbook = Workbook(write_only=True)
    temp_path: Path | None = None
    try:
        book = xlrd.open_workbook(source, on_demand=True)
        for source_sheet in book.sheets():
            sheet = workbook.create_sheet(source_sheet.name)
            for row_index in range(source_sheet.nrows):
                values = [
                    _xlrd_value(book, source_sheet.cell(row_index, column_index))
                    for column_index in range(source_sheet.ncols)
                ]
                sheet.append(values)
        if not workbook.worksheets:
            workbook.create_sheet("Sheet1")
        with NamedTemporaryFile(
            mode="wb",
            prefix=f".{target.name}.",
            suffix=".tmp",
            dir=target.parent,
            delete=False,
        ) as temp:
            temp_path = Path(temp.name)
        workbook.save(temp_path)
        os.replace(temp_path, target)
    except Exception as exc:
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)
        target.unlink(missing_ok=True)
        raise WorkbookRepairError("Старый файл Excel не удалось открыть") from exc
    finally:
        workbook.close()
        if book is not None:
            book.release_resources()


def _xlrd_value(book: xlrd.book.Book, cell: xlrd.sheet.Cell):
    if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
        return None
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate.xldate_as_datetime(cell.value, book.datemode)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return xlrd.error_text_from_code.get(cell.value, "#VALUE!")
    return cell.value


def _convert_spreadsheetml(source: Path, target: Path) -> None:
    try:
        root = ElementTree.parse(source).getroot()
    except (OSError, ElementTree.ParseError) as exc:
        raise WorkbookRepairError("XML-файл Excel повреждён") from exc

    workbook = Workbook(write_only=True)
    temp_path: Path | None = None
    namespaces = {
        "ss": SPREADSHEETML_NAMESPACE,
        "x": "urn:schemas-microsoft-com:office:excel",
    }
    try:
        for index, source_sheet in enumerate(root.findall("ss:Worksheet", namespaces), start=1):
            title = source_sheet.attrib.get(
                f"{{{SPREADSHEETML_NAMESPACE}}}Name", f"Sheet{index}"
            )
            sheet = workbook.create_sheet(title)
            table = source_sheet.find("ss:Table", namespaces)
            if table is None:
                continue
            expected_row = 1
            for source_row in table.findall("ss:Row", namespaces):
                row_number = int(
                    source_row.attrib.get(
                        f"{{{SPREADSHEETML_NAMESPACE}}}Index", expected_row
                    )
                )
                while expected_row < row_number:
                    sheet.append([])
                    expected_row += 1
                values: list[object | None] = []
                expected_column = 1
                for source_cell in source_row.findall("ss:Cell", namespaces):
                    column_number = int(
                        source_cell.attrib.get(
                            f"{{{SPREADSHEETML_NAMESPACE}}}Index", expected_column
                        )
                    )
                    while expected_column < column_number:
                        values.append(None)
                        expected_column += 1
                    data = source_cell.find("ss:Data", namespaces)
                    values.append(_spreadsheetml_value(data))
                    expected_column += 1
                sheet.append(values)
                expected_row += 1
        if not workbook.worksheets:
            raise WorkbookRepairError("XML-файл не содержит листов Excel")
        with NamedTemporaryFile(
            mode="wb",
            prefix=f".{target.name}.",
            suffix=".tmp",
            dir=target.parent,
            delete=False,
        ) as temp:
            temp_path = Path(temp.name)
        workbook.save(temp_path)
        os.replace(temp_path, target)
    except Exception as exc:
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)
        target.unlink(missing_ok=True)
        if isinstance(exc, WorkbookRepairError):
            raise
        raise WorkbookRepairError("XML-файл Excel не удалось преобразовать") from exc
    finally:
        workbook.close()


def _spreadsheetml_value(data: ElementTree.Element | None):
    if data is None:
        return None
    value = data.text
    data_type = data.attrib.get(f"{{{SPREADSHEETML_NAMESPACE}}}Type", "String")
    if value is None:
        return None
    if data_type == "Number":
        number = float(value)
        return int(number) if number.is_integer() else number
    if data_type == "Boolean":
        return value != "0"
    return value
