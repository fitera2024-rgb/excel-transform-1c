from __future__ import annotations

import re
from collections.abc import Iterable
from dataclasses import replace
from datetime import date, datetime
from typing import Any

from openpyxl.utils import get_column_letter

from .models import CandidateRange, IndicatorType, MONTH_NAMES, SourceRow


BUSINESS_ALIASES: dict[str, set[str]] = {
    "reporting_unit": {"подразделение (цфо 1)", "единица отчета", "единица отчёта"},
    "expense_type": {"тип расходов"},
    "department": {"департамент (цфо 2)", "департамент"},
    "organization_type": {"вид организации"},
    "cfo": {"отдел", "цфо", "отдел / цфо"},
    "tax": {"налогообложение", "налог"},
    "expense_group": {"группа расходов", "группа"},
    "article": {"статья", "исходная статья"},
}

INCOME_ALIASES: dict[str, set[str]] = {
    "revenue_type": {"тип доходов"},
    "revenue_group": {"группа дохода", "группа доходов"},
    "article": {"статья", "исходная статья"},
    "analytics": {"аналитика", "аналитики"},
}

INDICATOR_ALIASES: dict[str, set[str]] = {
    "indicator_type": {"тип показателя"},
    "revenue_group": {"группа дохода", "группа доходов", "группа раскрытия"},
    "formula_condition": {"условие формулы", "условия формулы"},
    "analytics": {"аналитика", "аналитика1", "аналитики"},
    "nomenclature": {"номенклатура", "инт номенклатура"},
    "unit": {"единица измерения", "ед. изм.", "единица"},
    "counterparty": {"контрагент"},
    "input_sales_channel": {"инт канал сбыта", "канал сбыта"},
    "sales_network": {"сеть"},
    "sales_region": {"регион продаж"},
}

MONTH_ALIASES: dict[int, set[str]] = {
    1: {"январь", "янв", "01"},
    2: {"февраль", "фев", "02"},
    3: {"март", "мар", "03"},
    4: {"апрель", "апр", "04"},
    5: {"май", "05"},
    6: {"июнь", "июн", "06"},
    7: {"июль", "июл", "07"},
    8: {"август", "авг", "08"},
    9: {"сентябрь", "сен", "09"},
    10: {"октябрь", "окт", "10"},
    11: {"ноябрь", "ноя", "11"},
    12: {"декабрь", "дек", "12"},
}

INTALEV_INDICATOR_ALIASES = {"показатели", "показатель"}
INTALEV_METADATA_CFO = re.compile(r"(?im)^\s*цфо\s*:\s*(.+?)\s*$")
INTALEV_PERIOD = re.compile(
    r"^\s*(\d{2})\.(\d{2})\.(\d{4})\s*-\s*"
    r"(\d{2})\.(\d{2})\.(\d{4})\s*$"
)
TECHNICAL_TOTAL = re.compile(r"(?:^|\s)итого(?:\s|$)", re.IGNORECASE)
TECHNICAL_RATIO = re.compile(r"^(?:%|рентабельност)", re.IGNORECASE)
EXPENSE_WORD = re.compile(r"расход", re.IGNORECASE)
BDR_REVENUE_GROUP = "Выручка_продажи внешние"
BDR_REVENUE_ARTICLES = (
    "Опт",
    "Розница",
    "HoReCa",
    "Сети ДВ",
    "Сети Федеральные",
    "Дискаунтеры ДВ",
    "Дискаунтеры Федеральные",
)
BDR_INCOME_START = "Выручка ИТОГО"
BDR_EXPENSE_START = "Расходы по основной деятельности ИТОГО"
BDR_KPI_TAIL_START = "EBITDA"
BDR_INCOME_ANCHORS = frozenset(
    {
        BDR_INCOME_START,
        "Прочие доходы по основной деятельности",
        "Валовая прибыль",
    }
)
BDR_EXPENSE_ANCHORS = frozenset(
    {
        "Административные расходы",
        "Коммерческие расходы",
        "Расходы на транспортную логистику",
        "Расходы на складскую логистику",
    }
)
BDR_KPI_HEAD_ANCHORS = frozenset(
    {
        "Оборот в кг",
        "Выручка за 1 кг",
        "Себестоимость 1 кг",
        "Итого расходов на 1 кг",
        "Валовая прибыль на 1 кг",
    }
)
BDR_KPI_TAIL_ANCHORS = frozenset({BDR_KPI_TAIL_START, "Операционная прибыль"})


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("ё", "е").strip().casefold()
    return re.sub(r"\s+", " ", text)


def _column_schema(values: Iterable[Any]) -> dict[str, int] | None:
    headers = {index: normalize_header(value) for index, value in enumerate(values, start=1)}
    columns: dict[str, int] = {}
    for field, aliases in BUSINESS_ALIASES.items():
        normalized_aliases = {normalize_header(alias) for alias in aliases}
        matches = [index for index, value in headers.items() if value in normalized_aliases]
        if len(matches) != 1:
            return None
        columns[field] = matches[0]
    for month, aliases in MONTH_ALIASES.items():
        normalized_aliases = {normalize_header(alias) for alias in aliases}
        matches = [index for index, value in headers.items() if value in normalized_aliases]
        if len(matches) != 1:
            return None
        columns[f"month_{month}"] = matches[0]
    for field, aliases in INDICATOR_ALIASES.items():
        normalized_aliases = {normalize_header(alias) for alias in aliases}
        matches = [index for index, value in headers.items() if value in normalized_aliases]
        if len(matches) > 1:
            return None
        if matches:
            columns[field] = matches[0]
    return columns


def _income_column_schema(values: Iterable[Any]) -> dict[str, int] | None:
    """Recognize the owner income range by its complete structural header."""

    headers = {
        index: normalize_header(value)
        for index, value in enumerate(values, start=1)
    }
    columns: dict[str, int] = {}
    for field, aliases in INCOME_ALIASES.items():
        normalized_aliases = {normalize_header(alias) for alias in aliases}
        matches = [
            index for index, value in headers.items() if value in normalized_aliases
        ]
        if len(matches) != 1:
            return None
        columns[field] = matches[0]
    for month, aliases in MONTH_ALIASES.items():
        normalized_aliases = {normalize_header(alias) for alias in aliases}
        matches = [
            index for index, value in headers.items() if value in normalized_aliases
        ]
        if len(matches) != 1:
            return None
        columns[f"month_{month}"] = matches[0]

    # Dedicated input analytics remain optional, exact columns. The four
    # required income columns above are already authoritative for their fields.
    for field, aliases in INDICATOR_ALIASES.items():
        if field in {"revenue_group", "analytics"}:
            continue
        normalized_aliases = {normalize_header(alias) for alias in aliases}
        matches = [
            index for index, value in headers.items() if value in normalized_aliases
        ]
        if len(matches) > 1:
            return None
        if matches:
            columns[field] = matches[0]
    return columns


def _bdr_plan_month_schemas(
    rows: list[tuple[Any, ...]],
) -> list[tuple[dict[str, int], int, int]]:
    """Return exact January–December plan headers from the scanned sheet head."""

    result: list[tuple[dict[str, int], int, int]] = []
    for row_index, row in enumerate(rows[:-1]):
        next_row = rows[row_index + 1]
        for start in range(0, max(len(row) - 11, 0)):
            periods = [_date_month(row[start + offset]) for offset in range(12)]
            if any(period is None for period in periods):
                continue
            years = {period[0] for period in periods if period is not None}
            months = [period[1] for period in periods if period is not None]
            if len(years) != 1 or months != list(range(1, 13)):
                continue
            if not all(
                normalize_header(next_row[start + offset]) == "план"
                for offset in range(12)
            ):
                continue
            result.append(
                (
                    {f"month_{month}": start + month for month in range(1, 13)},
                    next(iter(years)),
                    row_index + 1,
                )
            )
    return result


def _bdr_department_column(
    rows: list[tuple[Any, ...]],
    month_header_row: int,
    first_month_column: int,
) -> int | None:
    """Find one exact BDR ``Отдел`` column next to the detected month schema."""

    matches: set[int] = set()
    first_header_row = max(month_header_row - 2, 1)
    last_header_row = min(month_header_row + 1, len(rows))
    for row_number in range(first_header_row, last_header_row + 1):
        row = rows[row_number - 1]
        for column in range(1, min(first_month_column, len(row) + 1)):
            if normalize_header(row[column - 1]) == "отдел":
                matches.add(column)
    return next(iter(matches)) if len(matches) == 1 else None


def _bdr_row_label(
    row: tuple[Any, ...], label_columns: tuple[int, ...]
) -> tuple[str, int] | None:
    labels = [
        (str(row[column - 1]).strip(), column)
        for column in label_columns
        if column <= len(row)
        and isinstance(row[column - 1], str)
        and str(row[column - 1]).strip()
    ]
    return labels[-1] if labels else None


def _has_bdr_month_value(row: tuple[Any, ...], columns: dict[str, int]) -> bool:
    return any(
        column <= len(row) and row[column - 1] not in (None, "")
        for field, column in columns.items()
        if field.startswith("month_")
    )


def _one_exact_row(
    labelled_rows: list[tuple[int, str]], label: str
) -> int | None:
    matches = [row_number for row_number, value in labelled_rows if value == label]
    return matches[0] if len(matches) == 1 else None


def _one_exact_row_between(
    labelled_rows: list[tuple[int, str]],
    label: str,
    *,
    after: int,
    before: int,
) -> int | None:
    """Return the single exact anchor inside an already proven block order."""

    matches = [
        row_number
        for row_number, value in labelled_rows
        if value == label and after < row_number < before
    ]
    return matches[0] if len(matches) == 1 else None


def _bdr_full_candidate(
    sheet: Any,
    scanned_rows: list[tuple[Any, ...]],
    max_scan_rows: int = 2000,
) -> CandidateRange | None:
    """Detect one complete BDR by its ordered business blocks and plan periods."""

    matches: list[CandidateRange] = []
    for month_columns, source_year, month_header_row in _bdr_plan_month_schemas(
        scanned_rows
    ):
        first_month_column = min(month_columns.values())
        department_column = _bdr_department_column(
            scanned_rows,
            month_header_row,
            first_month_column,
        )
        label_columns = tuple(
            column
            for column in range(1, first_month_column)
            if column != department_column
        )
        if not label_columns:
            continue
        last_scan_row = min(sheet.max_row or max_scan_rows, max_scan_rows)
        body = list(
            sheet.iter_rows(
                min_row=month_header_row + 2,
                max_row=last_scan_row,
                min_col=1,
                max_col=max(month_columns.values()),
                values_only=True,
            )
        )
        labelled_rows: list[tuple[int, str]] = []
        business_rows: list[int] = []
        for row_number, row in enumerate(body, start=month_header_row + 2):
            labelled = _bdr_row_label(row, label_columns)
            if labelled is None:
                continue
            label, _ = labelled
            labelled_rows.append((row_number, label))
            if _has_bdr_month_value(row, month_columns):
                business_rows.append(row_number)

        income_start = _one_exact_row(labelled_rows, BDR_INCOME_START)
        kpi_tail_start = _one_exact_row(labelled_rows, BDR_KPI_TAIL_START)
        expense_start = (
            _one_exact_row_between(
                labelled_rows,
                BDR_EXPENSE_START,
                after=income_start,
                before=kpi_tail_start,
            )
            if income_start is not None and kpi_tail_start is not None
            else None
        )
        if (
            income_start is None
            or expense_start is None
            or kpi_tail_start is None
            or not (income_start < expense_start < kpi_tail_start)
            or not business_rows
        ):
            continue

        exact_labels = {label for _, label in labelled_rows}
        if not (
            BDR_INCOME_ANCHORS.issubset(exact_labels)
            and BDR_EXPENSE_ANCHORS.issubset(exact_labels)
            and len(BDR_KPI_HEAD_ANCHORS.intersection(exact_labels)) >= 3
            and BDR_KPI_TAIL_ANCHORS.issubset(exact_labels)
        ):
            continue

        first_data_row = min(business_rows)
        last_data_row = max(business_rows)
        if not (first_data_row < income_start and kpi_tail_start <= last_data_row):
            continue

        header = scanned_rows[month_header_row - 1]
        reporting_values = [
            (str(header[column - 1]).strip(), column)
            for column in label_columns
            if column <= len(header)
            and isinstance(header[column - 1], str)
            and str(header[column - 1]).strip()
        ]
        source_cfo = reporting_values[0][0] if len(reporting_values) == 1 else ""
        reporting_unit_cell = (
            f"{get_column_letter(reporting_values[0][1])}{month_header_row}"
            if len(reporting_values) == 1
            else ""
        )
        matches.append(
            CandidateRange(
                candidate_id="",
                sheet=sheet.title,
                header_row=month_header_row,
                first_data_row=first_data_row,
                last_data_row=last_data_row,
                columns={
                    **month_columns,
                    **({"department": department_column} if department_column else {}),
                },
                source_kind="bdr_full",
                source_cfo=source_cfo,
                source_year=source_year,
                indicator_blocks=(
                    (IndicatorType.KPI, first_data_row, income_start - 1),
                    (IndicatorType.REVENUE, income_start, expense_start - 1),
                    (IndicatorType.EXPENSE, expense_start, kpi_tail_start - 1),
                    (IndicatorType.KPI, kpi_tail_start, last_data_row),
                ),
                label_columns=label_columns,
                reporting_unit_cell=reporting_unit_cell,
            )
        )

    # Multiple full schemas are an ambiguity, not an invitation to take the first.
    return matches[0] if len(matches) == 1 else None


def _bdr_value_label_rows(
    sheet: Any,
    candidate: CandidateRange,
) -> dict[int, str]:
    """Return exact KPI/revenue labels keyed by their source row."""

    result: dict[int, str] = {}
    rows = sheet.iter_rows(
        min_row=candidate.first_data_row,
        max_row=candidate.last_data_row,
        min_col=1,
        max_col=max(candidate.label_columns),
        values_only=True,
    )
    for row_number, row in enumerate(rows, start=candidate.first_data_row):
        if _bdr_indicator_type(candidate, row_number) not in {
            IndicatorType.KPI,
            IndicatorType.REVENUE,
        }:
            continue
        labelled = _bdr_row_label(row, candidate.label_columns)
        if labelled is not None:
            result[row_number] = labelled[0]
    return result


def _attach_exact_bdr_value_source(
    workbook: Any,
    candidate: CandidateRange,
    summaries: list[CandidateRange],
) -> CandidateRange:
    """Bind department-aware rows to one exact summary indicator/month grid."""

    same_year = [
        summary
        for summary in summaries
        if summary.source_year == candidate.source_year
    ]
    if len(same_year) != 1:
        return candidate
    summary = same_year[0]
    target_rows = _bdr_value_label_rows(workbook[candidate.sheet], candidate)
    source_rows = _bdr_value_label_rows(workbook[summary.sheet], summary)
    mapping: dict[int, int] = {
        target_row: target_row
        for target_row, label in target_rows.items()
        if source_rows.get(target_row) == label
    }

    target_by_label: dict[str, list[int]] = {}
    source_by_label: dict[str, list[int]] = {}
    for row_number, label in target_rows.items():
        target_by_label.setdefault(label, []).append(row_number)
    for row_number, label in source_rows.items():
        source_by_label.setdefault(label, []).append(row_number)
    for label, target_row_numbers in target_by_label.items():
        source_row_numbers = source_by_label.get(label, [])
        if len(target_row_numbers) == 1 and len(source_row_numbers) == 1:
            mapping.setdefault(target_row_numbers[0], source_row_numbers[0])

    target_labels = set(target_rows.values())
    if not BDR_KPI_HEAD_ANCHORS.intersection(target_labels) or not (
        BDR_KPI_TAIL_ANCHORS.issubset(target_labels)
    ):
        return candidate
    if len(mapping) < len(BDR_KPI_HEAD_ANCHORS.intersection(target_labels)) + len(
        BDR_KPI_TAIL_ANCHORS
    ):
        return candidate
    return replace(
        candidate,
        bdr_value_sheet=summary.sheet,
        bdr_value_columns={
            field: column
            for field, column in summary.columns.items()
            if field.startswith("month_")
        },
        bdr_value_rows=tuple(sorted(mapping.items())),
    )


def detect_candidate_ranges(
    workbook: Any,
    scan_rows: int = 100,
    scan_columns: int = 100,
) -> list[CandidateRange]:
    raw: list[tuple[Any, int, dict[str, int], str, str, int | None]] = []
    full_bdr_candidates: list[CandidateRange] = []
    for sheet in workbook.worksheets:
        max_row = min(sheet.max_row or scan_rows, scan_rows)
        max_column = min(sheet.max_column or scan_columns, scan_columns)
        scanned_rows = list(
            sheet.iter_rows(
                min_row=1,
                max_row=max_row,
                min_col=1,
                max_col=max_column,
                values_only=True,
            )
        )
        full_bdr = _bdr_full_candidate(sheet, scanned_rows)
        if full_bdr is not None:
            full_bdr_candidates.append(full_bdr)
        for row_number, values in enumerate(scanned_rows, start=1):
            schema = _column_schema(values)
            if schema:
                raw.append((sheet, row_number, schema, "prepared_budget", "", None))
                continue
            income_schema = _income_column_schema(values)
            if income_schema:
                raw.append(
                    (
                        sheet,
                        row_number,
                        income_schema,
                        "prepared_income_budget",
                        "",
                        None,
                    )
                )
                continue
            intalev = _intalev_schema(values)
            if intalev:
                source_cfo = _metadata_cfo(sheet, row_number)
                source_year = _period_year(values, intalev)
                raw.append(
                    (
                        sheet,
                        row_number,
                        intalev,
                        "intalev_opiu",
                        source_cfo,
                        source_year,
                    )
                )

        for parent_row, columns, source_year in _bdr_revenue_schemas(scanned_rows):
            raw.append(
                (
                    sheet,
                    parent_row,
                    columns,
                    "bdr_revenue_summary",
                    "",
                    source_year,
                )
            )

    if full_bdr_candidates:
        # A real owner workbook can contain both a department-aware planning
        # grid and an aggregate BDR summary.  The planning grid remains the
        # business candidate because it proves ``Отдел`` exactly; the summary
        # is retained only as the exact saved-value source for matching
        # KPI/revenue indicator/month pairs.
        department_candidates = [
            candidate
            for candidate in full_bdr_candidates
            if "department" in candidate.columns
        ]
        summary_candidates = [
            candidate
            for candidate in full_bdr_candidates
            if "department" not in candidate.columns
        ]
        selected = department_candidates or full_bdr_candidates
        if department_candidates and summary_candidates:
            selected = [
                _attach_exact_bdr_value_source(
                    workbook,
                    candidate,
                    summary_candidates,
                )
                for candidate in department_candidates
            ]
        return [
            replace(candidate, candidate_id=f"candidate-{index}")
            for index, candidate in enumerate(selected, start=1)
        ]

    candidates: list[CandidateRange] = []
    for index, (sheet, header_row, columns, source_kind, source_cfo, source_year) in enumerate(raw):
        later_headers = [
            other_row
            for other_sheet, other_row, *_ in raw
            if other_sheet.title == sheet.title and other_row > header_row
        ]
        upper_bound = min(later_headers) - 1 if later_headers else sheet.max_row
        if source_kind == "intalev_opiu":
            if not _has_intalev_business_structure(
                sheet, header_row + 1, upper_bound, columns["article"]
            ):
                continue
            first_data_row, last_data_row = _intalev_data_bounds(
                sheet, header_row + 1, upper_bound, columns["article"]
            )
        elif source_kind == "bdr_revenue_summary":
            first_data_row = header_row + 1
            last_data_row = header_row + len(BDR_REVENUE_ARTICLES)
        else:
            first_data_row = header_row + 1
            last_data_row = _last_data_row(sheet, first_data_row, upper_bound, columns)
        if last_data_row < first_data_row:
            continue
        candidates.append(
            CandidateRange(
                candidate_id=f"candidate-{index + 1}",
                sheet=sheet.title,
                header_row=header_row,
                first_data_row=first_data_row,
                last_data_row=last_data_row,
                columns=columns,
                source_kind=source_kind,
                source_cfo=source_cfo,
                source_year=source_year,
            )
        )
    return candidates


def _date_month(value: Any) -> tuple[int, int] | None:
    if isinstance(value, (date, datetime)):
        return value.year, value.month
    return None


def _bdr_revenue_schemas(
    rows: list[tuple[Any, ...]],
) -> list[tuple[int, dict[str, int], int]]:
    """Find only the source-proven BDR revenue block by its full structure.

    The summary has no semantic labels for its two hierarchy columns. A block
    is accepted only when a contiguous January–December header is followed by
    twelve exact ``план`` markers and the body contains the exact revenue group
    plus every one of the seven exact child articles, once each. This avoids
    sheet-name, contains, first-candidate and position-only inference.
    """

    result: list[tuple[int, dict[str, int], int]] = []
    month_schemas: list[tuple[dict[str, int], int, int]] = []
    for row_index, row in enumerate(rows[:-1]):
        next_row = rows[row_index + 1]
        for start in range(0, max(len(row) - 11, 0)):
            periods = [_date_month(row[start + offset]) for offset in range(12)]
            if any(period is None for period in periods):
                continue
            years = {period[0] for period in periods if period is not None}
            months = [period[1] for period in periods if period is not None]
            if len(years) != 1 or months != list(range(1, 13)):
                continue
            if not all(
                normalize_header(next_row[start + offset]) == "план"
                for offset in range(12)
            ):
                continue
            year = next(iter(years))
            month_schemas.append(
                (
                    {
                        f"month_{month}": start + month
                        for month in range(1, 13)
                    },
                    year,
                    row_index + 1,
                )
            )

    expected_articles = set(BDR_REVENUE_ARTICLES)
    child_count = len(BDR_REVENUE_ARTICLES)
    for parent_index, parent in enumerate(rows):
        children = rows[parent_index + 1 : parent_index + 1 + child_count]
        if len(children) != child_count:
            continue
        group_columns = [
            index
            for index, value in enumerate(parent)
            if str(value or "").strip() == BDR_REVENUE_GROUP
        ]
        for group_index in group_columns:
            if not all(
                str(row[group_index] or "").strip() == BDR_REVENUE_GROUP
                for row in children
            ):
                continue
            for article_index in group_columns:
                if article_index == group_index:
                    continue
                articles = [str(row[article_index] or "").strip() for row in children]
                if len(set(articles)) != child_count or set(articles) != expected_articles:
                    continue
                for month_columns, source_year, month_header_row in month_schemas:
                    if parent_index + 1 <= month_header_row + 1:
                        continue
                    if not all(
                        any(
                            isinstance(row[column - 1], (int, float))
                            and not isinstance(row[column - 1], bool)
                            for column in month_columns.values()
                        )
                        for row in children
                    ):
                        continue
                    result.append(
                        (
                            parent_index + 1,
                            {
                                "revenue_group": group_index + 1,
                                "article": article_index + 1,
                                "input_sales_channel": article_index + 1,
                                **month_columns,
                            },
                            source_year,
                        )
                    )
    return result


def _intalev_schema(values: Iterable[Any]) -> dict[str, int] | None:
    row = list(values)
    indicators = [
        index
        for index, value in enumerate(row, start=1)
        if normalize_header(value) in INTALEV_INDICATOR_ALIASES
    ]
    if len(indicators) != 1:
        return None

    months: dict[int, int] = {}
    years: set[int] = set()
    for index, value in enumerate(row, start=1):
        period = _parse_period(value)
        if period is None:
            continue
        year, month = period
        if month in months:
            return None
        months[month] = index
        years.add(year)
    if set(months) != set(range(1, 13)) or len(years) != 1:
        return None
    return {
        "article": indicators[0],
        **{f"month_{month}": months[month] for month in range(1, 13)},
    }


def _parse_period(value: Any) -> tuple[int, int] | None:
    if isinstance(value, datetime):
        return value.year, value.month
    if isinstance(value, date):
        return value.year, value.month
    match = INTALEV_PERIOD.match(str(value or ""))
    if not match:
        return None
    start_day, start_month, start_year, _, end_month, end_year = map(int, match.groups())
    if start_day != 1 or start_month != end_month or start_year != end_year:
        return None
    return start_year, start_month


def _period_year(values: Iterable[Any], columns: dict[str, int]) -> int | None:
    row = list(values)
    period = _parse_period(row[columns["month_1"] - 1])
    return period[0] if period else None


def _metadata_cfo(sheet: Any, header_row: int) -> str:
    for row in sheet.iter_rows(
        min_row=1,
        max_row=max(header_row - 1, 1),
        min_col=1,
        max_col=min(sheet.max_column or 1, 16),
        values_only=True,
    ):
        for value in row:
            match = INTALEV_METADATA_CFO.search(str(value or ""))
            if match:
                return match.group(1).strip()
    return ""


def _row_outline_level(sheet: Any, row_number: int) -> int:
    row_dimensions = getattr(sheet, "row_dimensions", None)
    if row_dimensions is None:
        return 0
    try:
        return int(row_dimensions[row_number].outlineLevel or 0)
    except (AttributeError, KeyError, TypeError):
        return 0


def _next_nonempty_label(
    sheet: Any, row_number: int, last_row: int, label_column: int
) -> tuple[int, str] | None:
    for candidate_row in range(row_number + 1, last_row + 1):
        value = str(sheet.cell(candidate_row, label_column).value or "").strip()
        if value:
            return candidate_row, value
    return None


def _is_flat_expense_type_row(
    sheet: Any, row_number: int, last_row: int, label_column: int
) -> bool:
    cell = sheet.cell(row_number, label_column)
    name = str(cell.value or "").strip()
    if not name or not bool(cell.font.bold) or not EXPENSE_WORD.search(name):
        return False
    next_label = _next_nonempty_label(sheet, row_number, min(last_row, row_number + 3), label_column)
    return bool(
        next_label
        and normalize_header(next_label[1]).startswith("% расходов")
    )


def _has_intalev_business_structure(
    sheet: Any, first_row: int, upper_bound: int | None, label_column: int
) -> bool:
    """Reject generic 12-month metric tables that only resemble Intalev OPIU.

    This scan must remain sequential: random ``sheet.cell`` access on an
    ``openpyxl`` read-only worksheet reparses the XML stream and turns a
    several-megabyte workbook into a multi-minute operation.
    """

    last_row = upper_bound if upper_bound is not None else sheet.max_row
    labelled: list[tuple[int, Any, str]] = []
    rows = sheet.iter_rows(
        min_row=first_row,
        max_row=last_row,
        min_col=label_column,
        max_col=label_column,
        values_only=False,
    )
    for row_number, row in enumerate(rows, start=first_row):
        cell = row[0]
        name = str(cell.value or "").strip()
        if name:
            labelled.append((row_number, cell, name))

    has_expense_label = False
    has_hierarchy = False
    for index, (row_number, cell, name) in enumerate(labelled):
        if EXPENSE_WORD.search(name):
            has_expense_label = True
        indent = int(cell.alignment.indent or 0)
        if indent > 0 or _row_outline_level(sheet, row_number) > 0:
            has_hierarchy = True

        if not bool(cell.font.bold) or not EXPENSE_WORD.search(name):
            continue
        # The real flat OPIU layout places ``% расходов`` immediately after
        # an expense-section header. Allow intervening blank rows but not a
        # distant, unrelated ratio line.
        if index + 1 < len(labelled):
            next_row, _, next_name = labelled[index + 1]
            if (
                next_row <= row_number + 3
                and normalize_header(next_name).startswith("% расходов")
            ):
                return True
    return has_expense_label and has_hierarchy


def _intalev_data_bounds(
    sheet: Any, first_row: int, upper_bound: int | None, label_column: int
) -> tuple[int, int]:
    last_row = upper_bound if upper_bound is not None else sheet.max_row
    first_found: int | None = None
    last_found: int | None = None
    rows = sheet.iter_rows(
        min_row=first_row,
        max_row=last_row,
        min_col=label_column,
        max_col=label_column,
        values_only=True,
    )
    for row_number, row in enumerate(rows, start=first_row):
        if normalize_header(row[0]):
            if first_found is None:
                first_found = row_number
            last_found = row_number
    if first_found is None or last_found is None:
        return last_row + 1, last_row
    return first_found, last_found


def _first_intalev_data_row(
    sheet: Any, first_row: int, upper_bound: int | None, label_column: int
) -> int:
    return _intalev_data_bounds(sheet, first_row, upper_bound, label_column)[0]


def _last_intalev_data_row(
    sheet: Any, first_row: int, upper_bound: int | None, label_column: int
) -> int:
    return _intalev_data_bounds(sheet, first_row, upper_bound, label_column)[1]


def _last_data_row(
    sheet: Any,
    first_row: int,
    upper_bound: int | None,
    columns: dict[str, int],
) -> int:
    relevant = list(columns.values())
    min_column = min(relevant)
    max_column = max(relevant)
    last = first_row - 1
    empty_streak = 0
    row_options = {
        "min_row": first_row,
        "min_col": min_column,
        "max_col": max_column,
        "values_only": True,
    }
    if upper_bound is not None:
        row_options["max_row"] = upper_bound
    rows = sheet.iter_rows(**row_options)
    for row_number, row in enumerate(rows, start=first_row):
        values = [row[column - min_column] for column in relevant]
        if all(value is None for value in values):
            empty_streak += 1
            if empty_streak >= 3:
                break
            continue
        empty_streak = 0
        last = row_number
    return last


def read_source_rows(
    workbook: Any,
    candidate: CandidateRange,
    source_file: str,
    bdr_cached_formula_values: dict[str, Any] | None = None,
) -> list[SourceRow]:
    if candidate.source_kind == "bdr_full":
        return _read_bdr_full_rows(
            workbook,
            candidate,
            source_file,
            bdr_cached_formula_values or {},
        )
    if candidate.source_kind == "intalev_opiu":
        return _read_intalev_source_rows(workbook, candidate, source_file)
    if candidate.source_kind == "prepared_income_budget":
        return _read_income_source_rows(workbook, candidate, source_file)
    if candidate.source_kind == "bdr_revenue_summary":
        return _read_bdr_revenue_rows(workbook, candidate, source_file)

    sheet = workbook[candidate.sheet]
    result: list[SourceRow] = []
    relevant = list(candidate.columns.values())
    min_column = min(relevant)
    max_column = max(relevant)
    rows = sheet.iter_rows(
        min_row=candidate.first_data_row,
        max_row=candidate.last_data_row,
        min_col=min_column,
        max_col=max_column,
        values_only=True,
    )
    for row_number, row in enumerate(rows, start=candidate.first_data_row):
        def value(field: str) -> Any:
            return row[candidate.columns[field] - min_column]

        month_values = tuple(
            value(f"month_{month}")
            for month in range(1, 13)
        )
        shared_values = {
            field: value(field)
            for field in BUSINESS_ALIASES
        }
        indicator_values = {
            field: value(field) if field in candidate.columns else ""
            for field in INDICATOR_ALIASES
        }
        if all(
            item is None or item == ""
            for item in (*shared_values.values(), *indicator_values.values(), *month_values)
        ):
            continue
        cells = {
            field: f"{get_column_letter(column)}{row_number}"
            for field, column in candidate.columns.items()
        }
        result.append(
            SourceRow(
                source_file=source_file,
                sheet=candidate.sheet,
                row_number=row_number,
                months=month_values,
                cells=cells,
                **shared_values,
                **indicator_values,
            )
        )
    return result


def _bdr_indicator_type(candidate: CandidateRange, row_number: int) -> IndicatorType:
    matches = [
        indicator_type
        for indicator_type, first_row, last_row in candidate.indicator_blocks
        if first_row <= row_number <= last_row
    ]
    if len(matches) != 1:
        raise ValueError("Строка БДР не относится к одному определённому блоку")
    return matches[0]


def _read_bdr_full_rows(
    workbook: Any,
    candidate: CandidateRange,
    source_file: str,
    cached_formula_values: dict[str, Any],
) -> list[SourceRow]:
    sheet = workbook[candidate.sheet]
    max_column = max((*candidate.columns.values(), *candidate.label_columns))
    rows = sheet.iter_rows(
        min_row=candidate.first_data_row,
        max_row=candidate.last_data_row,
        min_col=1,
        max_col=max_column,
        values_only=True,
    )
    result: list[SourceRow] = []
    for row_number, row in enumerate(rows, start=candidate.first_data_row):
        labelled = _bdr_row_label(row, candidate.label_columns)
        if labelled is None:
            continue
        indicator, indicator_column = labelled
        indicator_type = _bdr_indicator_type(candidate, row_number)
        month_cells = tuple(
            f"{get_column_letter(candidate.columns[f'month_{month}'])}{row_number}"
            for month in range(1, 13)
        )
        month_values = tuple(
            cached_formula_values.get(
                cell,
                row[candidate.columns[f"month_{month}"] - 1],
            )
            for month, cell in enumerate(month_cells, start=1)
        )
        if not any(value not in (None, "") for value in month_values):
            continue
        article_cell = f"{get_column_letter(indicator_column)}{row_number}"
        department_column = candidate.columns.get("department")
        kpi_department = (
            row[department_column - 1]
            if indicator_type == IndicatorType.KPI
            and department_column is not None
            and department_column <= len(row)
            else ""
        )
        department_cell = (
            f"{get_column_letter(department_column)}{row_number}"
            if kpi_department and department_column is not None
            else candidate.reporting_unit_cell or article_cell
        )
        cells = {
            "article": article_cell,
            "indicator_type": article_cell,
            "cfo": department_cell,
            "reporting_unit": candidate.reporting_unit_cell or article_cell,
            **({"department": department_cell} if kpi_department else {}),
            **{
                f"month_{month}": month_cells[month - 1]
                for month in range(1, 13)
            },
        }
        result.append(
            SourceRow(
                source_file=source_file,
                sheet=candidate.sheet,
                row_number=row_number,
                reporting_unit=candidate.source_cfo,
                expense_type="",
                department=kpi_department,
                organization_type="",
                cfo=kpi_department or candidate.source_cfo,
                tax=None,
                expense_group="",
                article=indicator,
                months=month_values,
                cells=cells,
                indicator_type=indicator_type.value,
                source_kind="bdr_full",
            )
        )
    return result


def _read_income_source_rows(
    workbook: Any,
    candidate: CandidateRange,
    source_file: str,
) -> list[SourceRow]:
    sheet = workbook[candidate.sheet]
    result: list[SourceRow] = []
    relevant = list(candidate.columns.values())
    min_column = min(relevant)
    max_column = max(relevant)
    rows = sheet.iter_rows(
        min_row=candidate.first_data_row,
        max_row=candidate.last_data_row,
        min_col=min_column,
        max_col=max_column,
        values_only=True,
    )
    for row_number, row in enumerate(rows, start=candidate.first_data_row):
        def value(field: str) -> Any:
            return row[candidate.columns[field] - min_column]

        month_values = tuple(value(f"month_{month}") for month in range(1, 13))
        revenue_type = value("revenue_type")
        revenue_group = value("revenue_group")
        article = value("article")
        analytics = value("analytics")
        indicator_values = {
            field: value(field) if field in candidate.columns else ""
            for field in INDICATOR_ALIASES
        }
        indicator_values["indicator_type"] = "REVENUE"
        indicator_values["revenue_group"] = revenue_group
        indicator_values["analytics"] = analytics
        if all(
            item is None or item == ""
            for item in (
                revenue_type,
                revenue_group,
                article,
                analytics,
                *indicator_values.values(),
                *month_values,
            )
        ):
            continue

        cells = {
            field: f"{get_column_letter(column)}{row_number}"
            for field, column in candidate.columns.items()
        }
        cells["expense_type"] = cells["revenue_type"]
        cells["expense_group"] = cells["revenue_group"]
        result.append(
            SourceRow(
                source_file=source_file,
                sheet=candidate.sheet,
                row_number=row_number,
                reporting_unit=None,
                expense_type=revenue_type,
                department=None,
                organization_type=None,
                cfo=None,
                tax=None,
                expense_group=revenue_group,
                article=article,
                months=month_values,
                cells=cells,
                **indicator_values,
            )
        )
    return result


def _read_bdr_revenue_rows(
    workbook: Any,
    candidate: CandidateRange,
    source_file: str,
) -> list[SourceRow]:
    sheet = workbook[candidate.sheet]
    result: list[SourceRow] = []
    relevant = list(candidate.columns.values())
    min_column = min(relevant)
    max_column = max(relevant)
    rows = sheet.iter_rows(
        min_row=candidate.first_data_row,
        max_row=candidate.last_data_row,
        min_col=min_column,
        max_col=max_column,
        values_only=True,
    )
    for row_number, row in enumerate(rows, start=candidate.first_data_row):
        def value(field: str) -> Any:
            return row[candidate.columns[field] - min_column]

        revenue_group = value("revenue_group")
        article = value("article")
        input_sales_channel = value("input_sales_channel")
        month_values = tuple(value(f"month_{month}") for month in range(1, 13))
        cells = {
            field: f"{get_column_letter(column)}{row_number}"
            for field, column in candidate.columns.items()
        }
        cells["expense_group"] = cells["revenue_group"]
        result.append(
            SourceRow(
                source_file=source_file,
                sheet=candidate.sheet,
                row_number=row_number,
                reporting_unit=None,
                expense_type=None,
                department=None,
                organization_type=None,
                cfo=None,
                tax=None,
                expense_group=revenue_group,
                article=article,
                months=month_values,
                cells=cells,
                indicator_type="REVENUE",
                revenue_group=revenue_group,
                input_sales_channel=input_sales_channel,
            )
        )
    return result


def _intalev_month_values(sheet: Any, candidate: CandidateRange, row_number: int) -> tuple[Any, ...]:
    return tuple(
        0
        if sheet.cell(row_number, candidate.columns[f"month_{month}"]).value in (None, "")
        else sheet.cell(row_number, candidate.columns[f"month_{month}"]).value
        for month in range(1, 13)
    )


def _intalev_cells(
    candidate: CandidateRange,
    row_number: int,
    *,
    reporting_unit_column: int | None = None,
) -> dict[str, str]:
    label_column = candidate.columns["article"]
    article_cell = f"{get_column_letter(label_column)}{row_number}"
    reporting_cell = (
        f"{get_column_letter(reporting_unit_column)}{row_number}"
        if reporting_unit_column
        else "A2"
    )
    cfo_cell = "A2" if candidate.source_cfo else article_cell
    return {
        "reporting_unit": reporting_cell,
        "expense_type": article_cell,
        "department": article_cell,
        "organization_type": article_cell,
        "cfo": cfo_cell,
        "tax": article_cell,
        "expense_group": article_cell,
        "article": article_cell,
        **{
            f"month_{month}": (
                f"{get_column_letter(candidate.columns[f'month_{month}'])}{row_number}"
            )
            for month in range(1, 13)
        },
    }


def _read_hierarchical_intalev_rows(
    sheet: Any, candidate: CandidateRange, source_file: str
) -> list[SourceRow]:
    label_column = candidate.columns["article"]
    labelled: list[tuple[int, str, int]] = []
    for row_number in range(candidate.first_data_row, candidate.last_data_row + 1):
        cell = sheet.cell(row_number, label_column)
        name = str(cell.value or "").strip()
        if not name:
            continue
        indent = int(cell.alignment.indent or 0)
        outline = _row_outline_level(sheet, row_number)
        level = indent if indent else outline * 2
        labelled.append((row_number, name, level))

    stack: dict[int, str] = {}
    result: list[SourceRow] = []
    for index, (row_number, name, level) in enumerate(labelled):
        stack[level] = name
        for stale_level in [item for item in stack if item > level]:
            del stack[stale_level]

        next_level = labelled[index + 1][2] if index + 1 < len(labelled) else -1
        is_leaf = next_level <= level
        if not is_leaf or level == 0 or TECHNICAL_TOTAL.search(name):
            continue

        business_ancestors = [
            stack[item]
            for item in sorted(stack)
            if 0 < item < level and not TECHNICAL_TOTAL.search(stack[item])
        ]
        expense_type = business_ancestors[0] if business_ancestors else ""
        expense_group = business_ancestors[-1] if business_ancestors else ""
        if normalize_header(expense_group) == "<пустое значение>":
            expense_group = ""

        result.append(
            SourceRow(
                source_file=source_file,
                sheet=candidate.sheet,
                row_number=row_number,
                reporting_unit=None,
                expense_type=expense_type,
                department=None,
                organization_type=None,
                cfo=candidate.source_cfo,
                tax=None,
                expense_group=expense_group,
                article=name,
                months=_intalev_month_values(sheet, candidate, row_number),
                cells=_intalev_cells(candidate, row_number),
            )
        )
    return result


def _flat_section_rows(
    sheet: Any,
    candidate: CandidateRange,
) -> list[tuple[int, int, str]]:
    label_column = candidate.columns["article"]
    type_rows: list[tuple[int, str]] = []
    for row_number in range(candidate.first_data_row, candidate.last_data_row + 1):
        if _is_flat_expense_type_row(
            sheet, row_number, candidate.last_data_row, label_column
        ):
            type_rows.append((row_number, str(sheet.cell(row_number, label_column).value).strip()))
    result: list[tuple[int, int, str]] = []
    for index, (row_number, name) in enumerate(type_rows):
        end_row = (
            type_rows[index + 1][0] - 1
            if index + 1 < len(type_rows)
            else candidate.last_data_row
        )
        result.append((row_number, end_row, name))
    return result


def _read_flat_intalev_rows(
    sheet: Any, candidate: CandidateRange, source_file: str
) -> list[SourceRow]:
    """Read the real flat OPIU layout that encodes hierarchy by formatting.

    The source repeats the expense type in the column immediately to the left
    of ``Показатели``.  Bold rows are business groups; italic/regular rows are
    their leaves.  A bold row without following leaves is retained as a
    standalone business article so values are never silently lost.
    """

    label_column = candidate.columns["article"]
    parent_column = label_column - 1 if label_column > 1 else None
    reporting_unit_column = label_column - 2 if label_column > 2 else None
    result: list[SourceRow] = []

    for type_row, section_end, expense_type in _flat_section_rows(sheet, candidate):
        type_norm = normalize_header(expense_type)
        content: list[tuple[int, str, bool]] = []
        for row_number in range(type_row + 1, section_end + 1):
            cell = sheet.cell(row_number, label_column)
            name = str(cell.value or "").strip()
            if not name:
                continue
            parent_hint = (
                normalize_header(sheet.cell(row_number, parent_column).value)
                if parent_column
                else type_norm
            )
            # Rows after the final expense section (EBITDA, profit, ratios)
            # do not repeat the current expense type and are outside the
            # business hierarchy even though they share the same table.
            if parent_column and parent_hint != type_norm:
                continue
            normalized = normalize_header(name)
            if (
                normalized.startswith("% расходов")
                or TECHNICAL_TOTAL.search(name)
                or TECHNICAL_RATIO.search(normalized)
                or normalized == type_norm
            ):
                continue
            content.append((row_number, name, bool(cell.font.bold)))

        current_group = ""
        for index, (row_number, name, is_bold) in enumerate(content):
            if is_bold:
                has_leaf_children = False
                for _, _, next_is_bold in content[index + 1 :]:
                    if next_is_bold:
                        break
                    has_leaf_children = True
                    break
                if has_leaf_children:
                    current_group = name
                    continue
                expense_group = name
                article = name
            else:
                expense_group = current_group or name
                article = name

            reporting_unit = None
            if reporting_unit_column:
                reporting_unit = sheet.cell(row_number, reporting_unit_column).value
            result.append(
                SourceRow(
                    source_file=source_file,
                    sheet=candidate.sheet,
                    row_number=row_number,
                    reporting_unit=reporting_unit,
                    expense_type=expense_type,
                    department=None,
                    organization_type=None,
                    cfo=candidate.source_cfo,
                    tax=None,
                    expense_group=expense_group,
                    article=article,
                    months=_intalev_month_values(sheet, candidate, row_number),
                    cells=_intalev_cells(
                        candidate,
                        row_number,
                        reporting_unit_column=reporting_unit_column,
                    ),
                )
            )
    return result


def _read_intalev_source_rows(
    workbook: Any, candidate: CandidateRange, source_file: str
) -> list[SourceRow]:
    sheet = workbook[candidate.sheet]
    hierarchical = _read_hierarchical_intalev_rows(sheet, candidate, source_file)
    if hierarchical:
        return hierarchical
    return _read_flat_intalev_rows(sheet, candidate, source_file)
