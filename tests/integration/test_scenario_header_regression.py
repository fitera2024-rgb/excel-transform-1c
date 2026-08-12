from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook

from excel_transform_1c.adapters.references import parse_reference_workbook


pytestmark = pytest.mark.integration


def _scenario_export_with_header_word_in_data() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист_1"
    sheet.cell(7, 1, "Сценарии")
    sheet.cell(7, 19, "Код")

    rows = [
        ("Исполнение бюджета", "00001"),
        ("Резерв", "00002"),
        ("Факт", "00001"),
        ("Фактические данные", "00002"),
        ("ПЛАН 2026", "00010"),
        ("ПЛАН 2027", "00011"),
        ("Прогноз", "00003"),
        ("Бюджет", "00004"),
        ("Сценарий отчетности КИК", "00005"),
        ("Корректировка", "00006"),
        ("Оперативный план", "00007"),
        ("Стратегический план", "00008"),
    ]
    for row_number, (name, code) in enumerate(rows, start=8):
        sheet.cell(row_number, 1, name)
        sheet.cell(row_number, 19, code)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_scenario_name_containing_header_word_is_not_dropped() -> None:
    rows = parse_reference_workbook(
        _scenario_export_with_header_word_in_data(),
        "scenarios",
    )

    assert len(rows) == 12
    assert any(row["name"] == "Сценарий отчетности КИК" for row in rows)
