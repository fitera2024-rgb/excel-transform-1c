from io import BytesIO

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment

from excel_transform_1c.adapters.opiu_sources import build_catalog_from_source_bytes
from excel_transform_1c.application.service import WorkflowService
from excel_transform_1c.core.opiu_rules.opiu_rule_models import AUTO_MATCH
from tests.helpers.workbooks import reference_bytes


pytestmark = pytest.mark.integration


MXL_HEADERS = (
    "Код упрощенной формулы",
    "Наименование",
    "Описание",
    "Объект учета",
    "Регистр ИБ",
    "Счет",
    "Назначение расчетов",
    "Корр счет",
    "Потребитель расчета",
    "Способ получения",
    "Способ использования",
    "Код",
    "Не используется",
)


def _xlsx(build) -> bytes:
    workbook = Workbook()
    build(workbook.active)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def formula_bytes() -> bytes:
    def build(sheet):
        sheet.append(["Строка", "Сумма"])
        sheet.append(["Расходы по основной деятельности", "[TOTAL]"])
        for row, name, token in (
            (3, "Административные", "SRC_ADMIN"),
            (4, "Коммерческие", "SRC_COMMERCIAL"),
        ):
            sheet.cell(row, 1, name).alignment = Alignment(indent=2)
            sheet.cell(row, 2, f"[{token}]")

    return _xlsx(build)


def analytic_bytes() -> bytes:
    def build(sheet):
        sheet.append(["Строка", "Аналитика 1", "Аналитика 2"])
        sheet.append(["Расходы по основной деятельности", "Организационные единицы"])
        sheet.append(["Административные", "Организационные единицы", "ЦФО (казначейство)"])
        sheet.append(["Коммерческие", "Организационные единицы", "ЦФО (казначейство)"])

    return _xlsx(build)


def indicator_bytes() -> bytes:
    def build(sheet):
        headers = {
            1: "Показатели отчетов",
            9: "Группа аналитик",
            11: "Колонка",
            20: "Строка",
            39: "Нормализованный код",
            50: "Код",
        }
        for column, value in headers.items():
            sheet.cell(8, column, value)
        for row, line, code in (
            (9, "Административные", "IND-ADMIN"),
            (10, "Коммерческие", "IND-COMMERCIAL"),
        ):
            sheet.cell(row, 1, f"{line} сумма")
            sheet.cell(row, 9, "ОргЕдЦФО")
            sheet.cell(row, 11, "Сумма")
            sheet.cell(row, 20, line)
            sheet.cell(row, 39, code)
            sheet.cell(row, 50, code)

    return _xlsx(build)


def code_name_bytes(rows: list[tuple[str, str]]) -> bytes:
    def build(sheet):
        sheet.append(["Наименование", "Код"])
        for name, code in rows:
            sheet.append([name, code])

    return _xlsx(build)


def _mxl_cell(value: str, tail: str) -> str:
    escaped = value.replace('"', '""')
    return f'{{16,2,\n{{1,1,\n{{"#","{escaped}"}}\n}},0}},{tail}\n'


def mxl_bytes() -> bytes:
    rows = [
        MXL_HEADERS,
        (
            "SRC_ADMIN",
            "Источник административных",
            "Ист: С1 В ИЕРАРХИИ(Административные)\r\nПр: С1=Статья",
            "",
            "Журнал проводок МСФО",
            "26",
            "Отчет о прибылях и убытках",
            "",
            "Административные сумма",
            "Внутренние данные (регистр бухгалтерии)",
            "Для формул расчета",
            "SRC_ADMIN",
            "Нет",
        ),
        (
            "SRC_COMMERCIAL",
            "Источник коммерческих",
            "Ист: С1 В ИЕРАРХИИ(Коммерческие)\r\nПр: С1=Статья",
            "",
            "Журнал проводок МСФО",
            "44",
            "Отчет о прибылях и убытках",
            "",
            "Коммерческие сумма",
            "Внутренние данные (регистр бухгалтерии)",
            "Для формул расчета",
            "SRC_COMMERCIAL",
            "Нет",
        ),
    ]
    parts = []
    width = len(MXL_HEADERS)
    for row_number, row in enumerate(rows, start=1):
        for column, value in enumerate(row, start=1):
            tail = (
                f"{row_number},0,{width},0,"
                if column == width
                else f"{column},"
            )
            parts.append(_mxl_cell(value, tail))
    return b"MOXCEL\x00\x08\x00\x01\x00\x0c\x00\xef\xbb\xbf" + "".join(parts).encode("utf-8")


def source_bundle() -> dict[str, bytes]:
    return {
        "formulas_xlsx": formula_bytes(),
        "analytics_xlsx": analytic_bytes(),
        "indicators_xlsx": indicator_bytes(),
        "sources_mxl": mxl_bytes(),
        "regions_xlsx": code_name_bytes([("ДВ", "REG-1")]),
        "networks_xlsx": code_name_bytes(
            [("Административные", "NET-1"), ("Коммерческие", "NET-2")]
        ),
    }


def test_four_source_join_builds_complete_internal_registries():
    catalog = build_catalog_from_source_bytes(**source_bundle())

    assert len(catalog.formula_rules) == 3
    assert len(catalog.analytic_rules) == 3
    assert len(catalog.indicator_catalog) == 2
    assert len(catalog.source_rules) == 2
    assert len(catalog.region_catalog) == 1
    assert len(catalog.network_catalog) == 2
    assert len(catalog.rules) == 2
    assert {rule.disclosure_group for rule in catalog.rules} == {
        "Административные",
        "Коммерческие",
    }
    assert all(rule.report_indicator.endswith(" сумма") for rule in catalog.rules)
    assert len(catalog.unresolved) == 2


def test_service_persists_catalog_and_expands_exact_articles(tmp_path):
    service = WorkflowService(tmp_path / "runtime")
    service.upload_reference("erp_articles", reference_bytes("erp_articles"))

    catalog = service.upload_opiu_rule_sources(**source_bundle())
    restarted = WorkflowService(tmp_path / "runtime")
    rules = restarted.opiu_rules()

    assert len(catalog.rules) == 2
    assert {(rule.disclosure_group, rule.article) for rule in rules} == {
        ("Административные", "Интернет"),
        ("Административные", "Удалить"),
        ("Коммерческие", "Реклама"),
    }
    result = restarted.opiu_rules()
    assert result[0].rule_id.startswith("opiu-")

    from excel_transform_1c.core.opiu_rules.opiu_indicator_resolver import (
        OPIUIndicatorResolver,
    )

    match = OPIUIndicatorResolver(rules).resolve(
        disclosure_group="Административные",
        article="Интернет",
        organization="Организация",
        cfo="ЦФО",
    )
    assert match.status == AUTO_MATCH
    assert match.rule and match.rule.indicator_code == "IND-ADMIN"
