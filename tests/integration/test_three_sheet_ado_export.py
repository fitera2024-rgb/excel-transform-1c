from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from excel_transform_1c.adapters.excel import (
    ADO_INDICATOR_HEADERS,
    ADO_OPIU_HEADERS,
    EXPORT_HEADERS,
    export_opiu_light,
)
from excel_transform_1c.core.models import PreviewRecord


pytestmark = pytest.mark.integration


def _record() -> PreviewRecord:
    return PreviewRecord(
        record_id="record-1",
        source_row=7,
        month=1,
        year=2026,
        reporting_unit="АЮ",
        organization="4 Владивосток → 4 Владивосток (000000041)",
        scenario="ПЛАН 2026",
        department="Департамент обеспечения",
        organization_type="ТК",
        cfo="АЮ Отдел обеспечения",
        expense_type="Административные расходы",
        expense_group="Хозяйственные расходы",
        source_article="Перчатки",
        erp_code="00-000169",
        erp_article_name="Перчатки",
        tax="БЕЗ НДС",
        amount=Decimal("123.45"),
        erp_department="АЮ Отдел обеспечения",
        cfo_code="000000175",
        organization_unit="4 Владивосток",
        organization_unit_code="000000041",
    )


def _indicator_record() -> PreviewRecord:
    record = _record()
    record.indicator_match_status = "matched"
    record.indicator = "Административные расходы"
    record.sales_channel = "Основной канал"
    return record


def test_export_keeps_legacy_sheet_and_adds_two_ado_sheets():
    payload = export_opiu_light([_record()])
    workbook = load_workbook(BytesIO(payload), data_only=True)
    try:
        assert workbook.sheetnames == ["OPIU Light", "ОПИУ", "Показатели"]

        legacy = workbook["OPIU Light"]
        assert tuple(cell.value for cell in legacy[1]) == EXPORT_HEADERS
        assert legacy.max_row == 2
        assert legacy["P2"].value == "00-000169"
        assert legacy["S2"].value == 123.45

        ado = workbook["ОПИУ"]
        assert tuple(cell.value for cell in ado[1]) == ADO_OPIU_HEADERS
        assert ado.max_row == 2
        assert tuple(cell.value for cell in ado[2]) == (
            "4 Владивосток",
            "000000041",
            "ПЛАН 2026",
            2026,
            1,
            "01.2026",
            "Департамент обеспечения",
            "АЮ Отдел обеспечения",
            "АЮ Отдел обеспечения",
            "000000175",
            "Административные расходы",
            "00-000169",
            "Перчатки",
            None,
            None,
            None,
            None,
            123.45,
        )
        assert ado["L2"].data_type == "s"
        assert ado["L2"].number_format == "@"
        assert isinstance(ado["R2"].value, (int, float))

        indicators = workbook["Показатели"]
        assert tuple(cell.value for cell in indicators[1]) == ADO_INDICATOR_HEADERS
        assert indicators.max_row == 1
    finally:
        workbook.close()


def test_ado_opiu_keeps_rows_when_reference_codes_are_missing():
    record = _record()
    record.erp_code = ""
    record.erp_article_name = ""

    payload = export_opiu_light([record])
    workbook = load_workbook(BytesIO(payload), data_only=True)
    try:
        ado = workbook["ОПИУ"]
        assert ado.max_row == 2
        assert ado["J2"].value == "000000175"
        assert ado["L2"].value is None
        assert ado["M2"].value == "Перчатки"
        assert ado["R2"].value == 123.45
    finally:
        workbook.close()


def test_export_organization_name_without_code() -> None:
    payload = export_opiu_light([_indicator_record()])
    workbook = load_workbook(BytesIO(payload), data_only=True)
    try:
        for sheet_name in ("OPIU Light", "ОПИУ", "Показатели"):
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in sheet[1]]
            organization_column = headers.index("Организация") + 1
            assert sheet.cell(2, organization_column).value == "4 Владивосток"
    finally:
        workbook.close()


def test_export_cfo_code_separate() -> None:
    payload = export_opiu_light([_record()])
    workbook = load_workbook(BytesIO(payload), data_only=True)
    try:
        for sheet_name in ("OPIU Light", "ОПИУ"):
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in sheet[1]]
            cfo_column = headers.index("ЦФО") + 1
            cfo_code_column = headers.index("Код ЦФО") + 1
            assert sheet.cell(2, cfo_column).value == "АЮ Отдел обеспечения"
            cfo_code = sheet.cell(2, cfo_code_column)
            assert cfo_code.value == "000000175"
            assert cfo_code.data_type == "s"
            assert cfo_code.number_format == "@"
    finally:
        workbook.close()


def test_export_root_organization_code() -> None:
    record = _indicator_record()
    record.organization = "Контекст → Не головная организация (999999999)"
    payload = export_opiu_light([record])
    workbook = load_workbook(BytesIO(payload), data_only=True)
    try:
        for sheet_name in ("OPIU Light", "ОПИУ", "Показатели"):
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in sheet[1]]
            code_column = headers.index("Код организации") + 1
            code = sheet.cell(2, code_column)
            assert code.value == "000000041"
            assert code.data_type == "s"
            assert code.number_format == "@"
    finally:
        workbook.close()
