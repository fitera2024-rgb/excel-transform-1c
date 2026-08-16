from io import BytesIO

import pytest
from openpyxl import load_workbook

from excel_transform_1c.adapters.excel import EXPORT_HEADERS, detect_path, read_path
from excel_transform_1c.adapters.persistence import LocalStore
from excel_transform_1c.application.service import WorkflowService
from excel_transform_1c.core.models import STATUS_ATTENTION, STATUS_SKIPPED
from excel_transform_1c.core.transform import manual_mapping_key
from tests.helpers.workbooks import (
    real_reference_bytes,
    reference_bytes,
    workbook_bytes,
    write_cached_formula_fixture,
)


pytestmark = pytest.mark.integration


def configured_service(tmp_path) -> WorkflowService:
    service = WorkflowService(tmp_path / "runtime")
    for kind in ("erp_articles", "organizations", "scenarios", "intalev_cfos"):
        service.upload_reference(kind, reference_bytes(kind))
    service.store.save_cfo_mappings(
        {"code:INT-CFO-1": "cfo", "code:INT-CFO-2": "other"}
    )
    return service


def default_context(service: WorkflowService, months=None, reporting_unit="ПС"):
    scenario = service.store.list_scenarios()[0]
    return service.build_context(
        reporting_unit,
        "ps",
        scenario.scenario_id,
        2026,
        months or [],
    )


def test_known_real_erp_reference_exports_are_loaded_directly(tmp_path):
    service = WorkflowService(tmp_path / "runtime")
    assert service.upload_reference("erp_articles", real_reference_bytes("erp_articles")) == 3
    assert service.upload_reference("organizations", real_reference_bytes("organizations")) == 5
    assert service.upload_reference("scenarios", real_reference_bytes("scenarios")) == 2

    articles = {item.code: item.path for item in service.erp_articles()}
    assert articles["ERP-001"] == ("Административные", "Связь", "Интернет")
    assert articles["ERP-002"] == ("Коммерческие", "Маркетинг", "Реклама")
    organizations = service.organization_nodes()
    assert next(item for item in organizations if item.code == "ORG-3").parent_id == "ORG-2"
    assert any("!!!Удалить" in item.full_path for item in organizations)
    scenarios = service.store.list_scenarios()
    assert any(item.name == "ПЛАН 2026" and item.year == 2026 for item in scenarios)
    assert any(item.name == "Факт" and item.year == 0 for item in scenarios)


def test_arbitrary_sheet_name_and_24_record_workflow(tmp_path):
    service = configured_service(tmp_path)
    pending = service.prepare_upload(
        "synthetic.xlsx",
        workbook_bytes(sheet_name="Совсем другое имя"),
        default_context(service),
    )
    run = service.process_upload(pending.upload_id, pending.candidates[0].candidate_id)
    assert run.candidate.sheet == "Совсем другое имя"
    assert len(run.records) == 24


def test_two_candidates_require_explicit_selection_and_single_flight(tmp_path):
    service = configured_service(tmp_path)
    pending = service.prepare_upload(
        "synthetic.xlsx",
        workbook_bytes(two_candidates=True),
        default_context(service),
    )
    assert len(pending.candidates) == 2
    first = service.process_upload(pending.upload_id, pending.candidates[1].candidate_id)
    second = service.process_upload(pending.upload_id, pending.candidates[1].candidate_id)
    assert first.run_id == second.run_id
    assert len(service.runs) == 1


def test_broken_and_no_range_workbooks_are_blocked(tmp_path):
    service = configured_service(tmp_path)
    with pytest.raises(ValueError, match="повреждён"):
        service.prepare_upload("bad.xlsx", b"not an xlsx", default_context(service))
    pending = service.prepare_upload(
        "empty.xlsx",
        workbook_bytes(no_range=True),
        default_context(service),
    )
    assert pending.candidates == []


def test_cached_formula_value_is_read_without_excel_recalculation(tmp_path):
    path = tmp_path / "cached.xlsx"
    write_cached_formula_fixture(path)
    candidate = detect_path(path)[0]
    rows = read_path(path, candidate, "cached.xlsx")
    assert rows[0].months[0] == 100


def test_local_persistence_survives_restart_and_keeps_mapping(tmp_path):
    database = tmp_path / "local.db"
    first = LocalStore(database)
    scenario = first.add_scenario("ПЛАН 2027", 2027, "synthetic")
    key = ("ОтчетОПрибыляхИУбытках", "Тип", "Группа", "Статья")
    first.save_manual_mapping(key, "ERP-001")
    second = LocalStore(database)
    restarted = next(
        item for item in second.list_scenarios() if item.name == "ПЛАН 2027"
    )
    assert restarted.scenario_id == scenario.scenario_id
    assert second.load_manual_mappings()[key] == "ERP-001"


def test_month_filter_does_not_destroy_full_result_and_export_schema_is_business_only(tmp_path):
    service = configured_service(tmp_path)
    pending = service.prepare_upload(
        "synthetic.xlsx",
        workbook_bytes(),
        default_context(service, [1, 2]),
    )
    run = service.process_upload(pending.upload_id, pending.candidates[0].candidate_id)
    assert len(run.records) == 24
    assert len(run.visible_records()) == 4
    payload = service.export_run(run.run_id)
    workbook = load_workbook(BytesIO(payload), data_only=True)
    sheet = workbook["OPIU Light"]
    assert tuple(cell.value for cell in sheet[1]) == EXPORT_HEADERS
    assert sheet.max_row == 5
    forbidden = {"SHA", "BatchID", "Proof JSON", "Internal path", "RUN-ID"}
    assert not forbidden.intersection(EXPORT_HEADERS)


def test_skipped_month_stays_in_preview_and_export_with_pointer(tmp_path):
    service = configured_service(tmp_path)
    pending = service.prepare_upload(
        "synthetic.xlsx",
        workbook_bytes(monthly_error=True),
        default_context(service),
    )
    run = service.process_upload(pending.upload_id, pending.candidates[0].candidate_id)
    assert len(run.records) == 24
    skipped = next(
        record
        for record in run.records
        if record.source_row == 3 and record.month == 5
    )
    assert skipped.status == STATUS_SKIPPED
    assert skipped.amount is None
    assert "M3" in skipped.comment

    payload = service.export_run(run.run_id)
    workbook = load_workbook(BytesIO(payload), data_only=True)
    sheet = workbook["OPIU Light"]
    exported = [
        tuple(cell.value for cell in row)
        for row in sheet.iter_rows(min_row=2)
        if row[21].value == 3 and row[5].value == 5
    ]
    assert len(exported) == 1
    assert exported[0][18] is None
    assert exported[0][19] == STATUS_SKIPPED
    assert "M3" in exported[0][20]


def test_field_specific_correction_keeps_other_missing_field_attention(tmp_path):
    service = configured_service(tmp_path)
    pending = service.prepare_upload(
        "synthetic.xlsx",
        workbook_bytes(department_error=True, cfo_error=True),
        default_context(service),
    )
    run = service.process_upload(pending.upload_id, pending.candidates[0].candidate_id)
    before = [
        issue
        for issue in run.unresolved_issues
        if issue.pointer.row == 3 and issue.kind == "shared-field"
    ]
    assert {item.pointer.field for item in before} >= {"department", "cfo"}

    service.correct_row(
        run.run_id,
        3,
        {"department": "Группа → ПС → ЦФО 1"},
    )
    after = [
        issue
        for issue in run.unresolved_issues
        if issue.pointer.row == 3 and issue.kind == "shared-field"
    ]
    assert "department" not in {item.pointer.field for item in after}
    assert "cfo" in {item.pointer.field for item in after}
    assert all(
        record.status == STATUS_ATTENTION
        for record in run.records
        if record.source_row == 3
    )


def test_path_change_recalculates_and_invalidates_old_erp_code(tmp_path):
    service = configured_service(tmp_path)
    pending = service.prepare_upload(
        "synthetic.xlsx",
        workbook_bytes(),
        default_context(service),
    )
    run = service.process_upload(pending.upload_id, pending.candidates[0].candidate_id)
    assert all(
        record.erp_code == "ERP-001"
        for record in run.records
        if record.source_row == 3
    )

    service.correct_row(run.run_id, 3, {"expense_group": "Маркетинг"})
    affected = [record for record in run.records if record.source_row == 3]
    assert all(record.erp_code == "" for record in affected)
    assert any(
        issue.kind == "erp-mapping" and issue.pointer.row == 3
        for issue in run.unresolved_issues
    )


def test_path_and_erp_change_saves_mapping_for_new_path(tmp_path):
    service = configured_service(tmp_path)
    pending = service.prepare_upload(
        "synthetic.xlsx",
        workbook_bytes(),
        default_context(service),
    )
    run = service.process_upload(pending.upload_id, pending.candidates[0].candidate_id)

    service.correct_row(
        run.run_id,
        3,
        {
            "expense_group": "Маркетинг",
            "source_article": "Реклама",
            "erp_code": "ERP-002",
        },
    )
    new_key = manual_mapping_key("Административные", "Маркетинг", "Реклама")
    assert service.store.load_manual_mappings()[new_key] == "ERP-002"
    assert all(
        record.erp_code == "ERP-002"
        for record in run.records
        if record.source_row == 3
    )


def test_manual_mapping_conflict_with_exact_match_stays_visible(tmp_path):
    service = configured_service(tmp_path)
    conflict_key = manual_mapping_key("Административные", "Связь", "Интернет")
    service.store.save_manual_mapping(conflict_key, "ERP-002")
    pending = service.prepare_upload(
        "synthetic.xlsx",
        workbook_bytes(),
        default_context(service),
    )
    run = service.process_upload(pending.upload_id, pending.candidates[0].candidate_id)
    affected = [record for record in run.records if record.source_row == 3]
    assert all(record.erp_code == "ERP-002" for record in affected)
    assert all(record.status == STATUS_ATTENTION for record in affected)
    assert any(
        "конфликтует" in issue.description
        for issue in run.unresolved_issues
        if issue.pointer.row == 3 and issue.kind == "erp-mapping"
    )


def test_reporting_unit_conflict_continues_with_attention(tmp_path):
    service = configured_service(tmp_path)
    pending = service.prepare_upload(
        "synthetic.xlsx",
        workbook_bytes(reporting_unit="ПС"),
        default_context(service, reporting_unit="АЮ"),
    )
    run = service.process_upload(pending.upload_id, pending.candidates[0].candidate_id)
    assert len(run.records) == 24
    assert all(record.status == STATUS_ATTENTION for record in run.records)
    issue = next(
        issue
        for issue in run.unresolved_issues
        if issue.kind == "context-reporting-unit"
    )
    assert issue.pointer.cell == "A3"


def test_user_correction_updates_preview_and_registry_without_rerun(tmp_path):
    service = configured_service(tmp_path)
    pending = service.prepare_upload(
        "synthetic.xlsx",
        workbook_bytes(missing_mapping=True),
        default_context(service),
    )
    run = service.process_upload(pending.upload_id, pending.candidates[0].candidate_id)
    before_count = len(run.unresolved_issues)
    before_id = run.run_id
    service.correct_row(run.run_id, 3, {"erp_code": "ERP-001"})
    assert run.run_id == before_id
    assert run.rerun_count == 0
    assert all(
        record.erp_code == "ERP-001"
        for record in run.records
        if record.source_row == 3
    )
    assert len(run.unresolved_issues) < before_count
    assert service.store.load_manual_mappings()


def test_manual_mapping_reuses_only_same_accepted_key(tmp_path):
    service = configured_service(tmp_path)
    pending = service.prepare_upload(
        "first.xlsx",
        workbook_bytes(missing_mapping=True),
        default_context(service),
    )
    first = service.process_upload(pending.upload_id, pending.candidates[0].candidate_id)
    service.correct_row(first.run_id, 3, {"erp_code": "ERP-001"})

    pending_two = service.prepare_upload(
        "second.xlsx",
        workbook_bytes(
            sheet_name="Другой synthetic лист",
            missing_mapping=True,
        ),
        default_context(service),
    )
    second = service.process_upload(
        pending_two.upload_id,
        pending_two.candidates[0].candidate_id,
    )
    assert all(
        record.erp_code == "ERP-001"
        for record in second.records
        if record.source_row == 3
    )
    assert not any(
        issue.kind == "erp-mapping" and issue.pointer.row == 3
        for issue in second.unresolved_issues
    )
