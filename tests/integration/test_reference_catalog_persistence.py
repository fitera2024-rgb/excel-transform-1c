from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.styles import Alignment

from excel_transform_1c.application.service import WorkflowService


def _bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _split_header_reference(kind: str, *, supplement: bool = False) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист_1"

    if kind == "erp_articles":
        sheet.cell(3, 1, "Иерархия статей доходов и расходов")
        sheet.cell(6, 15, "Код элемента")
        rows = [
            (8, "Административные", 0, None),
            (9, "Связь", 2, None),
            (10, "Интернет" if not supplement else "Телефония", 4, None),
            (11, None, 0, "ERP-001" if not supplement else "ERP-003"),
        ]
        for row, name, level, code in rows:
            if name is not None:
                cell = sheet.cell(row, 1, name)
                cell.alignment = Alignment(indent=level)
            if code:
                sheet.cell(row, 15, code)

    elif kind == "organizations":
        sheet.cell(2, 1, "Иерархия организаций")
        sheet.cell(5, 39, "Код справочника")
        rows = (
            [
                (8, "Группа", "ORG-1", 0),
                (9, "ПС", "ORG-2", 1),
                (10, "ЦФО 1", "ORG-3", 2),
            ]
            if not supplement
            else [
                (8, "Группа", "ORG-1", 0),
                (9, "ПС", "ORG-2", 1),
                (10, "ЦФО 1 обновлён", "ORG-3", 2),
                (11, "ЦФО 2", "ORG-4", 2),
            ]
        )
        for row, name, code, level in rows:
            cell = sheet.cell(row, 1, name)
            cell.alignment = Alignment(indent=level)
            sheet.cell(row, 39, code)

    elif kind == "scenarios":
        sheet.cell(4, 1, "Сценарии бюджетирования")
        sheet.cell(7, 19, "Код записи")
        rows = (
            [
                (9, "ПЛАН 2026", 10),
                (10, "Факт", 1),
            ]
            if not supplement
            else [
                (9, "ПЛАН 2026", 10),
                (10, "ПЛАН 2027", 11),
            ]
        )
        for row, name, code in rows:
            sheet.cell(row, 1, name)
            code_cell = sheet.cell(row, 19, code)
            code_cell.number_format = "00000"
    else:
        raise AssertionError(kind)

    return _bytes(workbook)




def _shared_strings_case_mismatched_cfo_reference() -> bytes:
    parts = {
        "[Content_Types].xml": """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>
</Types>""",
        "_rels/.rels": """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>""",
        "xl/workbook.xml": """<?xml version="1.0" encoding="UTF-8"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets><sheet name="TDSheet" sheetId="1" r:id="rId1"/></sheets>
</workbook>""",
        "xl/_rels/workbook.xml.rels": """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>
</Relationships>""",
        "xl/worksheets/sheet1.xml": """<?xml version="1.0" encoding="UTF-8"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetData>
    <row r="1"><c r="A1" t="s"><v>0</v></c><c r="B1" t="s"><v>1</v></c><c r="C1" t="s"><v>2</v></c></row>
    <row r="2"><c r="A2" t="s"><v>3</v></c><c r="B2" t="s"><v>4</v></c><c r="C2" t="s"><v>5</v></c></row>
  </sheetData>
</worksheet>""",
        # Deliberately capitalized differently from the relationship/content type.
        "xl/SharedStrings.xml": """<?xml version="1.0" encoding="UTF-8"?>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="6" uniqueCount="6">
  <si><t>Код ЦФО Инталев</t></si>
  <si><t>Наименование ЦФО Инталев</t></si>
  <si><t>Полный путь ЦФО</t></si>
  <si><t>0000000007</t></si>
  <si><t>ЦД/ЦЗ Фонд развития</t></si>
  <si><t>ЦД/ЦЗ Фонд развития</t></si>
</sst>""",
    }
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        for name, payload in parts.items():
            archive.writestr(name, payload)
    return output.getvalue()


def test_reference_import_repairs_shared_strings_case_and_persists_cfo(tmp_path):
    runtime = tmp_path / "runtime"
    service = WorkflowService(runtime)

    imported = service.upload_reference(
        "intalev_cfos", _shared_strings_case_mismatched_cfo_reference()
    )

    assert imported == 1
    assert service.reference_counts()["intalev_cfos"] == 16
    imported_item = next(
        item for item in service.intalev_cfos()
        if item.source_key == "code:0000000007"
    )
    assert imported_item.name == "ЦД/ЦЗ Фонд развития"
    restarted_item = next(
        item for item in WorkflowService(runtime).intalev_cfos()
        if item.source_key == "code:0000000007"
    )
    assert restarted_item.source_key == "code:0000000007"

def test_new_store_starts_with_packaged_baselines_and_keeps_them_on_restart(tmp_path):
    runtime = tmp_path / "runtime"
    first = WorkflowService(runtime)

    assert first.reference_counts() == {
        "erp_articles": 271,
        "organizations": 357,
        "scenarios": 12,
        "intalev_cfos": 16,
        "article_indicators": 208,
        "opiu_formulas": 517,
        "opiu_analytics": 517,
        "regions": 22,
        "sales_networks": 233,
        "opiu_report_indicators": 683,
        "opiu_source_rules": 310,
    }
    assert first.store.catalog_source("erp_articles") == "baseline"
    assert first.store.catalog_source("organizations") == "baseline"
    assert first.store.catalog_source("scenarios") == "baseline"
    assert first.store.catalog_source("intalev_cfos") == "baseline"
    assert first.store.catalog_source("article_indicators") == "baseline"

    restarted = WorkflowService(runtime)
    assert restarted.reference_counts() == first.reference_counts()
    assert restarted.store.catalog_source("scenarios") == "baseline"


def test_first_explicit_upload_supplements_its_packaged_baseline(tmp_path):
    service = WorkflowService(tmp_path / "runtime")

    service.upload_reference("organizations", _split_header_reference("organizations"))

    assert service.reference_counts()["organizations"] == 360
    assert service.reference_counts()["erp_articles"] == 271
    assert service.reference_counts()["scenarios"] == 12
    assert service.reference_counts()["intalev_cfos"] == 16
    assert service.store.catalog_source("organizations") == "user"

def test_real_exports_accept_split_multiline_header_rows(tmp_path):
    service = WorkflowService(tmp_path / "runtime")

    assert service.upload_reference(
        "erp_articles", _split_header_reference("erp_articles")
    ) == 1
    assert service.upload_reference(
        "organizations", _split_header_reference("organizations")
    ) == 3
    assert service.upload_reference(
        "scenarios", _split_header_reference("scenarios")
    ) == 2

    assert any(article.code == "ERP-001" for article in service.erp_articles())
    imported_nodes = {
        node.code for node in service.organization_nodes()
        if node.code.startswith("ORG-")
    }
    assert imported_nodes == {"ORG-1", "ORG-2", "ORG-3"}
    plan = next(
        scenario
        for scenario in service.store.list_scenarios()
        if scenario.name == "ПЛАН 2026"
    )
    assert plan.erp_code == "00010"


def test_organization_catalog_is_global_persistent_and_supplemented(tmp_path):
    runtime = tmp_path / "runtime"
    first = WorkflowService(runtime)

    first.upload_reference(
        "organizations", _split_header_reference("organizations")
    )
    first.upload_reference(
        "organizations",
        _split_header_reference("organizations", supplement=True),
    )

    assert first.reference_counts()["organizations"] == 361
    assert {
        node.code: node.name for node in first.organization_nodes()
    }["ORG-3"] == "ЦФО 1 обновлён"

    restarted = WorkflowService(runtime)
    assert restarted.reference_counts()["organizations"] == 361
    assert {
        node.code for node in restarted.allowed_organization_nodes()
        if node.code.startswith("ORG-")
    } == {"ORG-1", "ORG-2", "ORG-3", "ORG-4"}


def test_scenario_catalog_is_loaded_once_and_incrementally_supplemented(tmp_path):
    runtime = tmp_path / "runtime"
    first = WorkflowService(runtime)

    first.upload_reference("scenarios", _split_header_reference("scenarios"))
    original_plan = next(
        scenario
        for scenario in first.store.list_scenarios()
        if scenario.name == "ПЛАН 2026"
    )

    first.upload_reference(
        "scenarios",
        _split_header_reference("scenarios", supplement=True),
    )

    scenarios = first.store.list_scenarios()
    assert {"ПЛАН 2026", "ПЛАН 2027", "Факт"}.issubset(
        {scenario.name for scenario in scenarios}
    )
    assert next(
        scenario for scenario in scenarios if scenario.name == "ПЛАН 2026"
    ).scenario_id == original_plan.scenario_id

    restarted = WorkflowService(runtime)
    assert {"ПЛАН 2026", "ПЛАН 2027", "Факт"}.issubset(
        {scenario.name for scenario in restarted.store.list_scenarios()}
    )


def test_exact_user_update_of_baseline_survives_restart_without_losing_catalog(tmp_path):
    runtime = tmp_path / "runtime"
    service = WorkflowService(runtime)
    original = service.store.load_reference("organizations")[0]
    updated = dict(original)
    updated["name"] = f"{original['name']} · локальное уточнение"
    updated["full_path"] = f"{original['full_path']} · локальное уточнение"

    service.store.replace_reference("organizations", [updated])

    assert service.reference_counts()["organizations"] == 357
    assert next(
        item for item in service.store.load_reference("organizations")
        if item["node_id"] == original["node_id"]
    )["name"].endswith("локальное уточнение")

    restarted = WorkflowService(runtime)
    assert restarted.reference_counts()["organizations"] == 357
    persisted = next(
        item for item in restarted.store.load_reference("organizations")
        if item["node_id"] == original["node_id"]
    )
    assert persisted["name"].endswith("локальное уточнение")
    assert restarted.store.catalog_source("organizations") == "user"
