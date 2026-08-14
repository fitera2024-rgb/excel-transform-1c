from __future__ import annotations

import os
import shutil
import subprocess
from io import BytesIO
from pathlib import Path

import pytest
import xlwt
from openpyxl import load_workbook

from excel_transform_1c.adapters.excel import detect_path
from excel_transform_1c.adapters.protected_ooxml import decrypt_protected_ooxml
from excel_transform_1c.adapters.workbook_repair import (
    WorkbookFormat,
    default_working_path,
    detect_workbook_format,
    prepare_workbook,
)
from tests.helpers.workbooks import HEADERS, protected_workbook_bytes, workbook_bytes
from tests.integration.test_workflow import configured_service, default_context


def _legacy_budget_bytes() -> bytes:
    output = BytesIO()
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Legacy budget")
    sheet.write(0, 0, "Synthetic fixture")
    for column, value in enumerate(HEADERS):
        sheet.write(1, column, value)
    rows = [
        ["ПС", "Административные", "Департамент 1", "ТК", "ЦФО 1", 0.2, "Связь", "Интернет", 100],
        ["ПС", "Коммерческие", "Департамент 2", "ТК", "ЦФО 2", "БЕЗ НДС", "Маркетинг", "Реклама", 0],
    ]
    for row_number, values in enumerate(rows, start=2):
        for column, value in enumerate([*values, *([0] * 11)]):
            sheet.write(row_number, column, value)
    workbook.save(output)
    return output.getvalue()


def test_legacy_biff_disguised_as_xlsx_completes_intake_and_snapshot(tmp_path):
    service = configured_service(tmp_path)
    original = _legacy_budget_bytes()

    pending = service.prepare_upload(
        "legacy-budget.xlsx",
        original,
        default_context(service),
    )
    working_copy = default_working_path(pending.original_path)
    run = service.process_upload(pending.upload_id, pending.candidates[0].candidate_id)
    run_path = service.run_dir / run.run_id

    assert pending.is_protected is False
    assert len(pending.candidates) == 1
    assert working_copy != pending.original_path
    assert working_copy.exists()
    assert pending.original_path.read_bytes() == original
    assert (run_path / "source-original.xlsx").read_bytes() == original
    assert len(run.records) == 24



def test_true_xls_extension_completes_workflow_without_filename_heuristic(tmp_path):
    service = configured_service(tmp_path)
    original = _legacy_budget_bytes()

    pending = service.prepare_upload(
        "legacy-budget.xls",
        original,
        default_context(service),
    )
    run = service.process_upload(pending.upload_id, pending.candidates[0].candidate_id)

    assert pending.original_path.suffix == ".xls"
    assert pending.original_path.read_bytes() == original
    assert len(run.records) == 24

def test_plain_ooxml_intake_uses_exact_separate_working_copy(tmp_path):
    source = tmp_path / "plain.xlsx"
    original = workbook_bytes()
    source.write_bytes(original)

    prepared = prepare_workbook(source)

    assert prepared.working_path != source
    assert source.read_bytes() == original
    assert prepared.working_path.read_bytes() == original
    assert len(detect_path(prepared.working_path)) == 1


def test_protected_ay_and_pv_synthetic_regression(tmp_path):
    password = "synthetic-ay-pv-password"
    for label in ("ay", "pv"):
        source = tmp_path / f"{label}.xlsx"
        target = tmp_path / f"{label}-working.xlsx"
        source.write_bytes(protected_workbook_bytes(workbook_bytes(), password))

        assert detect_workbook_format(source) is WorkbookFormat.ENCRYPTED_OOXML
        decrypt_protected_ooxml(source, target, password)
        assert detect_workbook_format(target) is WorkbookFormat.OOXML
        assert len(detect_path(target)) == 1


@pytest.mark.parametrize(
    "environment_name",
    ["EXCEL_INTAKE_REAL_CFO_FILE", "EXCEL_INTAKE_REAL_OPIU_FILE"],
)
def test_real_unprotected_workbook_passes_intake_when_available(
    tmp_path,
    environment_name,
):
    configured_path = os.environ.get(environment_name)
    if not configured_path:
        pytest.skip(f"{environment_name} is not configured")
    source = Path(configured_path)
    if not source.is_file():
        pytest.skip(f"{environment_name} is not available")

    staged = tmp_path / "real-input.xlsx"
    shutil.copyfile(source, staged)
    prepared = prepare_workbook(staged)
    workbook = load_workbook(prepared.working_path, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames
    finally:
        workbook.close()


@pytest.mark.parametrize(
    "environment_name",
    ["EXCEL_INTAKE_REAL_AY_FILE", "EXCEL_INTAKE_REAL_PV_FILE"],
)
def test_real_protected_workbook_passes_intake_when_available(
    tmp_path,
    environment_name,
):
    configured_path = os.environ.get(environment_name)
    password = os.environ.get("EXCEL_INTAKE_REAL_PASSWORD")
    if not configured_path or password is None:
        pytest.skip("protected real-file smoke is not configured")
    source = Path(configured_path)
    if not source.is_file():
        pytest.skip("protected real-file smoke is not available")

    staged = tmp_path / "protected-input.xlsx"
    working = tmp_path / "working.xlsx"
    shutil.copyfile(source, staged)
    decrypt_protected_ooxml(staged, working, password)
    prepared = prepare_workbook(working)
    workbook = load_workbook(prepared.working_path, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames
    finally:
        workbook.close()


def test_source_excel_files_are_not_tracked_in_git():
    repository = Path(__file__).resolve().parents[2]
    completed = subprocess.run(
        [
            "git",
            "-C",
            str(repository),
            "ls-files",
            "*.xls",
            "*.xlsx",
            "*.xlsm",
            "*.xlsb",
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    assert completed.stdout.strip() == ""
