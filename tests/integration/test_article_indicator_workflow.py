from io import BytesIO

import pytest
from openpyxl import load_workbook

from excel_transform_1c.adapters.excel import ADO_INDICATOR_HEADERS
from excel_transform_1c.adapters.references import parse_reference_workbook
from excel_transform_1c.application import service as service_module
from excel_transform_1c.application.service import WorkflowService
from tests.helpers.workbooks import (
    indicator_classifier_bytes,
    reference_bytes,
    workbook_bytes,
)


pytestmark = pytest.mark.integration


def configured_service(tmp_path) -> WorkflowService:
    service = WorkflowService(tmp_path / "runtime")
    for kind in ("erp_articles", "organizations", "scenarios", "intalev_cfos"):
        service.upload_reference(kind, reference_bytes(kind))
    service.store.save_cfo_mappings(
        {"code:INT-CFO-1": "cfo", "code:INT-CFO-2": "other"}
    )
    return service


def process_run(service: WorkflowService):
    scenario = service.store.list_scenarios()[0]
    context = service.build_context("ПС", "ps", scenario.scenario_id, 2026, [])
    pending = service.prepare_upload("synthetic.xlsx", workbook_bytes(), context)
    return service.process_upload(
        pending.upload_id,
        pending.candidates[0].candidate_id,
    )


def workbook_rows(payload: bytes, sheet_name: str) -> tuple[tuple[object, ...], ...]:
    workbook = load_workbook(BytesIO(payload), data_only=True)
    try:
        sheet = workbook[sheet_name]
        return tuple(tuple(cell.value for cell in row) for row in sheet.iter_rows())
    finally:
        workbook.close()


def test_classifier_parser_and_persistence_use_direct_exact_keys(tmp_path):
    payload = parse_reference_workbook(
        indicator_classifier_bytes(),
        "article_indicators",
    )
    assert payload[0] == {
        "erp_code": "ERP-001",
        "article_path": "Административные → Связь → Интернет",
        "article_name": "Интернет",
        "indicator": "Услуги связи",
        "sales_channel": "Основной канал",
    }

    first = WorkflowService(tmp_path / "runtime")
    assert first.upload_indicator_classifier(indicator_classifier_bytes()) == 2
    second = WorkflowService(tmp_path / "runtime")

    assert second.article_indicator_rules() == first.article_indicator_rules()
    assert len(second.article_indicator_rules()) == 2


def test_classifier_supplement_repeats_search_in_current_run_without_reading_excel(
    tmp_path,
    monkeypatch,
):
    service = configured_service(tmp_path)
    run = process_run(service)
    assert service.indicator_counts(run.run_id) == {
        "automatic": 0,
        "attention": 2,
        "not_found": 2,
    }
    assert run.indicator_classifier_loaded is False

    def unexpected_read(*_args, **_kwargs):
        raise AssertionError("Исходный Excel не должен читаться повторно")

    monkeypatch.setattr(service_module, "read_path", unexpected_read)
    original_run = service.get_run(run.run_id)

    service.upload_indicator_classifier(indicator_classifier_bytes(), run.run_id)

    assert service.get_run(run.run_id) is original_run
    assert run.rerun_count == 0
    assert run.indicator_classifier_loaded is True
    assert service.indicator_counts(run.run_id) == {
        "automatic": 2,
        "attention": 0,
        "not_found": 0,
    }


def test_ambiguous_and_missing_classifier_results_remain_unapplied(tmp_path):
    service = configured_service(tmp_path)
    run = process_run(service)
    ambiguous = indicator_classifier_bytes(
        [
            {
                "erp_code": "OTHER-1",
                "article_name": "Интернет",
                "indicator": "Первый",
                "sales_channel": "Канал",
            },
            {
                "erp_code": "OTHER-2",
                "article_name": "Интернет",
                "indicator": "Второй",
                "sales_channel": "Канал",
            },
        ]
    )

    service.upload_indicator_classifier(ambiguous, run.run_id)

    assert service.indicator_counts(run.run_id) == {
        "automatic": 0,
        "attention": 2,
        "not_found": 1,
    }
    internet = next(record for record in run.records if record.source_article == "Интернет")
    assert internet.indicator == ""
    assert internet.sales_channel == ""


def test_indicator_export_populates_third_sheet_without_changing_first_two(tmp_path):
    service = configured_service(tmp_path)
    run = process_run(service)
    before = service.export_run(run.run_id)

    service.upload_indicator_classifier(indicator_classifier_bytes(), run.run_id)
    after = service.export_run(run.run_id)

    assert workbook_rows(after, "OPIU Light") == workbook_rows(before, "OPIU Light")
    assert workbook_rows(after, "ОПИУ") == workbook_rows(before, "ОПИУ")

    workbook = load_workbook(BytesIO(after), data_only=True)
    try:
        assert workbook.sheetnames == ["OPIU Light", "ОПИУ", "Показатели"]
        indicators = workbook["Показатели"]
        assert tuple(cell.value for cell in indicators[1]) == ADO_INDICATOR_HEADERS
        assert indicators.max_row == 25
        assert {cell.value for cell in indicators["G"][1:]} == {
            "Услуги связи",
            "Маркетинговые расходы",
        }
        assert any(cell.value == 0 for cell in indicators["H"][1:])
    finally:
        workbook.close()
