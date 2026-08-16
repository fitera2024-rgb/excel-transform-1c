from io import BytesIO
import os
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from excel_transform_1c.application.service import WorkflowService
from excel_transform_1c.adapters.excel import ADO_INDICATOR_HEADERS
from excel_transform_1c.core.detection import detect_candidate_ranges
from excel_transform_1c.core.indicator_matching import (
    INDICATOR_MATCHED,
    INDICATOR_MISSING,
)
from excel_transform_1c.core.models import IndicatorType, STATUS_ATTENTION
from tests.helpers.workbooks import HEADERS


pytestmark = pytest.mark.integration


def _expense_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Расходы"
    sheet.append(HEADERS)
    sheet.append(
        [
            "ПС",
            "Административные расходы",
            "Департамент",
            "ТК",
            "ЦФО",
            "БЕЗ НДС",
            "Командировочные",
            "Проживание",
            100,
            *([0] * 11),
        ]
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _income_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Доходы"
    sheet.append(
        [
            "ТИП ДОХОДОВ",
            "ГРУППА ДОХОДОВ",
            "СТАТЬЯ",
            "АНАЛИТИКА",
            "ЯНВАРЬ",
            "ФЕВРАЛЬ",
            "МАРТ",
            "АПРЕЛЬ",
            "МАЙ",
            "ИЮНЬ",
            "ИЮЛЬ",
            "АВГУСТ",
            "СЕНТЯБРЬ",
            "ОКТЯБРЬ",
            "НОЯБРЬ",
            "ДЕКАБРЬ",
        ]
    )
    sheet.append(
        [
            "Продажи",
            "Выручка_продажи внешние",
            "HoReCa",
            "Аби Продакт",
            200,
            *([0] * 11),
        ]
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _bdr_revenue_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "БДР 2026 ИТОГ"
    for month in range(1, 13):
        column = 22 + month
        sheet.cell(2, column, datetime(2026, month, 1))
        sheet.cell(3, column, "план")
    group = "Выручка_продажи внешние"
    articles = (
        "Опт",
        "Розница",
        "HoReCa",
        "Сети ДВ",
        "Сети Федеральные",
        "Дискаунтеры ДВ",
        "Дискаунтеры Федеральные",
    )
    sheet.cell(56, 1, group)
    sheet.cell(56, 7, group)
    for offset, article in enumerate(articles, start=1):
        row = 56 + offset
        sheet.cell(row, 1, group)
        sheet.cell(row, 7, article)
        for month in range(1, 13):
            sheet.cell(row, 22 + month, offset * month)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _quantity_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Количество"
    sheet.append([*HEADERS, "Тип показателя", "Номенклатура", "Единица измерения"])
    sheet.append(
        [
            "ПС",
            "Количественные показатели",
            "Департамент",
            "ТК",
            "ЦФО",
            "?",
            "Продукция",
            "Товар А",
            5,
            *([0] * 11),
            "QUANTITY",
            "Товар А",
            "кг",
        ]
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _clean_context(service: WorkflowService):
    organization = service.organization_nodes()[0]
    scenario = service.store.list_scenarios()[0]
    return service.build_context(
        "ПС",
        organization.node_id,
        scenario.scenario_id,
        2026,
        [],
    )


def test_clean_install_applies_packaged_mxl_link_without_inventing_channel(tmp_path):
    service = WorkflowService(tmp_path / "runtime")
    context = _clean_context(service)
    pending = service.prepare_upload("expense.xlsx", _expense_workbook(), context)
    run = service.process_upload(
        pending.upload_id,
        pending.candidates[0].candidate_id,
    )

    assert run.indicator_classifier_loaded is True
    assert service.indicator_counts(run.run_id) == {
        "automatic": 1,
        "attention": 0,
        "not_found": 0,
    }
    record = run.records[0]
    assert record.indicator == "Административные расходы"
    assert record.sales_channel == ""
    assert record.indicator_match_status == INDICATOR_MATCHED
    assert record.indicator_match_reason == ""

    exported = load_workbook(BytesIO(service.export_run(run.run_id)), data_only=True)
    try:
        assert exported.sheetnames == ["OPIU Light", "ОПИУ", "Показатели"]
        indicators = exported["Показатели"]
        assert tuple(cell.value for cell in indicators[1]) == ADO_INDICATOR_HEADERS
        assert indicators.max_row == 13
        assert {cell.value for cell in indicators["K"][1:]} == {"Расход"}
        assert {cell.value for cell in indicators["L"][1:]} == {None}
        assert {cell.value for cell in indicators["M"][1:]} == {
            "Административные расходы"
        }
        assert sum(cell.value for cell in indicators["N"][1:]) == 100
    finally:
        exported.close()


def test_clean_install_resolves_structural_income_from_packaged_rules(tmp_path):
    service = WorkflowService(tmp_path / "runtime")
    pending = service.prepare_upload(
        "bdr-2026.xlsx",
        _bdr_revenue_workbook(),
        _clean_context(service),
    )

    assert len(pending.candidates) == 1
    assert pending.candidates[0].source_kind == "bdr_revenue_summary"
    assert pending.candidates[0].source_year == 2026
    run = service.process_upload(
        pending.upload_id,
        pending.candidates[0].candidate_id,
    )

    assert service.indicator_counts(run.run_id) == {
        "automatic": 7,
        "attention": 0,
        "not_found": 0,
    }
    record = run.records[0]
    assert record.indicator_type == IndicatorType.REVENUE
    assert record.indicator == "Опт"
    assert record.sales_channel == "Опт"
    assert record.input_sales_channel == "Опт"
    assert record.erp_code == ""
    assert record.indicator_match_status == INDICATOR_MATCHED

    exported = load_workbook(BytesIO(service.export_run(run.run_id)), data_only=True)
    try:
        assert exported.sheetnames == ["OPIU Light", "ОПИУ", "Показатели"]
        indicators = exported["Показатели"]
        expected = {
            "Опт",
            "Розница",
            "HoReCa",
            "Сети ДВ",
            "Сети Федеральные",
            "Дискаунтеры ДВ",
            "Дискаунтеры Федеральные",
        }
        assert {cell.value for cell in indicators["K"][1:]} == {"Доход"}
        assert {cell.value for cell in indicators["L"][1:]} == expected
        assert {cell.value for cell in indicators["M"][1:]} == expected
        assert sum(cell.value for cell in indicators["N"][1:]) == sum(
            float(record.amount) for record in run.records
        )
    finally:
        exported.close()


def test_bdr_revenue_schema_rejects_non_exact_article_set():
    workbook = load_workbook(BytesIO(_bdr_revenue_workbook()))
    try:
        workbook["БДР 2026 ИТОГ"]["G63"] = "дискаунтеры федеральные"

        assert not [
            candidate
            for candidate in detect_candidate_ranges(workbook)
            if candidate.source_kind == "bdr_revenue_summary"
        ]
    finally:
        workbook.close()


def test_clean_install_keeps_unproven_quantity_at_attention(tmp_path):
    service = WorkflowService(tmp_path / "runtime")
    pending = service.prepare_upload(
        "quantity.xlsx",
        _quantity_workbook(),
        _clean_context(service),
    )
    run = service.process_upload(
        pending.upload_id,
        pending.candidates[0].candidate_id,
    )

    assert service.indicator_counts(run.run_id) == {
        "automatic": 0,
        "attention": 1,
        "not_found": 1,
    }
    record = run.records[0]
    assert record.indicator_type == IndicatorType.QUANTITY
    assert record.status == STATUS_ATTENTION
    assert record.indicator_match_status == INDICATOR_MISSING
    assert record.indicator == ""
    assert "Точная связь количества" in record.indicator_match_reason

    exported = load_workbook(BytesIO(service.export_run(run.run_id)), data_only=True)
    try:
        assert exported.sheetnames == ["OPIU Light", "ОПИУ", "Показатели"]
        assert exported["Показатели"].max_row == 1
    finally:
        exported.close()


def test_real_owner_budget_revenue_smoke_when_available(tmp_path):
    configured = os.environ.get("CODEX06_REAL_BUDGET_FILE")
    source = Path(configured) if configured else None
    if source is None or not source.is_file():
        pytest.skip("Реальная owner-книга бюджета недоступна")

    service = WorkflowService(tmp_path / "runtime")
    pending = service.prepare_upload(
        source.name,
        source.read_bytes(),
        _clean_context(service),
    )
    candidates = [
        candidate
        for candidate in pending.candidates
        if candidate.source_kind == "bdr_revenue_summary"
        and candidate.sheet == "БДР 2026 ИТОГ"
    ]
    assert len(candidates) == 1
    expense_candidates = [
        candidate
        for candidate in pending.candidates
        if candidate.source_kind == "prepared_budget"
        and candidate.sheet == "загрузка ERP расходы"
    ]
    assert len(expense_candidates) == 1

    run = service.process_upload(pending.upload_id, candidates[0].candidate_id)

    assert len({record.source_row for record in run.records}) == 7
    assert len(run.records) == 84
    assert service.indicator_counts(run.run_id) == {
        "automatic": 7,
        "attention": 0,
        "not_found": 0,
    }
    assert {record.indicator for record in run.records} == {
        "Опт",
        "Розница",
        "HoReCa",
        "Сети ДВ",
        "Сети Федеральные",
        "Дискаунтеры ДВ",
        "Дискаунтеры Федеральные",
    }
    assert all(record.sales_channel == record.input_sales_channel for record in run.records)
    assert all(record.erp_code == "" for record in run.records)

    exported = load_workbook(BytesIO(service.export_run(run.run_id)), data_only=True)
    try:
        assert exported.sheetnames == ["OPIU Light", "ОПИУ", "Показатели"]
        exported_total = sum(
            cell.value or 0 for cell in exported["Показатели"]["N"][1:]
        )
        source_total = sum(float(record.amount) for record in run.records)
        assert exported_total == pytest.approx(source_total)
    finally:
        exported.close()

    expense_run = service.process_upload(
        pending.upload_id,
        expense_candidates[0].candidate_id,
    )
    matched_expenses = [
        record
        for record in expense_run.records
        if record.indicator_match_status == INDICATOR_MATCHED
    ]
    assert matched_expenses
    assert all(record.indicator_type == IndicatorType.EXPENSE for record in matched_expenses)
    assert all(record.sales_channel == "" for record in matched_expenses)

    expense_export = load_workbook(
        BytesIO(service.export_run(expense_run.run_id)),
        data_only=True,
    )
    try:
        assert expense_export.sheetnames == ["OPIU Light", "ОПИУ", "Показатели"]
        indicators = expense_export["Показатели"]
        assert indicators.max_row > 1
        assert {cell.value for cell in indicators["L"][1:]} == {None}
    finally:
        expense_export.close()
