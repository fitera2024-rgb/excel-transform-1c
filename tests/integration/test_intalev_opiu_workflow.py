from io import BytesIO
import os
from pathlib import Path

import pytest
from openpyxl import load_workbook

from excel_transform_1c.adapters.excel import EXPORT_HEADERS
from excel_transform_1c.application.service import WorkflowService
from tests.helpers.workbooks import intalev_opiu_bytes, reference_bytes


pytestmark = pytest.mark.integration


def _service(tmp_path) -> WorkflowService:
    service = WorkflowService(tmp_path / "runtime")
    for kind in ("erp_articles", "organizations", "scenarios", "intalev_cfos"):
        service.upload_reference(kind, reference_bytes(kind))
    return service


def _context(service: WorkflowService):
    scenario = service.store.add_scenario("Факт", 2025, erp_code="00001", erp_confirmed=True)
    return service.build_context("ПС", "ps", scenario.scenario_id, 2025, [])


def test_intalev_preview_and_valid_export_preserve_business_values(tmp_path):
    service = _service(tmp_path)
    pending = service.prepare_upload(
        "synthetic-intalev.xlsx",
        intalev_opiu_bytes(sheet_name="Не TDSheet", monthly_error=True),
        _context(service),
    )
    assert len(pending.candidates) == 1
    run = service.process_upload(pending.upload_id, pending.candidates[0].candidate_id)
    assert len(run.records) == 36
    assert all(record.source_cfo == "ЦД/ЦЗ Фонд развития" for record in run.records)

    payload = service.export_run(run.run_id)
    workbook = load_workbook(BytesIO(payload), data_only=True)
    try:
        sheet = workbook["OPIU Light"]
        assert tuple(cell.value for cell in sheet[1]) == EXPORT_HEADERS
        assert sheet.max_row == 37
        values = [tuple(cell.value for cell in row) for row in sheet.iter_rows(min_row=2)]
        assert sum(row[20] == 0 for row in values) == 34
        assert any(row[20] == -10 for row in values)
        skipped = next(row for row in values if row[23] == 10 and row[5] == 5)
        assert skipped[20] is None
        assert skipped[21] == "Пропущено"
        assert "I10" in skipped[22]
        assert all(row[10] == "ЦД/ЦЗ Фонд развития" for row in values)
        assert sheet.freeze_panes == "A2"
    finally:
        workbook.close()


def test_existing_ayu_and_pv_prepared_budget_inputs_do_not_regress(tmp_path):
    service = _service(tmp_path)
    for reporting_unit in ("АЮ", "ПВ"):
        context = _context(service)
        context = context.__class__(
            reporting_unit,
            context.organization_node_id,
            context.organization_name,
            context.scenario_id,
            context.scenario_name,
            context.scenario_year,
            context.scenario_erp_confirmed,
            context.year,
            context.selected_months,
        )
        from tests.helpers.workbooks import workbook_bytes

        pending = service.prepare_upload(
            f"{reporting_unit}.xlsx",
            workbook_bytes(reporting_unit=reporting_unit),
            context,
        )
        run = service.process_upload(pending.upload_id, pending.candidates[0].candidate_id)
        assert run.candidate.source_kind == "prepared_budget"
        assert len(run.records) == 24


def test_real_intalev_file_when_available(tmp_path):
    configured = os.environ.get("EXCEL_INTAKE_REAL_OPIU_FILE")
    if not configured:
        pytest.skip("Exact real-file handoff EXCEL_INTAKE_REAL_OPIU_FILE не задан")
    source = Path(configured)
    if not source.is_file():
        pytest.skip("Exact real-file handoff недоступен")

    service = _service(tmp_path)
    pending = service.prepare_upload(source.name, source.read_bytes(), _context(service))
    assert pending.candidates
    candidate = pending.candidates[0]
    assert candidate.source_kind == "intalev_opiu"
    assert candidate.header_row == 4
    assert candidate.first_data_row == 7
    assert candidate.last_data_row >= candidate.first_data_row
    run = service.process_upload(pending.upload_id, candidate.candidate_id)
    assert len(run.records) == len({record.source_row for record in run.records}) * 12
