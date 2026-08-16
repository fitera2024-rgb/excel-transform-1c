import re
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from excel_transform_1c.adapters.excel import (
    ADO_INDICATOR_HEADERS,
    ADO_OPIU_HEADERS,
    EXPORT_HEADERS,
)
from excel_transform_1c.ui.app import create_app
from tests.helpers.workbooks import (
    BDR_FULL_INDICATORS,
    bdr_formula_workbook_bytes,
    erp_organization_hierarchy_bytes,
)


pytestmark = pytest.mark.integration


def test_bdr_full_start_preview_export_stop(tmp_path) -> None:
    app = create_app(tmp_path / "runtime")
    with TestClient(app) as client:
        response = client.post(
            "/references",
            data={"kind": "organizations"},
            files={
                "reference_file": (
                    "ОрганизациииерархияЕРП.xlsx",
                    erp_organization_hierarchy_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            follow_redirects=True,
        )
        assert response.status_code == 200
        scenario = next(
            item
            for item in client.app.state.workflow.store.list_scenarios()
            if item.name == "ПЛАН 2026"
        )
        reporting_unit = "АЮ Административный Отдел"
        response = client.post(
            "/uploads",
            data={
                "reporting_unit": reporting_unit,
                "organization_node_id": "000000001",
                "scenario_id": scenario.scenario_id,
                "year": "2026",
                "period_selector_present": "1",
                "all_year": "1",
            },
            files={
                "budget_file": (
                    "БДР 2026 ИТОГ.xlsx",
                    bdr_formula_workbook_bytes(reporting_unit=reporting_unit),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            follow_redirects=True,
        )
        assert response.status_code == 200
        assert 'data-testid="bdr-diagnostics"' in response.text
        assert "Прочитано строк БДР:" in response.text
        assert "Доходные показатели:" in response.text
        assert "Расходные показатели:" in response.text
        assert "KPI найдено:" in response.text
        assert "KPI с организацией:" in response.text
        assert "KPI с периодом:" in response.text
        assert "KPI со значением:" in response.text
        assert "KPI экспортировано:" in response.text
        assert "Все прочитанные показатели имеют экспортируемые значения" in response.text
        assert 'data-testid="indicator-classifier-summary"' not in response.text

        match = re.search(r"/runs/([a-f0-9]+)", str(response.url))
        assert match
        export = client.get(f"/runs/{match.group(1)}/export")
        assert export.status_code == 200
        workbook = load_workbook(BytesIO(export.content), data_only=True)
        try:
            assert workbook.sheetnames == ["OPIU Light", "ОПИУ", "Показатели"]
            assert tuple(cell.value for cell in workbook["OPIU Light"][1]) == EXPORT_HEADERS
            assert tuple(cell.value for cell in workbook["ОПИУ"][1]) == ADO_OPIU_HEADERS
            assert tuple(cell.value for cell in workbook["Показатели"][1]) == ADO_INDICATOR_HEADERS
            expected_source_rows = len(BDR_FULL_INDICATORS) - 5 + 1
            assert workbook["OPIU Light"].max_row == expected_source_rows * 12 + 1
            assert workbook["Показатели"].max_row > 1
            headers = {
                cell.value: cell.column
                for cell in workbook["OPIU Light"][1]
            }
            for required in (
                "Организация",
                "Код организации",
                "Департамент",
                "Отдел",
                "ЦФО",
                "Код ЦФО",
                "Тип показателя",
                "Показатель",
                "Период",
                "Значение",
            ):
                assert required in headers
            first = workbook["OPIU Light"][2]
            assert first[headers["Организация"] - 1].value == 'ООО "Айс Юнион"'
            assert first[headers["Код организации"] - 1].value == "000000001"
            assert first[headers["Департамент"] - 1].value == "Административный департамент"
            assert first[headers["Отдел"] - 1].value == "АЮ Административный Отдел"
            assert first[headers["ЦФО"] - 1].value == "АЮ Административный Отдел"
            assert first[headers["Код ЦФО"] - 1].value == "000000173"
            assert first[headers["Тип показателя"] - 1].value == "KPI"
            assert first[headers["Показатель"] - 1].value == "Оборот в кг"
            assert first[headers["Период"] - 1].value == "01.2026"
            assert first[headers["Значение"] - 1].value == 593845

            expense = next(
                row
                for row in workbook["OPIU Light"].iter_rows(min_row=2)
                if row[headers["Исходное название статьи"] - 1].value == "Интернет"
                and row[headers["Месяц"] - 1].value == 1
            )
            assert expense[headers["Тип расходов"] - 1].value == "Административные расходы"
            assert expense[headers["Группа расходов"] - 1].value == "Связь"
            assert expense[headers["ERP-код статьи"] - 1].value == "00-000069"
            assert expense[headers["Департамент"] - 1].value == "Административный департамент"
            assert expense[headers["ЦФО"] - 1].value == "АЮ Административный Отдел"
            assert expense[headers["Код ЦФО"] - 1].value == "000000173"
            assert expense[headers["Значение"] - 1].value == 1
        finally:
            workbook.close()
