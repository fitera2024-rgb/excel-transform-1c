import json
import re
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from excel_transform_1c.adapters.excel import (
    ADO_INDICATOR_HEADERS,
    ADO_OPIU_HEADERS,
    EXPORT_HEADERS,
)
from excel_transform_1c.core.models import (
    IndicatorType,
    STATUS_ATTENTION,
    TAX_NOT_REQUIRED,
)
from excel_transform_1c.ui.app import create_app
from tests.helpers.workbooks import (
    reference_bytes,
    revenue_quantity_classifier_bytes,
    revenue_quantity_workbook_bytes,
)


pytestmark = pytest.mark.integration


def test_start_service_upload_rules_preview_confirm_export_stop_service(tmp_path):
    # START_SERVICE / STOP_SERVICE are provided by TestClient's lifespan context.
    with TestClient(create_app(tmp_path / "runtime")) as client:
        for kind in ("erp_articles", "organizations", "scenarios", "intalev_cfos"):
            response = client.post(
                "/references",
                data={"kind": kind},
                files={
                    "reference_file": (
                        f"synthetic-{kind}.xlsx",
                        reference_bytes(kind),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                follow_redirects=True,
            )
            assert response.status_code == 200

        service = client.app.state.workflow
        scenario = next(item for item in service.store.list_scenarios() if item.year == 2026)
        response = client.post(
            "/uploads",
            data={
                "reporting_unit": "ПС",
                "organization_node_id": "ps",
                "scenario_id": scenario.scenario_id,
                "year": "2026",
                "period_selector_present": "1",
                "all_year": "1",
            },
            files={
                "budget_file": (
                    "revenue-quantity.xlsx",
                    revenue_quantity_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            follow_redirects=True,
        )
        assert response.status_code == 200
        match = re.search(r"/runs/([a-f0-9]+)", str(response.url))
        assert match
        run_id = match.group(1)
        run = service.get_run(run_id)
        unresolved_by_row = {record.source_row: record for record in run.records}
        assert unresolved_by_row[4].status == STATUS_ATTENTION
        assert unresolved_by_row[5].status == STATUS_ATTENTION
        assert "не найдено" in unresolved_by_row[4].indicator_match_reason
        assert "не найдена" in unresolved_by_row[5].indicator_match_reason

        response = client.post(
            f"/runs/{run_id}/indicator-classifier",
            files={
                "classifier_file": (
                    "indicator-rules.xlsx",
                    revenue_quantity_classifier_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            follow_redirects=True,
        )
        assert response.status_code == 200
        for indicator in (
            "Административные расходы",
            "Выручка",
            "Количество продукции",
        ):
            assert indicator in response.text
        for label in ("Расход", "Доход", "Количество"):
            assert label in response.text
        assert "proof JSON" not in response.text
        assert "internal ID" not in response.text

        by_source_row = {record.source_row: record for record in run.records}
        assert by_source_row[3].indicator_type == IndicatorType.EXPENSE
        assert by_source_row[4].indicator_type == IndicatorType.REVENUE
        assert by_source_row[5].indicator_type == IndicatorType.QUANTITY
        revenue = by_source_row[4]
        assert revenue.counterparty == 'ООО "Покупатель"'
        assert revenue.input_sales_channel == "Сети Федеральные"
        assert revenue.sales_network == "Сеть 1"
        assert revenue.sales_region == "Приморский край"
        assert revenue.indicator == "Выручка"
        assert revenue.sales_channel == "Основной канал"

        response = client.post(
            f"/runs/{run_id}/confirm-tax-not-required",
            data={"confirmed": "1", "source_rows": json.dumps([5])},
            follow_redirects=True,
        )
        assert response.status_code == 200
        assert "Налогообложение отмечено как не требующееся: 1 строк" in response.text
        assert all(
            record.tax == TAX_NOT_REQUIRED
            for record in run.records
            if record.source_row == 5
        )

        response = client.get(f"/runs/{run_id}/export")
        assert response.status_code == 200
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        try:
            assert workbook.sheetnames == ["OPIU Light", "ОПИУ", "Показатели"]
            assert tuple(cell.value for cell in workbook["OPIU Light"][1]) == EXPORT_HEADERS
            assert tuple(cell.value for cell in workbook["ОПИУ"][1]) == ADO_OPIU_HEADERS
            assert (
                tuple(cell.value for cell in workbook["Показатели"][1])
                == ADO_INDICATOR_HEADERS
            )
            indicators = [
                cell.value for cell in workbook["Показатели"]["M"][1:]
            ]
            assert set(indicators) == {
                "Административные расходы",
                "Выручка",
                "Количество продукции",
            }
        finally:
            workbook.close()
