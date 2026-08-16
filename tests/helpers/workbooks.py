from __future__ import annotations

import base64
import hashlib
import re
from datetime import datetime
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.styles import Alignment
from msoffcrypto.format.ooxml import OOXMLFile


HEADERS = [
    "ПОДРАЗДЕЛЕНИЕ (ЦФО 1)",
    "ТИП РАСХОДОВ",
    "ДЕПАРТАМЕНТ (ЦФО 2)",
    "Вид организации",
    "ОТДЕЛ",
    "НАЛОГООБЛОЖЕНИЕ",
    "ГРУППА РАСХОДОВ",
    "СТАТЬЯ",
    "Январь",
    "Февраль",
    "Март",
    "Апрель",
    "Май",
    "Июнь",
    "Июль",
    "Август",
    "Сентябрь",
    "Октябрь",
    "Ноябрь",
    "Декабрь",
]


BDR_FULL_INDICATORS = (
    (10, "Оборот в кг"),
    (11, "Выручка за 1 кг"),
    (12, "Итого расходов на 1 кг"),
    (13, "Валовая прибыль на 1 кг"),
    (20, "Выручка ИТОГО"),
    (21, "Прочие доходы по основной деятельности"),
    (22, "Валовая прибыль"),
    (30, "Расходы по основной деятельности ИТОГО"),
    (31, "Административные расходы"),
    (32, "Коммерческие расходы"),
    (33, "Расходы на транспортную логистику"),
    (34, "Расходы на складскую логистику"),
    (40, "EBITDA"),
    (41, "Операционная прибыль"),
)


def bdr_full_workbook_bytes(
    *,
    reporting_unit: str = "АЮ Административный Отдел",
    with_internal_prepared_range: bool = True,
    error_indicator: str = "",
) -> bytes:
    """Synthetic whole-BDR business document with all three required blocks."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Произвольный сводный лист"
    sheet.cell(2, 1, reporting_unit)
    sheet.cell(2, 5, "Отдел")
    for month in range(1, 13):
        column = 22 + month
        sheet.cell(2, column, datetime(2026, month, 1))
        sheet.cell(3, column, "план")
    for row_number, indicator in BDR_FULL_INDICATORS:
        sheet.cell(row_number, 5, reporting_unit)
        sheet.cell(row_number, 7, indicator)
        for month in range(1, 13):
            sheet.cell(
                row_number,
                22 + month,
                "#N/A" if indicator == error_indicator else row_number * month,
            )

    if with_internal_prepared_range:
        internal = workbook.create_sheet("загрузка ERP расходы")
        internal.append(HEADERS)
        internal.append(
            [
                reporting_unit,
                "Административные расходы",
                "Административный департамент",
                "ТК",
                reporting_unit,
                "БЕЗ НДС",
                "Связь",
                "Интернет",
                1,
                *([0] * 11),
            ]
        )

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def bdr_formula_workbook_bytes(
    *,
    reporting_unit: str = "АЮ Отдел обеспечения",
    indicator: str = "Оборот в кг",
    month: int = 1,
    cached_value: int = 593845,
) -> bytes:
    """Whole-BDR fixture whose KPI month is a formula with a saved result."""

    raw = bdr_full_workbook_bytes(reporting_unit=reporting_unit)
    row_number = next(
        row for row, name in BDR_FULL_INDICATORS if name == indicator
    )
    coordinate = f"{_column_letter(22 + month)}{row_number}"
    return _replace_numeric_cell_with_cached_formula(
        raw,
        coordinate,
        row_number * month,
        f"SUM({cached_value},0)",
        cached_value,
    )


def bdr_split_kpi_value_workbook_bytes(
    *,
    reporting_unit: str = "АЮ Отдел обеспечения",
    cached_value: int = 593845,
) -> bytes:
    """Department-aware BDR plus one exact aggregate KPI value sheet."""

    workbook = Workbook()
    planning = workbook.active
    planning.title = "Строки БДР с отделом"
    planning.cell(2, 1, reporting_unit)
    planning.cell(2, 5, "Отдел")
    for month in range(1, 13):
        planning.cell(2, 9 + month, datetime(2026, month, 1))
        planning.cell(3, 9 + month, "план")
    for row_number, indicator in BDR_FULL_INDICATORS:
        planning.cell(row_number, 5, reporting_unit)
        planning.cell(row_number, 7, indicator)
        for month in range(1, 13):
            planning.cell(row_number, 9 + month, 0)

    summary = workbook.create_sheet("Сводные значения БДР")
    summary.cell(2, 1, "АЮ")
    for month in range(1, 13):
        summary.cell(2, 22 + month, datetime(2026, month, 1))
        summary.cell(3, 22 + month, "план")
    for row_number, indicator in BDR_FULL_INDICATORS:
        summary.cell(row_number, 7, indicator)
        for month in range(1, 13):
            summary.cell(row_number, 22 + month, row_number * month)
            summary.cell(row_number, 22 + month).number_format = "#,##0"
    # The owner summary contains a repeated expense total after the KPI tail.
    # It must not invalidate the exact expense-block anchor before EBITDA.
    summary.cell(42, 7, "Расходы по основной деятельности ИТОГО")
    for month in range(1, 13):
        summary.cell(42, 22 + month, 42 * month)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    raw = output.getvalue()
    row_number = next(
        row for row, name in BDR_FULL_INDICATORS if name == "Оборот в кг"
    )
    raw = _replace_numeric_cell_with_cached_formula(
        raw,
        f"W{row_number}",
        row_number,
        f"SUM({cached_value},0)",
        cached_value,
        sheet_name="xl/worksheets/sheet2.xml",
    )
    revenue_per_kg_row = next(
        row for row, name in BDR_FULL_INDICATORS if name == "Выручка за 1 кг"
    )
    return _replace_numeric_cell_with_cached_formula(
        raw,
        f"W{revenue_per_kg_row}",
        revenue_per_kg_row,
        "469.5283132972408",
        469.5283132972408,
        sheet_name="xl/worksheets/sheet2.xml",
    )


def _replace_numeric_cell_with_cached_formula(
    workbook_bytes: bytes,
    coordinate: str,
    original_value: int,
    formula: str,
    cached_value: int | float,
    *,
    sheet_name: str = "xl/worksheets/sheet1.xml",
) -> bytes:
    with ZipFile(BytesIO(workbook_bytes), "r") as source:
        members = {name: source.read(name) for name in source.namelist()}
    xml = members[sheet_name].decode("utf-8")
    pattern = re.compile(
        rf'<c r="{re.escape(coordinate)}"(?P<attrs>[^>]*)>'
        rf'<v>{re.escape(str(original_value))}</v></c>'
    )
    match = pattern.search(xml)
    if match is None:
        raise AssertionError(f"Fixture cell not found: {coordinate}")
    style = re.search(r'\s+s="[^"]+"', match.group("attrs"))
    style_attribute = style.group(0) if style else ""
    replacement = (
        f'<c r="{coordinate}"{style_attribute}>'
        f'<f>{formula}</f><v>{cached_value}</v></c>'
    )
    members[sheet_name] = pattern.sub(replacement, xml, count=1).encode("utf-8")
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as target:
        for name, data in members.items():
            target.writestr(name, data)
    return output.getvalue()


def _column_letter(column: int) -> str:
    letters = ""
    current = column
    while current:
        current, remainder = divmod(current - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def workbook_bytes(
    *,
    sheet_name: str = "Произвольное имя",
    two_candidates: bool = False,
    no_range: bool = False,
    monthly_error: bool = False,
    shared_error: bool = False,
    department_error: bool = False,
    cfo_error: bool = False,
    missing_mapping: bool = False,
    negative: bool = False,
    reporting_unit: str = "ПС",
    tax_error: bool = False,
    second_cfo: str = "ЦФО 2",
    first_department: str = "Департамент 1",
    first_cfo: str = "ЦФО 1",
) -> bytes:
    workbook = Workbook()
    first = workbook.active
    first.title = sheet_name
    if no_range:
        first.append(["Это не загрузочный диапазон", "Значение"])
        first.append(["x", 1])
    else:
        _append_candidate(
            first,
            monthly_error,
            shared_error,
            department_error,
            cfo_error,
            missing_mapping,
            negative,
            reporting_unit,
            tax_error,
            second_cfo,
            first_department,
            first_cfo,
        )
        if two_candidates:
            second = workbook.create_sheet("Второй диапазон")
            _append_candidate(
                second,
                False,
                False,
                False,
                False,
                False,
                False,
                reporting_unit,
                False,
                second_cfo,
                first_department,
                first_cfo,
            )
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def large_workbook_bytes(
    inert_rows: int = 25_000,
    inert_columns: int = 8,
) -> bytes:
    """Build a multi-MiB XLSX with only two processable business rows."""

    workbook = Workbook(write_only=True)
    business = workbook.create_sheet("Небольшой business range")
    _append_candidate(business, False, False, False, False, False, False, "ПС")

    inert = workbook.create_sheet("Большой inert synthetic лист")
    for row in range(inert_rows):
        values = []
        for column in range(inert_columns):
            digest = hashlib.blake2s(
                f"inert:{row}:{column}".encode(),
                digest_size=18,
            ).digest()
            values.append(base64.urlsafe_b64encode(digest).decode("ascii"))
        inert.append(values)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def protected_workbook_bytes(content: bytes, password: str) -> bytes:
    output = BytesIO()
    OOXMLFile(BytesIO(content)).encrypt(password, output)
    return output.getvalue()


def _append_candidate(
    sheet,
    monthly_error: bool,
    shared_error: bool,
    department_error: bool,
    cfo_error: bool,
    missing_mapping: bool,
    negative: bool,
    reporting_unit: str,
    tax_error: bool = False,
    second_cfo: str = "ЦФО 2",
    first_department: str = "Департамент 1",
    first_cfo: str = "ЦФО 1",
) -> None:
    sheet.append(["Синтетический fixture: вымышленные данные"])
    sheet.append(HEADERS)
    months_a = [0] * 12
    months_a[0] = -10 if negative else 100
    if monthly_error:
        months_a[4] = "#REF!"
    sheet.append(
        [
            reporting_unit,
            "Административные",
            "" if department_error else first_department,
            "ТК",
            "" if cfo_error else first_cfo,
            "?" if tax_error else 0.2,
            "Связь",
            "Интернет" if not missing_mapping else "Нет в ERP",
            *months_a,
        ]
    )
    sheet.append(
        [
            reporting_unit,
            "Коммерческие",
            "" if shared_error else "Департамент 2",
            "ТК",
            second_cfo,
            "БЕЗ НДС",
            "Маркетинг",
            "Реклама",
            *([0] * 12),
        ]
    )


def reference_bytes(kind: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Synthetic"
    if kind == "erp_articles":
        sheet.append(
            [
                "Код",
                "Официальное наименование",
                "Тип расходов",
                "Группа расходов",
                "Исходная статья",
            ]
        )
        sheet.append(["ERP-001", "Интернет ERP", "Административные", "Связь", "Интернет"])
        sheet.append(["ERP-002", "Реклама ERP", "Коммерческие", "Маркетинг", "Реклама"])
        sheet.append(["ERP-DEL", "Удалить (видимая запись)", "Административные", "Прочие", "Удалить"])
    elif kind == "organizations":
        sheet.append(["ID", "Код", "Наименование", "Родитель ID", "Полный путь"])
        sheet.append(["root", "ORG-1", "Группа", "", "Группа"])
        sheet.append(["ps", "ORG-2", "ПС", "root", "Группа → ПС"])
        sheet.append(["cfo", "ORG-3", "ЦФО 1", "ps", "Группа → ПС → ЦФО 1"])
        sheet.append(["del", "ORG-4", "!!!Удалить", "ps", "Группа → ПС → !!!Удалить"])
        sheet.append(["other", "ORG-5", "Сосед", "root", "Группа → Сосед"])
    elif kind == "scenarios":
        sheet.append(["Наименование", "Год", "ERP-код", "Комментарий"])
        sheet.append(["ПЛАН_2026", 2026, "00010", "synthetic"])
    elif kind == "intalev_cfos":
        sheet.append(["Код ЦФО Инталев", "Наименование ЦФО Инталев", "Полный путь ЦФО"])
        sheet.append(["INT-CFO-1", "ЦФО 1", "Инталев → ЦФО 1"])
        sheet.append(["INT-CFO-2", "ЦФО 2", "Инталев → ЦФО 2"])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def indicator_classifier_bytes(
    rows: list[dict[str, object]] | None = None,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Классификатор"
    sheet.append(
        [
            "ERP-код статьи",
            "Полный путь статьи",
            "Статья",
            "Показатель",
            "Канал сбыта",
        ]
    )
    source_rows = rows or [
        {
            "erp_code": "ERP-001",
            "article_path": "Административные → Связь → Интернет",
            "article_name": "Интернет",
            "indicator": "Услуги связи",
            "sales_channel": "Основной канал",
        },
        {
            "erp_code": "ERP-002",
            "article_path": "Коммерческие → Маркетинг → Реклама",
            "article_name": "Реклама",
            "indicator": "Маркетинговые расходы",
            "sales_channel": "Основной канал",
        },
    ]
    for row in source_rows:
        sheet.append(
            [
                row.get("erp_code", ""),
                row.get("article_path", ""),
                row.get("article_name", ""),
                row.get("indicator", ""),
                row.get("sales_channel", ""),
            ]
        )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def real_reference_bytes(kind: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист_1"

    if kind == "erp_articles":
        sheet.cell(1, 1, "Статья доходов и расходов")
        sheet.cell(1, 15, "Код")
        _hierarchy_cell(sheet, 2, "Административные", 0)
        _hierarchy_cell(sheet, 3, "Связь", 2)
        _hierarchy_cell(sheet, 4, "Интернет", 4)
        sheet.cell(5, 15, "ERP-001")
        _hierarchy_cell(sheet, 6, "Коммерческие", 0)
        _hierarchy_cell(sheet, 7, "Маркетинг", 2)
        _hierarchy_cell(sheet, 8, "Реклама", 4)
        sheet.cell(9, 15, "ERP-002")
        _hierarchy_cell(sheet, 10, "!!!Удалить", 0)
        _hierarchy_cell(sheet, 11, "Прочие", 2)
        _hierarchy_cell(sheet, 12, "Удалить", 4)
        sheet.cell(13, 15, "ERP-DEL")
    elif kind == "organizations":
        sheet.cell(1, 1, "Организации")
        sheet.cell(1, 39, "Код")
        rows = [
            (2, "Группа", "ORG-1", 0),
            (3, "ПС", "ORG-2", 1),
            (4, "ЦФО 1", "ORG-3", 2),
            (5, "!!!Удалить", "ORG-4", 2),
            (6, "Сосед", "ORG-5", 1),
        ]
        for row, name, code, level in rows:
            _hierarchy_cell(sheet, row, name, level)
            sheet.cell(row, 39, code)
    elif kind == "scenarios":
        sheet.cell(7, 1, "Сценарии")
        sheet.cell(7, 19, "Код")
        sheet.cell(8, 1, "ПЛАН 2026")
        sheet.cell(8, 19, "00010")
        sheet.cell(9, 1, "Факт")
        sheet.cell(9, 19, "00001")

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def erp_organization_hierarchy_bytes(
    *,
    cfo_code: str = "000000173",
    cfo_name: str = "АЮ Административный Отдел",
    source_department: str = "Административный департамент",
    organization_name: str = 'ООО "Айс Юнион"',
    organization_code: str = "000000001",
) -> bytes:
    """Synthetic structural analogue of the ERP organization hierarchy export."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист_1"
    sheet.cell(7, 1, "Организация")
    sheet.cell(8, 7, "Головная организация")
    sheet.cell(8, 32, "Верхний уровень иерархии")
    sheet.cell(8, 39, "Код")

    sheet.cell(9, 1, organization_name)
    sheet.cell(10, 1, cfo_name)
    sheet.cell(11, 1, source_department)
    sheet.cell(11, 7, cfo_name)
    sheet.cell(11, 32, organization_name)
    sheet.cell(11, 39, cfo_code)

    # The real export may list the coded organization element after its
    # subordinate CFO rows. The explicit parent column, not row order, is the
    # relationship authority for enrichment.
    sheet.cell(12, 1, organization_name)
    sheet.cell(13, 7, organization_name)
    sheet.cell(13, 32, "4 Владивосток")
    sheet.cell(13, 39, organization_code)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def revenue_quantity_workbook_bytes() -> bytes:
    """Synthetic prepared input covering all three structural indicator types."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ОПИУ synthetic"
    sheet.append(["Синтетический fixture: расходы, доходы и количества"])
    sheet.append(
        [
            *HEADERS,
            "Тип показателя",
            "Группа дохода",
            "Условия формулы",
            "Аналитики",
            "Номенклатура",
            "Единица измерения",
            "Контрагент",
            "ИНТ канал сбыта",
            "Сеть",
            "Регион продаж",
        ]
    )
    rows = [
        (
            [
                "ПС",
                "Административные",
                "Департамент 1",
                "ТК",
                "ЦФО 1",
                "20%",
                "Связь",
                "Интернет",
                100,
                *([0] * 11),
            ],
            ["EXPENSE", "", "", "", "", "", "", "", "", ""],
        ),
        (
            [
                "ПС",
                "Прочие доходы",
                "Департамент 2",
                "ТК",
                "ЦФО 2",
                "БЕЗ НДС",
                "Продажи",
                "Продажа товара",
                200,
                *([0] * 11),
            ],
            [
                "REVENUE",
                "Выручка от продаж",
                "Источник=Продажи",
                "Организационные единицы | ЦФО | ИНТ номенклатура",
                "Товар А",
                "",
                'ООО "Покупатель"',
                "Сети Федеральные",
                "Сеть 1",
                "Приморский край",
            ],
        ),
        (
            [
                "ПС",
                "Количественные показатели",
                "Департамент 2",
                "ТК",
                "ЦФО 2",
                "?",
                "Продукция",
                "Товар А",
                5,
                *([0] * 11),
            ],
            ["QUANTITY", "", "", "", "Товар А", "кг", "", "", "", ""],
        ),
    ]
    for business, indicators in rows:
        sheet.append([*business, *indicators])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def revenue_quantity_classifier_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Классификатор"
    sheet.append(
        [
            "ERP-код статьи",
            "Полный путь статьи",
            "Статья",
            "Показатель",
            "Канал сбыта",
            "Тип показателя",
            "Группа дохода",
            "Условия формулы",
            "Аналитики",
            "Номенклатура",
            "Единица измерения",
            "Контрагент",
            "ИНТ канал сбыта",
            "Сеть",
            "Регион продаж",
        ]
    )
    sheet.append(
        [
            "ERP-001",
            "Административные → Связь → Интернет",
            "Интернет",
            "Административные расходы",
            "Основной канал",
            "EXPENSE",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    sheet.append(
        [
            "",
            "",
            "Продажа товара",
            "Выручка",
            "Основной канал",
            "REVENUE",
            "Выручка от продаж",
            "Источник=Продажи",
            "Организационные единицы | ЦФО | ИНТ номенклатура",
            "Товар А",
            "",
            'ООО "Покупатель"',
            "Сети Федеральные",
            "Сеть 1",
            "Приморский край",
        ]
    )
    sheet.append(
        [
            "",
            "",
            "",
            "Количество продукции",
            "Основной канал",
            "QUANTITY",
            "",
            "",
            "",
            "Товар А",
            "кг",
            "",
            "",
            "",
            "",
        ]
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def intalev_opiu_bytes(
    *,
    sheet_name: str = "TDSheet",
    monthly_error: bool = False,
) -> bytes:
    """Build a fictional structural analogue of an Intalev annual OPIU export."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet["A1"] = "ОТЧЕТ 'ОТЧЕТ О ПРИБЫЛЯХ И УБЫТКАХ 2025'"
    sheet["A2"] = (
        "СЦЕНАРИЙ 1: Факт\n"
        "ПЕРИОДИЧНОСТЬ: месяц\n"
        "ЦФО: ЦД/ЦЗ Фонд развития"
    )
    sheet["A4"] = "Показатели"
    for month in range(1, 13):
        column = 4 + month
        sheet.cell(4, column, f"01.{month:02d}.2025 - 28.{month:02d}.2025")
        sheet.cell(5, column, "Факт")
        sheet.cell(6, column, "С НДС")

    _intalev_row(sheet, 7, "Расходы по основной деятельности ИТОГО", 0, [100] * 12)
    _intalev_row(sheet, 8, "Административные расходы", 2, [100] * 12)
    _intalev_row(sheet, 9, "Связь", 4, [100] * 12)
    first = [0] * 12
    first[0] = -10
    if monthly_error:
        first[4] = "#REF!"
    _intalev_row(sheet, 10, "Интернет", 6, first)
    _intalev_row(sheet, 11, "Телефония", 6, [0] * 12)
    _intalev_row(sheet, 12, "Маркетинг", 4, [0] * 12)
    _intalev_row(sheet, 13, "Реклама", 6, [0] * 12)
    _intalev_row(sheet, 14, "EBITDA", 0, [100] * 12)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _intalev_row(sheet, row: int, name: str, indent: int, months: list[object]) -> None:
    cell = sheet.cell(row, 1, name)
    cell.alignment = Alignment(indent=indent)
    sheet.row_dimensions[row].outlineLevel = indent // 2
    for month, value in enumerate(months, start=1):
        sheet.cell(row, 4 + month, value)


def _hierarchy_cell(sheet, row: int, value: str, level: int) -> None:
    cell = sheet.cell(row, 1, value)
    cell.alignment = Alignment(indent=level)


def write_cached_formula_fixture(path: Path) -> None:
    path.write_bytes(workbook_bytes())
    with ZipFile(path, "r") as source:
        members = {name: source.read(name) for name in source.namelist()}
    sheet_name = "xl/worksheets/sheet1.xml"
    xml = members[sheet_name].decode("utf-8")
    xml = xml.replace(
        '<c r="I3" t="n"><v>100</v></c>',
        '<c r="I3"><f>50+50</f><v>100</v></c>',
    )
    members[sheet_name] = xml.encode("utf-8")
    with ZipFile(path, "w", ZIP_DEFLATED) as target:
        for name, data in members.items():
            target.writestr(name, data)
