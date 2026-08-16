from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from excel_transform_1c.adapters.excel import detect_path, read_path
from excel_transform_1c.application.service import WorkflowService
from excel_transform_1c.core.detection import detect_candidate_ranges, read_source_rows
from excel_transform_1c.core.models import IndicatorType
from tests.helpers.workbooks import (
    bdr_formula_workbook_bytes,
    bdr_full_workbook_bytes,
    erp_organization_hierarchy_bytes,
)


DEPARTMENT = "АЮ Отдел обеспечения"
PARENT_DEPARTMENT = "Департамент обеспечения"
ORGANIZATION = "4 Владивосток"
ORGANIZATION_CODE = "000000041"
CFO_CODE = "000000175"


def _configured_service(tmp_path) -> WorkflowService:
    service = WorkflowService(tmp_path / "runtime")
    service.upload_reference(
        "organizations",
        erp_organization_hierarchy_bytes(
            cfo_code=CFO_CODE,
            cfo_name=DEPARTMENT,
            source_department=PARENT_DEPARTMENT,
            organization_name=ORGANIZATION,
            organization_code=ORGANIZATION_CODE,
        ),
    )
    return service


def _process_formula_bdr(tmp_path):
    service = _configured_service(tmp_path)
    scenario = next(
        item for item in service.store.list_scenarios() if item.name == "ПЛАН 2026"
    )
    context = service.build_context(
        DEPARTMENT,
        ORGANIZATION_CODE,
        scenario.scenario_id,
        2026,
        [],
    )
    pending = service.prepare_upload(
        "БДР 2026 ИТОГ.xlsx",
        bdr_formula_workbook_bytes(reporting_unit=DEPARTMENT),
        context,
    )
    run = service.process_upload(
        pending.upload_id,
        pending.candidates[0].candidate_id,
    )
    return service, run


def test_kpi_context_from_department() -> None:
    workbook = load_workbook(
        BytesIO(bdr_full_workbook_bytes(reporting_unit=DEPARTMENT)),
        data_only=True,
    )
    try:
        candidate = detect_candidate_ranges(workbook)[0]
        row = next(
            item
            for item in read_source_rows(workbook, candidate, "БДР 2026 ИТОГ.xlsx")
            if item.indicator_type == IndicatorType.KPI.value
            and item.article == "Оборот в кг"
        )
        assert candidate.columns["department"] == 5
        assert row.department == DEPARTMENT
        assert row.cfo == DEPARTMENT
        assert row.cells["department"] == "E10"
    finally:
        workbook.close()


def test_kpi_cfo_resolution(tmp_path) -> None:
    service, run = _process_formula_bdr(tmp_path)
    result = next(
        item
        for item in service.kpi_results(run.run_id)
        if item.indicator_name == "Оборот в кг" and item.period == "01.2026"
    )

    assert result.organization == ORGANIZATION
    assert result.organization_code == ORGANIZATION_CODE
    assert result.department == DEPARTMENT
    assert result.department_name == PARENT_DEPARTMENT
    assert result.cfo == DEPARTMENT
    assert result.cfo_code == CFO_CODE


def test_kpi_month_value_from_formula_cell(tmp_path) -> None:
    path = tmp_path / "БДР 2026 ИТОГ.xlsx"
    path.write_bytes(bdr_formula_workbook_bytes(reporting_unit=DEPARTMENT))
    candidate = detect_path(path)[0]
    row = next(
        item
        for item in read_path(path, candidate, path.name)
        if item.indicator_type == IndicatorType.KPI.value
        and item.article == "Оборот в кг"
    )

    assert row.months[0] == Decimal("593845")
    assert not str(row.months[0]).startswith("=")
    assert row.cells["month_1"] == "W10"


def test_kpi_export_mapping(tmp_path) -> None:
    service, run = _process_formula_bdr(tmp_path)
    exported = load_workbook(BytesIO(service.export_run(run.run_id)), data_only=False)
    try:
        for sheet_name in ("OPIU Light", "ОПИУ", "Показатели"):
            sheet = exported[sheet_name]
            headers = {cell.value: cell.column for cell in sheet[1]}
            for required in (
                "Организация",
                "Код организации",
                "Департамент",
                "Отдел",
                "ЦФО",
                "Код ЦФО",
                "Тип показателя",
                "Показатель",
                "Период",
                "Значение",
            ):
                assert required in headers
            row = next(
                cells
                for cells in sheet.iter_rows(min_row=2)
                if cells[headers["Тип показателя"] - 1].value == "KPI"
                and cells[headers["Показатель"] - 1].value == "Оборот в кг"
                and cells[headers["Период"] - 1].value == "01.2026"
            )
            assert row[headers["Организация"] - 1].value == ORGANIZATION
            assert row[headers["Код организации"] - 1].value == ORGANIZATION_CODE
            assert row[headers["Департамент"] - 1].value == PARENT_DEPARTMENT
            assert row[headers["Отдел"] - 1].value == DEPARTMENT
            assert row[headers["ЦФО"] - 1].value == DEPARTMENT
            assert row[headers["Код ЦФО"] - 1].value == CFO_CODE
            value_cell = row[headers["Значение"] - 1]
            assert value_cell.value == 593845
            assert value_cell.data_type == "n"
    finally:
        exported.close()
