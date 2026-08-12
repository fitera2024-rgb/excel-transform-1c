from io import BytesIO

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment

from excel_transform_1c.adapters.references import erp_articles, parse_reference_workbook
from excel_transform_1c.core.transform import ExactERPMapper


def _workbook() -> tuple[Workbook, object]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Синтетическая структура"
    sheet.cell(1, 1, "Иерархия статей доходов и расходов")
    sheet.cell(1, 15, "Код элемента")
    return workbook, sheet


def _hierarchy(
    sheet: object,
    row: int,
    name: str,
    *,
    indent: int = 0,
    outline: int = 0,
) -> None:
    cell = sheet.cell(row, 1, name)
    cell.alignment = Alignment(indent=indent)
    sheet.row_dimensions[row].outlineLevel = outline


def _code(sheet: object, row: int, code: str, analytics: str | None = None) -> None:
    if analytics is not None:
        sheet.cell(row, 1, analytics)
    sheet.cell(row, 15, code)


def _parse(workbook: Workbook) -> list[dict[str, str]]:
    output = BytesIO()
    workbook.save(output)
    return parse_reference_workbook(output.getvalue(), "erp_articles")


def test_code_row_analytics_does_not_replace_preceding_official_article() -> None:
    workbook, sheet = _workbook()
    _hierarchy(sheet, 2, "Синтетический тип", indent=0)
    _hierarchy(sheet, 3, "Группа поездок", indent=2, outline=1)
    _hierarchy(sheet, 4, "Официальная статья проживания", indent=4, outline=2)
    _code(sheet, 5, "ЦБ-000239", "Счет затрат 26")

    _hierarchy(sheet, 6, "Соседняя группа", indent=2, outline=1)
    _hierarchy(sheet, 7, "Соседняя статья", indent=4, outline=2)
    _code(sheet, 8, "SYN-002", "Техническая аналитика")

    parsed = {item["code"]: item for item in _parse(workbook)}

    assert parsed["ЦБ-000239"] == {
        "code": "ЦБ-000239",
        "name": "Официальная статья проживания",
        "expense_type": "Синтетический тип",
        "expense_group": "Группа поездок",
        "source_article": "Официальная статья проживания",
    }
    assert parsed["SYN-002"]["expense_group"] == "Соседняя группа"
    assert parsed["SYN-002"]["source_article"] == "Соседняя статья"


def test_outline_hierarchy_keeps_neighboring_groups_as_siblings() -> None:
    workbook, sheet = _workbook()
    _hierarchy(sheet, 2, "Синтетическая логистика", outline=0)
    _hierarchy(sheet, 3, "Первая группа оборудования", outline=1)
    _hierarchy(sheet, 4, "Первая статья", outline=2)
    _code(sheet, 5, "SYN-001")
    _hierarchy(sheet, 6, "Содержание оборудования", outline=1)
    _hierarchy(sheet, 7, "Синтетическая установка", outline=2)
    _code(sheet, 8, "00-000150")

    parsed = {item["code"]: item for item in _parse(workbook)}

    assert parsed["SYN-001"]["expense_group"] == "Первая группа оборудования"
    assert parsed["00-000150"] == {
        "code": "00-000150",
        "name": "Синтетическая установка",
        "expense_type": "Синтетическая логистика",
        "expense_group": "Содержание оборудования",
        "source_article": "Синтетическая установка",
    }


def test_delete_marked_branches_and_duplicate_paths_remain_visible() -> None:
    workbook, sheet = _workbook()
    _hierarchy(sheet, 2, "!!!Удалить", indent=0)
    _hierarchy(sheet, 3, "Удалить", indent=2)
    _hierarchy(sheet, 4, "Видимая статья", indent=4)
    _code(sheet, 5, "DEL-001")
    _hierarchy(sheet, 6, "Видимая статья", indent=4)
    _code(sheet, 7, "DEL-002")

    parsed = _parse(workbook)

    assert [item["code"] for item in parsed] == ["DEL-001", "DEL-002"]
    assert {
        (item["expense_type"], item["expense_group"], item["source_article"])
        for item in parsed
    } == {("!!!Удалить", "Удалить", "Видимая статья")}
    mapped, reason = ExactERPMapper(erp_articles(parsed)).resolve(
        "!!!Удалить", "Удалить", "Видимая статья"
    )
    assert mapped is None
    assert "нескольким ERP-кодам" in reason


def test_scaled_fixture_preserves_271_unique_erp_codes() -> None:
    workbook, sheet = _workbook()
    _hierarchy(sheet, 2, "Синтетический тип", indent=0)
    _hierarchy(sheet, 3, "Синтетическая группа", indent=2)

    row = 4
    for index in range(271):
        _hierarchy(sheet, row, f"Синтетическая статья {index:03d}", indent=4)
        _code(sheet, row + 1, f"SYN-{index:03d}", f"Аналитика {index:03d}")
        row += 2

    parsed = _parse(workbook)

    assert len(parsed) == 271
    assert len({item["code"] for item in parsed}) == 271
    assert parsed[0]["source_article"] == "Синтетическая статья 000"
    assert parsed[-1]["source_article"] == "Синтетическая статья 270"


@pytest.mark.parametrize(
    ("indent", "outline", "message"),
    [
        (1, 0, "неподдерживаемый отступ"),
        (4, 1, "противоречивые уровни"),
    ],
)
def test_unsupported_or_ambiguous_indent_fails_visibly(
    indent: int,
    outline: int,
    message: str,
) -> None:
    workbook, sheet = _workbook()
    _hierarchy(sheet, 2, "Синтетический тип", indent=0)
    _hierarchy(sheet, 3, "Неоднозначная группа", indent=indent, outline=outline)
    _code(sheet, 4, "SYN-001")

    with pytest.raises(ValueError, match=message):
        _parse(workbook)


def test_code_without_preceding_hierarchy_fails_visibly() -> None:
    workbook, sheet = _workbook()
    _code(sheet, 2, "SYN-001", "Техническая аналитика")

    with pytest.raises(ValueError, match="не имеет предшествующего узла"):
        _parse(workbook)
