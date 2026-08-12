from decimal import Decimal

from openpyxl import load_workbook
from io import BytesIO

from excel_transform_1c.adapters.persistence import LocalStore
from excel_transform_1c.core.access import effective_organization_nodes
from excel_transform_1c.core.detection import detect_candidate_ranges, read_source_rows
from excel_transform_1c.core.models import ERPArticle, OrganizationNode, RunContext, STATUS_ATTENTION, STATUS_OK
from excel_transform_1c.core.transform import ExactERPMapper, manual_mapping_key, normalize_tax, transform_rows
from tests.helpers.workbooks import workbook_bytes


def context(confirmed: bool = True) -> RunContext:
    return RunContext("ПС", "ps", "Группа → ПС", "s1", "ПЛАН 2026", 2026, confirmed, 2026)


def articles() -> list[ERPArticle]:
    return [
        ERPArticle("ERP-001", "Интернет ERP", "Административные", "Связь", "Интернет"),
        ERPArticle("ERP-002", "Реклама ERP", "Коммерческие", "Маркетинг", "Реклама"),
    ]


def source_rows(**kwargs):
    workbook = load_workbook(BytesIO(workbook_bytes(**kwargs)), data_only=True)
    candidate = detect_candidate_ranges(workbook)[0]
    return read_source_rows(workbook, candidate, "synthetic.xlsx")


def test_structural_schema_detection_ignores_sheet_name_and_finds_two_ranges():
    workbook = load_workbook(BytesIO(workbook_bytes(sheet_name="Никакой загрузки", two_candidates=True)), data_only=True)
    candidates = detect_candidate_ranges(workbook)
    assert [item.sheet for item in candidates] == ["Никакой загрузки", "Второй диапазон"]


def test_two_rows_normalize_to_24_records_including_zero():
    records, issues = transform_rows(source_rows(), context(), ExactERPMapper(articles()))
    assert len(records) == 24
    assert len([record for record in records if record.amount == Decimal("0")]) == 23
    assert not issues


def test_monthly_error_skips_only_one_month():
    records, issues = transform_rows(source_rows(monthly_error=True), context(), ExactERPMapper(articles()))
    assert len(records) == 23
    assert len([record for record in records if record.source_row == 3]) == 11
    assert [(issue.pointer.row, issue.pointer.month) for issue in issues] == [(3, 5)]


def test_shared_field_error_keeps_months_with_attention():
    records, issues = transform_rows(source_rows(shared_error=True), context(), ExactERPMapper(articles()))
    affected = [record for record in records if record.source_row == 4]
    assert len(affected) == 12
    assert all(record.status == STATUS_ATTENTION for record in affected)
    assert any(issue.kind == "shared-field" for issue in issues)


def test_tax_normalization_is_exact_and_zero_is_attention():
    assert normalize_tax(0.2) == ("20%", None)
    assert normalize_tax(0.22) == ("22%", None)
    assert normalize_tax(0)[1]
    assert normalize_tax("?")[1]
    assert normalize_tax("20%") == ("20%", None)


def test_negative_amount_is_preserved_with_attention():
    records, _ = transform_rows(source_rows(negative=True), context(), ExactERPMapper(articles()))
    negative = next(record for record in records if record.amount < 0)
    assert negative.amount == Decimal("-10")
    assert negative.status == STATUS_ATTENTION
    assert "Отрицательная сумма" in negative.comment


def test_exact_mapping_requires_unique_case_sensitive_full_path_and_conflict_blocks_saved_mapping():
    mapper = ExactERPMapper(articles())
    assert mapper.resolve("Административные", "Связь", "Интернет")[0].code == "ERP-001"
    assert mapper.resolve("административные", "Связь", "Интернет")[0] is None
    duplicated = articles() + [ERPArticle("ERP-X", "Duplicate", "Административные", "Связь", "Интернет")]
    assert "нескольким" in ExactERPMapper(duplicated).resolve("Административные", "Связь", "Интернет")[1]
    saved = {manual_mapping_key("Административные", "Связь", "Нет в ERP"): "ERP-002"}
    mapped, reason = ExactERPMapper(articles(), saved).resolve("Административные", "Связь", "Нет в ERP")
    assert mapped.code == "ERP-002" and reason is None
    missing, reason = ExactERPMapper([], saved).resolve("Административные", "Связь", "Нет в ERP")
    assert missing is None and "конфликтует" in reason


def test_manual_mapping_key_excludes_organization_department_and_cfo():
    assert manual_mapping_key("Тип", "Группа", "Статья") == (
        "ОтчетОПрибыляхИУбытках",
        "Тип",
        "Группа",
        "Статья",
    )


def test_scenario_alias_has_stable_local_identity(tmp_path):
    store = LocalStore(tmp_path / "local.db")
    first = store.add_scenario("ПЛАН_2026", 2026, erp_code="00010", erp_confirmed=True)
    second = store.add_scenario("ПЛАН 2026", 2026)
    assert first.scenario_id == second.scenario_id
    assert first.name == "ПЛАН 2026"


def test_organization_subtree_union_deduplicates_overlap_and_keeps_delete_node():
    nodes = [
        OrganizationNode("root", "1", "Root", None, "Root"),
        OrganizationNode("a", "2", "A", "root", "Root → A"),
        OrganizationNode("child", "3", "Child", "a", "Root → A → Child"),
        OrganizationNode("delete", "4", "!!!Удалить", "a", "Root → A → !!!Удалить"),
        OrganizationNode("b", "5", "B", "root", "Root → B"),
    ]
    allowed = effective_organization_nodes(nodes, ["a", "child"])
    assert [node.node_id for node in allowed] == ["a", "child", "delete"]


def test_unconfirmed_scenario_marks_records_attention():
    records, _ = transform_rows(source_rows(), context(False), ExactERPMapper(articles()))
    assert all(record.status == STATUS_ATTENTION for record in records)
    assert all("не подтверждён" in record.comment for record in records)
