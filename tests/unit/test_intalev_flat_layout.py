from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from excel_transform_1c.core.detection import detect_candidate_ranges, read_source_rows


def _period_header(month: int, year: int = 2025) -> str:
    # The parser intentionally accepts the same month at both period ends.
    return f"01.{month:02d}.{year} - 28.{month:02d}.{year}"


def _flat_intalev_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ОПИУ факт 2025"
    sheet["A1"] = "ОТЧЕТ О ПРИБЫЛЯХ И УБЫТКАХ"
    sheet["C4"] = "Показатели"
    for month in range(1, 13):
        sheet.cell(4, 3 + month, _period_header(month))

    def row(
        number: int,
        reporting_unit: str,
        expense_type: str,
        article: str,
        *,
        bold: bool = False,
        amount: int = 0,
    ) -> None:
        sheet.cell(number, 1, reporting_unit)
        sheet.cell(number, 2, expense_type)
        cell = sheet.cell(number, 3, article)
        cell.font = Font(bold=bold)
        for month in range(1, 13):
            sheet.cell(number, 3 + month, amount + month)

    row(7, "ПС", "Административные расходы", "Административные расходы", bold=True)
    row(8, "ПС", "Административные расходы", "% расходов")
    row(9, "ПС", "Административные расходы", "Связь", bold=True)
    row(10, "ПС", "Административные расходы", "Интернет", amount=10)
    row(11, "ПС", "Административные расходы", "Телефония", amount=20)
    row(12, "ПС", "Административные расходы", "Банковская комиссия", bold=True, amount=30)

    row(13, "ПС", "Коммерческие расходы", "Коммерческие расходы", bold=True)
    row(14, "ПС", "Коммерческие расходы", "% расходов")
    row(15, "ПС", "Коммерческие расходы", "Маркетинг", bold=True)
    row(16, "ПС", "Коммерческие расходы", "Реклама", amount=40)

    # A technical result shares the same 12-month table but no longer repeats
    # the current expense type in the parent column and must not be imported.
    row(17, "ПС", "Прибыль", "EBITDA", bold=True, amount=50)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _generic_metrics_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ТЭУ,ПРР,хранение"
    sheet["C4"] = "Показатели"
    for month in range(1, 13):
        sheet.cell(4, 3 + month, _period_header(month))
    metrics = ("Объём", "Тариф", "Средняя цена", "Количество рейсов")
    for offset, metric in enumerate(metrics, start=7):
        sheet.cell(offset, 3, metric)
        for month in range(1, 13):
            sheet.cell(offset, 3 + month, month)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_flat_intalev_layout_is_detected_and_preserves_business_hierarchy():
    workbook = load_workbook(BytesIO(_flat_intalev_bytes()), data_only=True)
    try:
        candidates = detect_candidate_ranges(workbook)
        assert len(candidates) == 1
        candidate = candidates[0]
        assert candidate.source_kind == "intalev_opiu"
        assert candidate.source_year == 2025

        rows = read_source_rows(workbook, candidate, "synthetic-flat-intalev.xlsx")
        assert [row.article for row in rows] == [
            "Интернет",
            "Телефония",
            "Банковская комиссия",
            "Реклама",
        ]
        assert [row.expense_type for row in rows] == [
            "Административные расходы",
            "Административные расходы",
            "Административные расходы",
            "Коммерческие расходы",
        ]
        assert [row.expense_group for row in rows] == [
            "Связь",
            "Связь",
            "Банковская комиссия",
            "Маркетинг",
        ]
        assert {row.reporting_unit for row in rows} == {"ПС"}
        assert all(len(row.months) == 12 for row in rows)
        assert "EBITDA" not in {row.article for row in rows}
    finally:
        workbook.close()


def test_generic_twelve_month_metrics_are_not_false_positive_intalev_opiu():
    workbook = load_workbook(BytesIO(_generic_metrics_bytes()), data_only=True)
    try:
        assert detect_candidate_ranges(workbook) == []
    finally:
        workbook.close()
