from io import BytesIO

from openpyxl import Workbook

from excel_transform_1c.adapters.references import (
    article_indicator_rules,
    parse_reference_workbook,
)
from excel_transform_1c.core.models import IndicatorType


def _classifier_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Статья",
            "Показатель",
            "Канал сбыта",
            "Тип показателя",
            "Группа дохода",
            "Условия формулы",
            "Контрагент",
            "ИНТ канал сбыта",
            "Сеть",
            "Регион продаж",
            "Номенклатура",
        ]
    )
    common = [
        "Продажа товара",
        "Выручка",
        "Сети Федеральные",
        "REVENUE",
        "Выручка от продаж",
        "Источник=Продажи",
        'ООО "Покупатель"',
        "Сети Федеральные",
        "Сеть 1",
    ]
    sheet.append([*common, "Приморский край", "Товар А"])
    sheet.append([*common, "Хабаровский край", "Товар А"])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_revenue_rule_import_preserves_separate_exact_analytics() -> None:
    payload = parse_reference_workbook(_classifier_bytes(), "article_indicators")
    rules = article_indicator_rules(payload)

    assert len(rules) == 2
    assert {rule.sales_region for rule in rules} == {
        "Приморский край",
        "Хабаровский край",
    }
    first = rules[0]
    assert first.indicator_type == IndicatorType.REVENUE
    assert first.analytics == ""
    assert first.counterparty == 'ООО "Покупатель"'
    assert first.input_sales_channel == "Сети Федеральные"
    assert first.sales_network == "Сеть 1"
    assert first.nomenclature == "Товар А"


def test_expense_classifier_does_not_require_sales_channel_column() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["ERP-код статьи", "Статья", "Показатель", "Тип показателя"])
    sheet.append(["ERP-001", "Интернет", "Услуги связи", "EXPENSE"])
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    payload = parse_reference_workbook(output.getvalue(), "article_indicators")

    assert payload[0]["sales_channel"] == ""


def test_revenue_rule_import_builds_full_path_without_formula_column() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Тип доходов",
            "Группа доходов",
            "Статья",
            "Показатель",
            "Тип показателя",
            "Аналитика",
        ]
    )
    sheet.append(
        [
            "Доходы от продаж",
            "Розничные продажи",
            "Продажа товара",
            "Выручка",
            "REVENUE",
            "Организационные единицы",
        ]
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    payload = parse_reference_workbook(output.getvalue(), "article_indicators")

    assert payload[0]["article_path"] == (
        "Доходы от продаж → Розничные продажи → Продажа товара"
    )
    assert payload[0]["formula_condition"] == ""
