from io import BytesIO

from openpyxl import Workbook, load_workbook

from excel_transform_1c.core.detection import detect_candidate_ranges, read_source_rows
from excel_transform_1c.core.models import ERPArticle, IndicatorType, RunContext
from excel_transform_1c.core.transform import ExactERPMapper, transform_rows
from tests.helpers.workbooks import HEADERS, workbook_bytes


def _budget_with_revenue_analytics(
    sales_channel_header: str = "ИНТ канал сбыта",
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Произвольный бюджет"
    sheet.append(
        [
            *HEADERS,
            "Контрагент",
            sales_channel_header,
            "Сеть",
            "Регион продаж",
        ]
    )
    sheet.append(
        [
            "ПС",
            "Доходы",
            "Продажи",
            "ТК",
            "ЦФО продаж",
            "БЕЗ НДС",
            "Выручка",
            "Продажа товара",
            100,
            *([0] * 11),
            'ООО "Покупатель"',
            "Сети Федеральные",
            "Сеть А",
            "Приморский край",
        ]
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _prepared_income_range() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    # The title is intentionally unrelated to income: the schema is authority.
    sheet.title = "Произвольный диапазон"
    sheet.append(
        [
            "ТИП ДОХОДОВ",
            "ГРУППА ДОХОДОВ",
            "СТАТЬЯ",
            "АНАЛИТИКА",
            *[name.upper() for name in (
                "Январь",
                "Февраль",
                "Март",
                "Апрель",
                "Май",
                "Июнь",
                "Июль",
                "Август",
                "Сентябрь",
                "Октябрь",
                "Ноябрь",
                "Декабрь",
            )],
        ]
    )
    common = ["Продажи", "Выручка_продажи внешние", "HoReCa"]
    sheet.append([*common, "Аби Продакт", 100, *([0] * 11)])
    sheet.append([*common, "", 0, *([0] * 11)])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _prepared_income_range_with_context() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Доход с доказанным контекстом"
    sheet.append(
        [
            "ПОДРАЗДЕЛЕНИЕ (ЦФО 1)",
            "ТИП ДОХОДОВ",
            "ДЕПАРТАМЕНТ (ЦФО 2)",
            "Вид организации",
            "ОТДЕЛ",
            "НАЛОГООБЛОЖЕНИЕ",
            "ГРУППА ДОХОДОВ",
            "СТАТЬЯ",
            "АНАЛИТИКА",
            *[name.upper() for name in (
                "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
            )],
        ]
    )
    sheet.append(
        [
            "АЮ",
            "Прочие доходы по основной деятельности",
            "Финансовый департамент",
            "ХК",
            "ХП Финансовый Отдел",
            "НДС 22%",
            "Услуги предоставленные внутрихолдинговые",
            "Услуги аренды",
            None,
            100,
            *([0] * 11),
        ]
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _context() -> RunContext:
    return RunContext(
        reporting_unit="ПС",
        organization_node_id="ps",
        organization_name="Группа → ПС",
        scenario_id="plan-2026",
        scenario_name="ПЛАН 2026",
        scenario_year=2026,
        scenario_erp_confirmed=True,
        year=2026,
    )


def test_prepared_budget_reads_exact_revenue_analytics_headers() -> None:
    workbook = load_workbook(BytesIO(_budget_with_revenue_analytics()), data_only=True)

    candidate = detect_candidate_ranges(workbook)[0]
    rows = read_source_rows(workbook, candidate, "synthetic-revenue.xlsx")

    assert len(rows) == 1
    row = rows[0]
    assert row.counterparty == 'ООО "Покупатель"'
    assert row.input_sales_channel == "Сети Федеральные"
    assert row.sales_network == "Сеть А"
    assert row.sales_region == "Приморский край"
    assert row.cells["counterparty"] == "U2"
    assert row.cells["input_sales_channel"] == "V2"
    assert row.cells["sales_network"] == "W2"
    assert row.cells["sales_region"] == "X2"


def test_prepared_income_range_uses_schema_not_sheet_name() -> None:
    workbook = load_workbook(BytesIO(_prepared_income_range()), data_only=True)

    candidates = detect_candidate_ranges(workbook)
    rows = read_source_rows(workbook, candidates[0], "synthetic-income.xlsx")

    assert len(candidates) == 1
    assert candidates[0].source_kind == "prepared_income_budget"
    assert len(rows) == 2
    assert {row.expense_group for row in rows} == {
        "Выручка_продажи внешние"
    }
    assert {row.expense_type for row in rows} == {"Продажи"}
    assert {row.revenue_group for row in rows} == {"Выручка_продажи внешние"}
    assert {row.article for row in rows} == {"HoReCa"}
    assert [row.analytics for row in rows] == ["Аби Продакт", None]
    assert all(len(row.months) == 12 for row in rows)
    assert all(row.indicator_type == "REVENUE" for row in rows)


def test_prepared_income_preserves_exact_optional_source_context() -> None:
    workbook = load_workbook(
        BytesIO(_prepared_income_range_with_context()), data_only=True
    )
    candidate = detect_candidate_ranges(workbook)[0]
    row = read_source_rows(workbook, candidate, "income-context.xlsx")[0]

    assert row.reporting_unit == "АЮ"
    assert row.department == "Финансовый департамент"
    assert row.organization_type == "ХК"
    assert row.cfo == "ХП Финансовый Отдел"
    assert row.tax == "НДС 22%"
    assert row.cells["department"] == "C2"
    assert row.cells["cfo"] == "E2"


def test_prepared_income_range_transforms_without_expense_only_fields() -> None:
    workbook = load_workbook(BytesIO(_prepared_income_range()), data_only=True)
    candidate = detect_candidate_ranges(workbook)[0]
    rows = read_source_rows(workbook, candidate, "synthetic-income.xlsx")

    records, issues = transform_rows(rows, _context(), ExactERPMapper([]))

    assert not issues
    assert len(records) == 24
    assert {record.indicator_type for record in records} == {IndicatorType.REVENUE}
    assert {record.tax for record in records} == {""}
    assert {record.analytics for record in records} == {"Аби Продакт", ""}


def test_plain_sales_channel_alias_is_read_exactly() -> None:
    workbook = load_workbook(
        BytesIO(_budget_with_revenue_analytics("Канал сбыта")),
        data_only=True,
    )

    candidate = detect_candidate_ranges(workbook)[0]
    row = read_source_rows(workbook, candidate, "synthetic-revenue.xlsx")[0]

    assert row.input_sales_channel == "Сети Федеральные"


def test_unlisted_sales_channel_header_is_not_guessed() -> None:
    workbook = load_workbook(
        BytesIO(_budget_with_revenue_analytics("Канал продаж")),
        data_only=True,
    )

    candidate = detect_candidate_ranges(workbook)[0]
    row = read_source_rows(workbook, candidate, "synthetic-revenue.xlsx")[0]

    assert "input_sales_channel" not in candidate.columns
    assert row.input_sales_channel == ""


def test_revenue_analytics_propagate_to_all_preview_months() -> None:
    workbook = load_workbook(BytesIO(_budget_with_revenue_analytics()), data_only=True)
    candidate = detect_candidate_ranges(workbook)[0]
    rows = read_source_rows(workbook, candidate, "synthetic-revenue.xlsx")
    mapper = ExactERPMapper(
        [
            ERPArticle(
                "REV-001",
                "Продажа товара",
                "Доходы",
                "Выручка",
                "Продажа товара",
            )
        ]
    )

    records, issues = transform_rows(rows, _context(), mapper)

    assert not issues
    assert len(records) == 12
    assert {record.counterparty for record in records} == {'ООО "Покупатель"'}
    assert {record.input_sales_channel for record in records} == {"Сети Федеральные"}
    assert {record.sales_network for record in records} == {"Сеть А"}
    assert {record.sales_region for record in records} == {"Приморский край"}
    assert {record.indicator_type for record in records} == {IndicatorType.REVENUE}
    assert all(record.sales_channel == "" for record in records)
    assert records[0].pointers["counterparty"].cell == "U2"
    assert records[0].pointers["input_sales_channel"].cell == "V2"


def test_single_input_sales_channel_classifies_revenue() -> None:
    workbook = load_workbook(BytesIO(_budget_with_revenue_analytics()), data_only=True)
    sheet = workbook["Произвольный бюджет"]
    sheet["U2"] = None
    sheet["W2"] = None
    sheet["X2"] = None
    candidate = detect_candidate_ranges(workbook)[0]
    rows = read_source_rows(workbook, candidate, "synthetic-revenue.xlsx")
    mapper = ExactERPMapper(
        [
            ERPArticle(
                "REV-001",
                "Продажа товара",
                "Доходы",
                "Выручка",
                "Продажа товара",
            )
        ]
    )

    records, _ = transform_rows(rows, _context(), mapper)

    assert {record.input_sales_channel for record in records} == {"Сети Федеральные"}
    assert {record.indicator_type for record in records} == {IndicatorType.REVENUE}


def test_legacy_expense_budget_keeps_revenue_analytics_optional() -> None:
    workbook = load_workbook(BytesIO(workbook_bytes()), data_only=True)

    candidate = detect_candidate_ranges(workbook)[0]
    rows = read_source_rows(workbook, candidate, "legacy-expense.xlsx")

    assert len(rows) == 2
    assert all(row.counterparty == "" for row in rows)
    assert all(row.input_sales_channel == "" for row in rows)
    assert all(row.sales_network == "" for row in rows)
    assert all(row.sales_region == "" for row in rows)
