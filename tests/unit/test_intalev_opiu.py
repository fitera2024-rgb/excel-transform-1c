from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

from excel_transform_1c.core.detection import detect_candidate_ranges, read_source_rows
from excel_transform_1c.core.models import ERPArticle, RunContext, STATUS_ATTENTION, STATUS_SKIPPED
from excel_transform_1c.core.transform import ExactERPMapper, transform_rows
from tests.helpers.workbooks import intalev_opiu_bytes


def _context() -> RunContext:
    return RunContext(
        "ПС",
        "ps",
        "Группа → ПС",
        "fact",
        "Факт",
        2025,
        True,
        2025,
    )


def _articles() -> list[ERPArticle]:
    return [
        ERPArticle("ERP-001", "Интернет ERP", "Административные расходы", "Связь", "Интернет"),
        ERPArticle("ERP-002", "Телефония ERP", "Административные расходы", "Связь", "Телефония"),
        ERPArticle("ERP-003", "Реклама ERP", "Административные расходы", "Маркетинг", "Реклама"),
    ]


def _rows(*, sheet_name: str = "TDSheet", monthly_error: bool = False):
    workbook = load_workbook(
        BytesIO(intalev_opiu_bytes(sheet_name=sheet_name, monthly_error=monthly_error)),
        data_only=True,
    )
    candidate = detect_candidate_ranges(workbook)[0]
    return workbook, candidate, read_source_rows(workbook, candidate, "synthetic-intalev.xlsx")


def test_structural_candidate_and_period_detection_ignore_sheet_name():
    workbook, candidate, rows = _rows(sheet_name="Совсем другое имя")
    try:
        assert candidate.sheet == "Совсем другое имя"
        assert candidate.source_kind == "intalev_opiu"
        assert candidate.header_row == 4
        assert candidate.first_data_row == 7
        assert candidate.last_data_row == 14
        assert candidate.source_year == 2025
        assert [candidate.columns[f"month_{month}"] for month in range(1, 13)] == list(
            range(5, 17)
        )
        assert len(rows) == 3
    finally:
        workbook.close()


def test_hierarchy_parsing_extracts_source_cfo_and_excludes_technical_totals():
    workbook, candidate, rows = _rows()
    try:
        assert candidate.source_cfo == "ЦД/ЦЗ Фонд развития"
        assert [(row.expense_type, row.expense_group, row.article) for row in rows] == [
            ("Административные расходы", "Связь", "Интернет"),
            ("Административные расходы", "Связь", "Телефония"),
            ("Административные расходы", "Маркетинг", "Реклама"),
        ]
        assert all(row.cfo == "ЦД/ЦЗ Фонд развития" for row in rows)
        assert not {"Расходы по основной деятельности ИТОГО", "EBITDA"}.intersection(
            row.article for row in rows
        )
    finally:
        workbook.close()


def test_maximum_completeness_preserves_zero_negative_and_monthly_error():
    workbook, _, rows = _rows(monthly_error=True)
    try:
        records, issues = transform_rows(rows, _context(), ExactERPMapper(_articles()))
    finally:
        workbook.close()

    assert len(records) == 36
    negative = next(record for record in records if record.amount == Decimal("-10"))
    assert negative.status == STATUS_ATTENTION
    assert "Отрицательная сумма" in negative.comment
    assert len([record for record in records if record.amount == Decimal("0")]) == 34
    skipped = next(record for record in records if record.source_row == 10 and record.month == 5)
    assert skipped.status == STATUS_SKIPPED
    assert skipped.amount is None
    assert "I10" in skipped.comment
    assert any(issue.kind == "monthly-error" and issue.pointer.cell == "I10" for issue in issues)


def test_structural_single_month_intalev_opiu_is_supported():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Любой лист"
    sheet["A2"] = "ЦФО: ЦМД Сахалин"
    sheet["A4"] = "Показатели"
    sheet["E4"] = "01.01.2025 - 31.01.2025"
    rows = (
        (7, "Расходы по основной деятельности ИТОГО", 0),
        (8, "Административные расходы", 2),
        (9, "Связь", 4),
        (10, "Интернет", 6),
        (11, "EBITDA", 0),
    )
    for row_number, name, indent in rows:
        sheet.cell(row_number, 1, name).alignment = Alignment(indent=indent)
        sheet.row_dimensions[row_number].outlineLevel = indent // 2
        sheet.cell(row_number, 5, 593845 if row_number == 10 else 0)

    candidates = detect_candidate_ranges(workbook)
    assert len(candidates) == 1
    candidate = candidates[0]
    assert candidate.source_kind == "intalev_opiu"
    assert candidate.source_year == 2025
    assert candidate.columns["month_1"] == 5
    assert "month_2" not in candidate.columns

    source_rows = read_source_rows(workbook, candidate, "monthly-intalev.xlsx")
    assert len(source_rows) == 1
    assert source_rows[0].cfo == "ЦМД Сахалин"
    assert source_rows[0].months[0] == 593845
    assert source_rows[0].months[1:] == (None,) * 11
    assert source_rows[0].cells["month_1"] == "E10"
    assert "month_2" not in source_rows[0].cells
    workbook.close()
