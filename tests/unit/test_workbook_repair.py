from __future__ import annotations

from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
import xlwt
from openpyxl import load_workbook

from excel_transform_1c.adapters.excel import detect_path
from excel_transform_1c.adapters.workbook_repair import (
    WorkbookFormat,
    WorkbookRepairError,
    detect_workbook_format,
    prepare_workbook,
    repair_ooxml,
)
from tests.helpers.workbooks import HEADERS, protected_workbook_bytes, workbook_bytes


def _legacy_biff_bytes() -> bytes:
    output = BytesIO()
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Legacy budget")
    sheet.write(0, 0, "Synthetic fixture")
    for column, value in enumerate(HEADERS):
        sheet.write(1, column, value)
    first = ["ПС", "Административные", "Департамент 1", "ТК", "ЦФО 1", 0.2, "Связь", "Интернет", 100]
    second = ["ПС", "Коммерческие", "Департамент 2", "ТК", "ЦФО 2", "БЕЗ НДС", "Маркетинг", "Реклама", 0]
    for row, values in enumerate((first, second), start=2):
        for column, value in enumerate([*values, *([0] * 11)]):
            sheet.write(row, column, value)
    workbook.save(output)
    return output.getvalue()


def _with_illegal_xml_character(content: bytes) -> bytes:
    source = BytesIO(content)
    target = BytesIO()
    changed = False
    with ZipFile(source) as input_archive, ZipFile(
        target, "w", compression=ZIP_DEFLATED
    ) as output_archive:
        for info in input_archive.infolist():
            payload = input_archive.read(info)
            if info.filename == "xl/worksheets/sheet1.xml":
                marker = b"<sheetData>"
                assert marker in payload
                payload = payload.replace(marker, b"<sheetData>\x01", 1)
                changed = True
            output_archive.writestr(info, payload)
    assert changed
    return target.getvalue()


def _unrecoverable_ooxml() -> bytes:
    output = BytesIO()
    with ZipFile(output, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", b"<Types><")
        archive.writestr("_rels/.rels", b"<Relationships/>")
        archive.writestr("xl/workbook.xml", b"<workbook><")
    return output.getvalue()


def _spreadsheetml_bytes() -> bytes:
    return b"""<?xml version="1.0"?>
<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"
 xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">
 <Worksheet ss:Name="XML budget"><Table>
  <Row><Cell><Data ss:Type="String">Synthetic</Data></Cell></Row>
  <Row><Cell><Data ss:Type="Number">42</Data></Cell></Row>
 </Table></Worksheet>
</Workbook>"""


def _without_central_directory(content: bytes) -> bytes:
    central_directory = content.find(b"PK\x01\x02")
    assert central_directory > 0
    return content[:central_directory]



def _shared_strings_case_mismatch_bytes() -> bytes:
    parts = {
        "[Content_Types].xml": """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>
</Types>""",
        "_rels/.rels": """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>""",
        "xl/workbook.xml": """<?xml version="1.0" encoding="UTF-8"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets><sheet name="TDSheet" sheetId="1" r:id="rId1"/></sheets>
</workbook>""",
        "xl/_rels/workbook.xml.rels": """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>
</Relationships>""",
        "xl/worksheets/sheet1.xml": """<?xml version="1.0" encoding="UTF-8"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetData><row r="1"><c r="A1" t="s"><v>0</v></c></row></sheetData>
</worksheet>""",
        "xl/SharedStrings.xml": """<?xml version="1.0" encoding="UTF-8"?>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="1" uniqueCount="1"><si><t>Значение</t></si></sst>""",
    }
    output = BytesIO()
    with ZipFile(output, "w", compression=ZIP_DEFLATED) as archive:
        for name, payload in parts.items():
            archive.writestr(name, payload)
    return output.getvalue()

def test_format_is_detected_from_signature_not_suffix(tmp_path):
    ooxml = tmp_path / "ordinary.xls"
    ooxml.write_bytes(workbook_bytes())
    assert detect_workbook_format(ooxml) is WorkbookFormat.OOXML

    legacy = tmp_path / "legacy.xlsx"
    legacy.write_bytes(_legacy_biff_bytes())
    assert detect_workbook_format(legacy) is WorkbookFormat.LEGACY_BIFF


def test_encrypted_ooxml_is_distinct_from_legacy_biff(tmp_path):
    legacy = tmp_path / "legacy.xlsx"
    legacy.write_bytes(_legacy_biff_bytes())
    encrypted = tmp_path / "encrypted.xls"
    encrypted.write_bytes(
        protected_workbook_bytes(workbook_bytes(), "synthetic-password")
    )

    assert detect_workbook_format(legacy) is WorkbookFormat.LEGACY_BIFF
    assert detect_workbook_format(encrypted) is WorkbookFormat.ENCRYPTED_OOXML


def test_legacy_spreadsheetml_uses_content_not_xlsx_suffix(tmp_path):
    source = tmp_path / "legacy-xml.xlsx"
    original = _spreadsheetml_bytes()
    source.write_bytes(original)

    prepared = prepare_workbook(source)

    assert prepared.format is WorkbookFormat.XML_SPREADSHEET
    assert prepared.working_path != source
    assert source.read_bytes() == original
    workbook = load_workbook(prepared.working_path, read_only=True, data_only=True)
    try:
        assert workbook["XML budget"]["A2"].value == 42
    finally:
        workbook.close()



def test_shared_strings_case_mismatch_is_repaired_into_openable_copy(tmp_path):
    source = tmp_path / "legacy-case.xlsx"
    original = _shared_strings_case_mismatch_bytes()
    source.write_bytes(original)

    prepared = prepare_workbook(source)

    assert prepared.format is WorkbookFormat.OOXML
    assert prepared.repaired is True
    assert source.read_bytes() == original
    assert prepared.working_path != source
    workbook = load_workbook(prepared.working_path, read_only=True, data_only=True)
    try:
        assert workbook["TDSheet"]["A1"].value == "Значение"
    finally:
        workbook.close()

def test_plain_ooxml_passes_without_change(tmp_path):
    source = tmp_path / "plain.xlsx"
    original = workbook_bytes()
    source.write_bytes(original)

    prepared = prepare_workbook(source)

    assert prepared.format is WorkbookFormat.OOXML
    assert prepared.working_path != source
    assert prepared.repaired is False
    assert source.read_bytes() == original
    assert prepared.working_path.read_bytes() == original


def test_legacy_input_keeps_original_and_uses_separate_working_copy(tmp_path):
    source = tmp_path / "renamed.xlsx"
    original = _legacy_biff_bytes()
    source.write_bytes(original)

    prepared = prepare_workbook(source)

    assert prepared.format is WorkbookFormat.LEGACY_BIFF
    assert prepared.working_path != source
    assert source.read_bytes() == original
    assert prepared.working_path.read_bytes() != original
    workbook = load_workbook(prepared.working_path, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == ["Legacy budget"]
    finally:
        workbook.close()


def test_ooxml_repair_is_idempotent(tmp_path):
    broken = tmp_path / "broken.xlsx"
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    broken.write_bytes(_with_illegal_xml_character(workbook_bytes()))

    first_result = repair_ooxml(broken, first)
    second_result = repair_ooxml(first, second)

    assert first_result.changed is True
    assert second_result.changed is False
    assert first.read_bytes() == second.read_bytes()


def test_recoverable_ooxml_opens_through_excel_adapter(tmp_path):
    source = tmp_path / "recoverable.xlsx"
    original = _with_illegal_xml_character(workbook_bytes())
    source.write_bytes(original)

    candidates = detect_path(source)
    prepared = prepare_workbook(source)

    assert len(candidates) == 1
    assert prepared.repaired is True
    assert prepared.working_path != source
    assert source.read_bytes() == original


def test_ooxml_with_missing_central_directory_is_rebuilt(tmp_path):
    source = tmp_path / "truncated-directory.xlsx"
    original = _without_central_directory(workbook_bytes())
    source.write_bytes(original)

    prepared = prepare_workbook(source)

    assert prepared.repaired is True
    assert prepared.working_path != source
    assert source.read_bytes() == original
    assert len(detect_path(prepared.working_path)) == 1


def test_unrecoverable_file_fails_closed(tmp_path):
    source = tmp_path / "unrecoverable.xlsx"
    source.write_bytes(_unrecoverable_ooxml())

    with pytest.raises(WorkbookRepairError, match="не может быть восстановлена"):
        prepare_workbook(source)
