from io import BytesIO

from openpyxl import load_workbook

from excel_transform_1c.adapters.references import (
    organization_nodes,
    parse_reference_workbook,
)
from excel_transform_1c.application.service import WorkflowService
from excel_transform_1c.core.detection import (
    BDR_REVENUE_ARTICLES,
    BDR_REVENUE_GROUP,
    detect_candidate_ranges,
    read_source_rows,
)
from excel_transform_1c.core.models import IndicatorType
from excel_transform_1c.core.organization_hierarchy import (
    ExactOrganizationHierarchyResolver,
)
from tests.helpers.workbooks import (
    BDR_FULL_INDICATORS,
    bdr_full_workbook_bytes,
    erp_organization_hierarchy_bytes,
)


def _candidate_and_rows():
    workbook = load_workbook(BytesIO(bdr_full_workbook_bytes()), data_only=True)
    candidates = detect_candidate_ranges(workbook)
    assert len(candidates) == 1
    candidate = candidates[0]
    rows = read_source_rows(workbook, candidate, "БДР 2026 ИТОГ.xlsx")
    return workbook, candidate, rows


def test_bdr_detect_income_block() -> None:
    workbook, candidate, rows = _candidate_and_rows()
    try:
        assert candidate.source_kind == "bdr_full"
        assert candidate.label == "БДР 2026 ИТОГ"
        assert {
            row.article
            for row in rows
            if row.indicator_type == IndicatorType.REVENUE.value
        } == {
            "Выручка ИТОГО",
            "Прочие доходы по основной деятельности",
            "Валовая прибыль",
        }
    finally:
        workbook.close()


def test_bdr_sales_channel_is_analytics_not_indicator() -> None:
    workbook = load_workbook(BytesIO(bdr_full_workbook_bytes()), data_only=True)
    sheet = workbook.active
    sheet.cell(23, 1, BDR_REVENUE_GROUP)
    sheet.cell(23, 7, "Опт")
    sheet.cell(14, 1, "Оборот в кг")
    sheet.cell(14, 7, "Розница")
    for month in range(1, 13):
        sheet.cell(23, 22 + month, month)
        sheet.cell(14, 22 + month, month)
    candidate = detect_candidate_ranges(workbook)[0]
    rows = read_source_rows(workbook, candidate, "БДР 2026 ИТОГ.xlsx")
    channel = next(row for row in rows if row.input_sales_channel == "Опт")
    assert channel.article == BDR_REVENUE_GROUP
    assert channel.indicator_type == IndicatorType.REVENUE.value
    assert channel.cells["article"] == "A23"
    assert channel.cells["input_sales_channel"] == "G23"
    kpi_channel = next(row for row in rows if row.input_sales_channel == "Розница")
    assert kpi_channel.article == "Оборот в кг"
    assert kpi_channel.indicator_type == IndicatorType.KPI.value
    assert kpi_channel.cells["article"] == "A14"
    workbook.close()


def test_bdr_seven_channel_block_uses_preceding_parent_indicator() -> None:
    workbook = load_workbook(BytesIO(bdr_full_workbook_bytes()), data_only=True)
    sheet = workbook.active
    sheet.cell(22, 7, "Валовая прибыль")
    channels = BDR_REVENUE_ARTICLES
    for row_number, channel in enumerate(channels, start=23):
        sheet.cell(row_number, 1, None)
        sheet.cell(row_number, 7, channel)
        for month in range(1, 13):
            sheet.cell(row_number, 22 + month, month)

    candidate = detect_candidate_ranges(workbook)[0]
    rows = read_source_rows(workbook, candidate, "БДР 2026 ИТОГ.xlsx")
    channel_rows = [
        row for row in rows
        if row.input_sales_channel in channels and row.row_number >= 23
    ]

    assert len(channel_rows) == 7
    assert {row.article for row in channel_rows} == {"Валовая прибыль"}
    assert {row.indicator_type for row in channel_rows} == {IndicatorType.REVENUE.value}
    assert {row.cells["article"] for row in channel_rows} == {"G22"}
    workbook.close()


def test_bdr_detect_expense_block() -> None:
    workbook, _, rows = _candidate_and_rows()
    try:
        expenses = [
            row
            for row in rows
            if row.indicator_type == IndicatorType.EXPENSE.value
        ]
        assert len(expenses) == 1
        assert expenses[0].expense_type == "Административные расходы"
        assert expenses[0].department == "Административный департамент"
        assert expenses[0].cfo == "АЮ Административный Отдел"
        assert expenses[0].expense_group == "Связь"
        assert expenses[0].article == "Интернет"
        assert expenses[0].months[0] == 1
        assert expenses[0].cells["month_1"] == "I2"
    finally:
        workbook.close()


def test_bdr_component_month_uses_excel_display_precision(tmp_path) -> None:
    workbook = load_workbook(BytesIO(bdr_full_workbook_bytes()), data_only=False)
    component = workbook["загрузка ERP расходы"]
    component["I2"] = 123.6
    component["I2"].number_format = "#,##0"
    path = tmp_path / "component-precision.xlsx"
    workbook.save(path)
    workbook.close()

    from excel_transform_1c.adapters.excel import detect_path, read_path

    candidate = detect_path(path)[0]
    expense = next(
        row for row in read_path(path, candidate, path.name)
        if row.source_kind == "prepared_budget"
    )
    assert expense.months[0] == 124


def test_bdr_component_confirmation_uses_run_row_identity(tmp_path) -> None:
    workbook = load_workbook(BytesIO(bdr_full_workbook_bytes()))
    workbook["загрузка ERP расходы"]["F2"] = None
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()

    service = WorkflowService(tmp_path / "runtime")
    service.upload_reference("organizations", erp_organization_hierarchy_bytes())
    scenario = next(
        item for item in service.store.list_scenarios() if item.name == "ПЛАН 2026"
    )
    reporting_unit = "АЮ Административный Отдел"
    context = service.build_context(
        reporting_unit,
        "000000001",
        scenario.scenario_id,
        2026,
        [],
    )
    pending = service.prepare_upload("БДР 2026 ИТОГ.xlsx", payload.getvalue(), context)
    run = service.process_upload(pending.upload_id, pending.candidates[0].candidate_id)
    expense = next(record for record in run.records if record.source_article == "Интернет")
    tax_issue = next(
        issue
        for issue in run.unresolved_issues
        if issue.kind == "tax" and issue.pointer.row == expense.source_row
    )

    assert expense.source_row != 2
    assert expense.source_excel_row == 2
    assert tax_issue.pointer.excel_row == 2
    assert service.tax_not_required_source_rows(run.run_id) == [expense.source_row]
    _, count = service.confirm_tax_not_required(run.run_id, [expense.source_row])
    assert count == 1


def test_bdr_detect_kpi_block() -> None:
    workbook, _, rows = _candidate_and_rows()
    try:
        kpis = {
            row.article
            for row in rows
            if row.indicator_type == IndicatorType.KPI.value
        }
        assert {
            "Оборот в кг",
            "Выручка за 1 кг",
            "Итого расходов на 1 кг",
            "Валовая прибыль на 1 кг",
            "EBITDA",
            "Операционная прибыль",
        }.issubset(kpis)
    finally:
        workbook.close()


def test_bdr_full_hierarchy_resolution() -> None:
    nodes = organization_nodes(
        parse_reference_workbook(
            erp_organization_hierarchy_bytes(),
            "organizations",
        )
    )

    resolution = ExactOrganizationHierarchyResolver(nodes).resolve_exact_department(
        "000000001",
        "АЮ Административный Отдел",
    )

    assert resolution is not None
    assert resolution.organization_unit == 'ООО "Айс Юнион"'
    assert resolution.organization_unit_code == "000000001"
    assert resolution.department == "Административный департамент"
    assert resolution.cfo == "АЮ Административный Отдел"
    assert resolution.cfo_code == "000000173"


def test_bdr_export_not_empty(tmp_path) -> None:
    service = WorkflowService(tmp_path / "runtime")
    service.upload_reference("organizations", erp_organization_hierarchy_bytes())
    scenario = next(
        item for item in service.store.list_scenarios() if item.name == "ПЛАН 2026"
    )
    reporting_unit = "АЮ Административный Отдел"
    context = service.build_context(
        reporting_unit,
        "000000001",
        scenario.scenario_id,
        2026,
        [],
    )
    pending = service.prepare_upload(
        "БДР 2026 ИТОГ.xlsx",
        bdr_full_workbook_bytes(reporting_unit=reporting_unit),
        context,
    )
    assert [candidate.label for candidate in pending.candidates] == ["БДР 2026 ИТОГ"]
    run = service.process_upload(
        pending.upload_id,
        pending.candidates[0].candidate_id,
    )

    diagnostics = service.bdr_diagnostics(run.run_id)
    assert diagnostics is not None
    expected_rows = len(BDR_FULL_INDICATORS) - 5 + 1
    assert diagnostics["rows_read"] == expected_rows
    assert diagnostics["monthly_cells_read"] == expected_rows * 12
    assert diagnostics["numeric_values"] == expected_rows * 12
    assert diagnostics["excel_errors"] == 0
    assert diagnostics["income"] == 3
    assert diagnostics["expense"] == 1
    assert diagnostics["kpi"] == 6
    assert diagnostics["kpi_found"] == 6
    assert diagnostics["kpi_with_organization"] == 6
    assert diagnostics["kpi_with_period"] == 6
    assert diagnostics["kpi_with_value"] == 6
    assert diagnostics["kpi_exported"] == 6
    assert diagnostics["matched"] == expected_rows
    assert diagnostics["exported"] == expected_rows
    assert diagnostics["exclusions"] == []
    assert len(service.export_run(run.run_id)) > 0


def test_bdr_non_exported_indicator_has_reason(tmp_path) -> None:
    service = WorkflowService(tmp_path / "runtime")
    service.upload_reference("organizations", erp_organization_hierarchy_bytes())
    scenario = next(
        item for item in service.store.list_scenarios() if item.name == "ПЛАН 2026"
    )
    reporting_unit = "АЮ Административный Отдел"
    context = service.build_context(
        reporting_unit,
        "000000001",
        scenario.scenario_id,
        2026,
        [],
    )
    pending = service.prepare_upload(
        "БДР 2026 ИТОГ.xlsx",
        bdr_full_workbook_bytes(
            reporting_unit=reporting_unit,
            error_indicator="Операционная прибыль",
        ),
        context,
    )
    run = service.process_upload(
        pending.upload_id,
        pending.candidates[0].candidate_id,
    )

    diagnostics = service.bdr_diagnostics(run.run_id)
    assert diagnostics is not None
    expected_rows = len(BDR_FULL_INDICATORS) - 5 + 1
    assert diagnostics["monthly_cells_read"] == expected_rows * 12
    assert diagnostics["numeric_values"] == (expected_rows - 1) * 12
    assert diagnostics["excel_errors"] == 12
    assert diagnostics["attention_rows"] >= 1
    assert diagnostics["exported"] == expected_rows - 1
    assert diagnostics["exclusions"] == [
        {
            "source_row": 41,
            "source_sheet": "Произвольный сводный лист",
            "indicator": "Операционная прибыль",
            "reason": "Ошибка Excel в месячной ячейке (Произвольный сводный лист!W41); "
            "Ошибка Excel в месячной ячейке (Произвольный сводный лист!X41); "
            "Ошибка Excel в месячной ячейке (Произвольный сводный лист!Y41); "
            "Ошибка Excel в месячной ячейке (Произвольный сводный лист!Z41); "
            "Ошибка Excel в месячной ячейке (Произвольный сводный лист!AA41); "
            "Ошибка Excel в месячной ячейке (Произвольный сводный лист!AB41); "
            "Ошибка Excel в месячной ячейке (Произвольный сводный лист!AC41); "
            "Ошибка Excel в месячной ячейке (Произвольный сводный лист!AD41); "
            "Ошибка Excel в месячной ячейке (Произвольный сводный лист!AE41); "
            "Ошибка Excel в месячной ячейке (Произвольный сводный лист!AF41); "
            "Ошибка Excel в месячной ячейке (Произвольный сводный лист!AG41); "
            "Ошибка Excel в месячной ячейке (Произвольный сводный лист!AH41)",
        }
    ]
